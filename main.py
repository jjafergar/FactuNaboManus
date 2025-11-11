# main.py (dashboard, histórico, config, validaciones, stepper, selector post-envío Macro + login con overlay y estilo unificado + borrar historial + fix login style/center v2 + fix splitter handle v3 + layout Enviar compacto + Font changes + Modern Login UI v2 Frameless + FIX QFont + Logo Tematizado)
# [CORREGIDO] Ahora usa macro_adapter.adapt_from_macro(path) con 1 solo argumento (lee hoja "CLIENTES" interna)
# [MODIFICADO] Dashboard mejorado con consulta dinámica por EMISOR y PERIODO.
import sys
import os
import json
import sqlite3
import webbrowser
import hashlib
import re
import glob
import pandas as pd
from datetime import datetime, timedelta, date
import platform
import ctypes

from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QLabel, QPushButton, QVBoxLayout,
    QHBoxLayout, QFileDialog, QTableWidget, QTableWidgetItem, QTextEdit,
    QProgressBar, QFrame, QToolButton, QLineEdit, QCheckBox, QSplitter,
    QListWidget, QStackedWidget, QDialog, QFormLayout, QMessageBox,
    QListWidgetItem, QHeaderView, QInputDialog, QComboBox, QGraphicsOpacityEffect,
    QGraphicsDropShadowEffect, QStyle, QColorDialog, QSizePolicy # <-- Iconos Sidebar (Aunque ahora no se usa para emojis)
)
from PySide6.QtGui import (
    QPixmap, QFont, QColor, QIcon, QKeySequence, QShortcut, QPalette, QScreen # Necesario para QScreen, QDesktopServices
)

from PySide6.QtGui import QDesktopServices
from PySide6.QtCore import (
QEasingCurve, QPoint, QPropertyAnimation, QRect, QSettings, QSize, QThread, QTimer, QUrl, Qt, QLocale, QTranslator
)

from worker import Worker, detect_available_browser
import macro_adapter

# --- Resource path helper for PyInstaller ---
def resource_path(relative_path):
    """Devuelve la ruta absoluta del recurso tanto en desarrollo como en los binarios."""
    if getattr(sys, "frozen", False):
        # Cuando está empaquetado preferimos la carpeta junto al ejecutable
        exec_dir = os.path.dirname(sys.executable)
        candidate = os.path.join(exec_dir, relative_path)
        if os.path.exists(candidate) or not hasattr(sys, "_MEIPASS"):
            return candidate
        # Recurso solo embebido (estático)
        return os.path.join(sys._MEIPASS, relative_path)

    # Ejecución en entorno de desarrollo
    base_path = os.path.abspath(os.path.dirname(__file__))
    return os.path.join(base_path, relative_path)


# --- Helpers ---
def _normalize_invoice_id(x):
    s = str(x).strip()
    # Si es numérico (incluye "25042.0"), devolvemos entero sin .0
    if re.fullmatch(r"\d+(?:\.0+)?", s):
        try:
            return str(int(float(s)))
        except Exception:
            return s
    # Si es alfanumérico (p.ej. "Int_25003"), devolvemos tal cual
    return s


# Formato monetario (es-ES): 3.976,42€
def format_eur(value) -> str:
    try:
        v = float(value)
    except Exception:
        return ""
    s = f"{v:,.2f}"
    s = s.replace(",", "X").replace(".", ",").replace("X", ".")
    return f"{s}€"

# [NUEVO] Helper para aplicar sombras
def apply_shadow(widget, blur=20, offset_y=4, color_str="#000000"):
    """Aplica un efecto de sombra sutil y moderno."""
    shadow = QGraphicsDropShadowEffect(widget)
    shadow.setBlurRadius(blur)
    # Sombra muy suave (alfa bajo)
    color = QColor(color_str)
    color.setAlpha(40)
    shadow.setColor(color)
    shadow.setOffset(0, offset_y)
    widget.setGraphicsEffect(shadow)


# [NUEVO] Windows Acrylic/Mica (blur/transparencia real)
def enable_windows_backdrop(win_id, dark_mode: bool = False):
    """
    Intenta activar Mica/Acrylic/Immersive Dark Mode en Windows 10/11.
    Silencioso si falla o no aplica.
    """
    try:
        if platform.system().lower() != "windows":
            return
        hwnd = int(win_id)
        DWMWA_USE_IMMERSIVE_DARK_MODE = 20  # BOOL
        DWMWA_MICA_EFFECT = 1029            # BOOL (Windows 11 antiguas)
        DWMWA_SYSTEMBACKDROP_TYPE = 38      # DWORD (Windows 11 22H2+)
        # 0=Auto, 1=None, 2=Mica, 3=Acrylic (algunas builds), 4=Tabbed
        backdrop_mica = ctypes.c_int(2)
        # backdrop_acrylic = ctypes.c_int(3)
        dark = ctypes.c_int(1 if dark_mode else 0)
        dwmapi = ctypes.windll.dwmapi
        # Dark chrome
        dwmapi.DwmSetWindowAttribute(hwnd, DWMWA_USE_IMMERSIVE_DARK_MODE, ctypes.byref(dark), ctypes.sizeof(dark))
        # System backdrop (prefer Mica)
        dwmapi.DwmSetWindowAttribute(hwnd, DWMWA_SYSTEMBACKDROP_TYPE, ctypes.byref(backdrop_mica), ctypes.sizeof(backdrop_mica))
        # Fallback for older Windows 11
        mica_on = ctypes.c_int(1)
        dwmapi.DwmSetWindowAttribute(hwnd, DWMWA_MICA_EFFECT, ctypes.byref(mica_on), ctypes.sizeof(mica_on))
        # If you prefer Acrylic, uncomment:
        # dwmapi.DwmSetWindowAttribute(hwnd, DWMWA_SYSTEMBACKDROP_TYPE, ctypes.byref(backdrop_acrylic), ctypes.sizeof(backdrop_acrylic))
    except Exception:
        pass

# --- Colores/Recursos ---
COLOR_PRIMARY = "#A0BF6E"
COLOR_SUCCESS = "#34C759"
COLOR_WARNING = "#FF9500"
COLOR_ERROR = "#FF3B30"
COLOR_BACKGROUND = "#F2F2F7"
COLOR_CARD = "#FFFFFF"
COLOR_TEXT = "#000000"
COLOR_SECONDARY_TEXT = "#8E8E93"
COLOR_BORDER = "#C6C6C8"
COLOR_SIDEBAR = "#FAFAFA"
COLOR_SIDEBAR_DARK = "#1C1C1E"
COLOR_DARK_BG = "#000000"
COLOR_DARK_CARD = "#1C1C1E"
COLOR_DARK_TEXT = "#FFFFFF"
COLOR_DARK_BORDER = "#38383A"

RESOURCE_DIR = resource_path("resources")
DB_PATH = resource_path("factunabo_history.db")
USERS_PATH = resource_path("users.json")


# --- DB ---
# [MODIFICADO] init_database para añadir la columna 'importe'
def init_database():
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    cursor.execute(
        """
        CREATE TABLE IF NOT EXISTS envios (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            fecha_envio TEXT NOT NULL,
            num_factura TEXT,
            empresa TEXT,
            estado TEXT,
            detalles TEXT,
            pdf_url TEXT,
            excel_path TEXT,
            pdf_local_path TEXT,
            importe REAL DEFAULT 0.0,
            cliente TEXT
        )
        """
    )
    # --- [NUEVO] Añadir columna 'importe' si la tabla ya existe ---
    try:
        cursor.execute("ALTER TABLE envios ADD COLUMN importe REAL DEFAULT 0.0")
    except sqlite3.OperationalError:
        pass
    # --- [FIN NUEVO] ---
    # --- [NUEVO] Añadir columna 'cliente' si la tabla ya existe ---
    try:
        cursor.execute("ALTER TABLE envios ADD COLUMN cliente TEXT")
    except sqlite3.OperationalError:
        pass
    # --- [FIN NUEVO] ---
    try:
        cursor.execute("ALTER TABLE envios ADD COLUMN pdf_local_path TEXT")
    except sqlite3.OperationalError:
        pass
    
    # --- [NUEVO] Crear índices para optimizar consultas ---
    try:
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_fecha_envio ON envios(fecha_envio)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_empresa ON envios(empresa)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_estado ON envios(estado)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_num_factura ON envios(num_factura)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_cliente ON envios(cliente)")
    except sqlite3.OperationalError as e:
        print(f"Warning: Error creando índices: {e}")
    # --- [FIN NUEVO] ---
    
    # --- [NUEVO] Crear tabla para cola offline ---
    try:
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS offline_queue (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                xml_content BLOB NOT NULL,
                num_factura TEXT NOT NULL,
                empresa TEXT NOT NULL,
                ejercicio TEXT,
                cliente_doc TEXT,
                api_key TEXT,
                fecha_creacion TEXT NOT NULL,
                intentos INTEGER DEFAULT 0,
                ultimo_intento TEXT,
                estado TEXT DEFAULT 'PENDIENTE'
            )
        """)
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_queue_estado ON offline_queue(estado)")
        cursor.execute("CREATE INDEX IF NOT EXISTS idx_queue_fecha ON offline_queue(fecha_creacion)")
    except sqlite3.OperationalError as e:
        print(f"Warning: Error creando tabla offline_queue: {e}")
    # --- [FIN NUEVO] ---
    
    conn.commit()
    conn.close()


# --- Componentes UI ---

# [TOTALMENTE CORREGIDO] AnimatedButton refactorizada para evitar conflictos de eventos
class AnimatedButton(QPushButton):
    """Botón con animación de "elevación" (sombra) al estilo macOS."""

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.setProperty("class", "AnimatedButton")
        # Asegurar que el texto sea visible
        self.setStyleSheet("color: white !important;")

        # Sombra base
        self._shadow = QGraphicsDropShadowEffect(self)
        self._shadow.setBlurRadius(18)
        color = QColor(0, 0, 0, 60) # Usamos un color de sombra fijo y seguro
        self._shadow.setColor(color)
        self._shadow.setOffset(0, 4)
        self.setGraphicsEffect(self._shadow)

        # Animación para el radio de la sombra (blur)
        self._anim_blur = QPropertyAnimation(self._shadow, b"blurRadius")
        self._anim_blur.setDuration(180)
        self._anim_blur.setEasingCurve(QEasingCurve.OutCubic)

        # Animación para el offset (Y)
        self._anim_offset = QPropertyAnimation(self._shadow, b"yOffset")
        self._anim_offset.setDuration(180)
        self._anim_offset.setEasingCurve(QEasingCurve.OutCubic)

    # --- [NUEVAS FUNCIONES HELPER] ---
    def _animate_hover_in(self):
        """Animates the button shadow to the 'hover' state."""
        self._anim_blur.stop()
        self._anim_blur.setStartValue(self._shadow.blurRadius())
        self._anim_blur.setEndValue(30)
        self._anim_blur.start()

        self._anim_offset.stop()
        self._anim_offset.setStartValue(self._shadow.yOffset())
        self._anim_offset.setEndValue(6) # Sube un poco
        self._anim_offset.start()

    def _animate_hover_out(self):
        """Animates the button shadow to the 'normal' state."""
        self._anim_blur.stop()
        self._anim_blur.setStartValue(self._shadow.blurRadius())
        self._anim_blur.setEndValue(18)
        self._anim_blur.start()

        self._anim_offset.stop()
        self._anim_offset.setStartValue(self._shadow.yOffset())
        self._anim_offset.setEndValue(4)
        self._anim_offset.start()
    # --- [FIN DE FUNCIONES HELPER] ---

    def enterEvent(self, e):
        # Al pasar el ratón, la sombra se expande y "sube"
        self._animate_hover_in()
        super().enterEvent(e)

    def leaveEvent(self, e):
        # Vuelve al estado normal
        self._animate_hover_out()
        super().leaveEvent(e)

    def mousePressEvent(self, e):
        # Al presionar, la sombra se contrae (efecto "click")
        self._anim_blur.stop()
        self._anim_blur.setStartValue(self._shadow.blurRadius())
        self._anim_blur.setEndValue(10)
        self._anim_blur.start()

        self._anim_offset.stop()
        self._anim_offset.setStartValue(self._shadow.yOffset())
        self._anim_offset.setEndValue(2)
        self._anim_offset.start()

        super().mousePressEvent(e)

    def mouseReleaseEvent(self, e):
        # Al soltar, vuelve al estado hover/normal
        super().mouseReleaseEvent(e) # Llama al padre primero

        # [CORRECCIÓN 1: DeprecationWarning]
        # e.pos() está obsoleto, usamos e.position().toPoint()
        current_pos = e.position().toPoint()

        # [CORRECCIÓN 2: TypeError]
        # No llames a enterEvent/leaveEvent. Llama a las funciones de animación.
        if self.rect().contains(current_pos):
            self._animate_hover_in() # El ratón sigue encima
        else:
            self._animate_hover_out() # El ratón se soltó fuera


class StatusChip(QLabel):
    def __init__(self, status, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.setProperty("class", "StatusChip")
        up = (status or "").upper()
        if up in ["ÉXITO", "SUCCESS"]:
            self.setProperty("status", "success")
        elif up in ["DUPLICADO", "DUPLICATE", "ATENCION"]:
            self.setProperty("status", "warning")
        else:
            self.setProperty("status", "NABO!") # O "error" si prefieres
        self.setText(up)
        self.setAlignment(Qt.AlignCenter)
        self.setFixedHeight(26)


class AnimatedNavList(QListWidget):
    """ListWidget con animación de hover que agranda el item (manejado por QSS)."""
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.setProperty("class", "NavList")
        # El efecto de hover se maneja completamente por QSS
        # No necesitamos lógica adicional aquí


class ModernTable(QTableWidget):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.setProperty("class", "ModernTable")
        self.setAlternatingRowColors(True)
        self.setMouseTracking(True)
        self.verticalHeader().setVisible(False)
        self.horizontalHeader().setStretchLastSection(True)

class TableTools(QWidget):
    """Barra de herramientas para tablas: búsqueda y densidad."""
    def __init__(self, table: QTableWidget, parent=None):
        super().__init__(parent)
        self._table = table
        row = QHBoxLayout(self)
        row.setContentsMargins(0, 0, 0, 8)
        row.setSpacing(8)
        self.search = QLineEdit(self)
        self.search.setPlaceholderText("Buscar…")
        self.search.textChanged.connect(self._apply_filter)
        self.compact_toggle = QCheckBox("Vista compacta", self)
        self.compact_toggle.setToolTip("Reduce el espaciado en tablas y controles para mostrar más información en menos espacio")
        self.compact_toggle.toggled.connect(self._toggle_density)
        row.addWidget(self.search)
        row.addStretch()
        row.addWidget(self.compact_toggle)

    def _toggle_density(self, checked: bool):
        """Activa/desactiva el modo vista compacta que reduce padding en tablas y controles."""
        win = self.window()
        if isinstance(win, QMainWindow):
            win.setProperty("density", "compact" if checked else "")
            # Refrescar estilos en toda la aplicación
            app = QApplication.instance()
            if app:
                for widget in app.allWidgets():
                    try:
                        widget.style().unpolish(widget)
                        widget.style().polish(widget)
                    except (RuntimeError, AttributeError):
                        pass

    def _apply_filter(self, text: str):
        txt = (text or "").strip().lower()
        for r in range(self._table.rowCount()):
            visible = False
            for c in range(self._table.columnCount()):
                it = self._table.item(r, c)
                if it and txt in str(it.text()).lower():
                    visible = True
                    break
            self._table.setRowHidden(r, not visible)

# --- Stepper ---
# [MODIFICADO] StepperWidget ahora anima la línea de progreso
class StepperWidget(QWidget):
    def __init__(self, steps, parent=None):
        super().__init__(parent)
        self.steps = steps
        self.current_step = 0
        self.init_ui()

    def init_ui(self):
        layout = QHBoxLayout(self)
        layout.setContentsMargins(20, 10, 20, 10)
        layout.setSpacing(0)
        self.step_labels = []
        self.step_circles = []
        self.step_lines = []
        self.line_anims = [] # [NUEVO]

        for i, step_name in enumerate(self.steps):
            circle = QLabel()
            circle.setFixedSize(40, 40)
            circle.setAlignment(Qt.AlignCenter)
            circle.setProperty("class", "StepCircle")
            circle.setProperty("state", "pending")
            circle.setText(str(i + 1))
            self.step_circles.append(circle)

            label = QLabel(step_name)
            label.setAlignment(Qt.AlignCenter)
            label.setProperty("class", "StepLabel")
            self.step_labels.append(label)

            step_container = QVBoxLayout()
            step_container.addWidget(circle, alignment=Qt.AlignCenter)
            step_container.addWidget(label)
            layout.addLayout(step_container)

            if i < len(self.steps) - 1:
                line = QFrame()
                line.setFrameShape(QFrame.HLine)
                line.setProperty("class", "StepLine")
                line.setProperty("state", "pending")
                line.setFixedHeight(2)
                line.setMinimumWidth(60)

                # [NUEVO] Preparar animación
                line.setMaximumWidth(0) # Empieza oculto
                anim = QPropertyAnimation(line, b"maximumWidth")
                anim.setDuration(400)
                anim.setEasingCurve(QEasingCurve.OutCubic)
                self.line_anims.append(anim)

                self.step_lines.append(line)
                layout.addWidget(line, alignment=Qt.AlignVCenter)

        self.set_step(0) # Inicializa el estado visual

    def set_step(self, step_index):
        self.current_step = step_index
        for i, circle in enumerate(self.step_circles):
            if i < step_index:
                circle.setProperty("state", "completed")
                circle.setText("✓")
            elif i == step_index:
                circle.setProperty("state", "active")
                circle.setText(str(i + 1))
            else:
                circle.setProperty("state", "pending")
                circle.setText(str(i + 1))
            circle.style().unpolish(circle)
            circle.style().polish(circle)

        for i, line in enumerate(self.step_lines):
            is_completed = (i < step_index)
            line.setProperty("state", "completed" if is_completed else "pending")
            line.style().unpolish(line)
            line.style().polish(line)

            # [NUEVO] Animar la línea
            anim = self.line_anims[i]
            anim.stop() # Detiene animación anterior si la hay

            target_width = line.minimumWidth() if is_completed else 0
            current_width = line.width()

            if current_width != target_width:
                anim.setStartValue(current_width)
                anim.setEndValue(target_width)
                anim.start()


# --- Login con overlay (estilo integrado) ---
class Overlay(QWidget):
    """Capa que oscurece la ventana y bloquea clicks mientras el login está visible."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setAttribute(Qt.WA_NoSystemBackground, False)
        self.setAttribute(Qt.WA_TransparentForMouseEvents, False)  # Captura eventos
        self.setStyleSheet("background-color: rgba(0,0,0,0.35);")
        self.hide()

    def resizeEvent(self, e):
        if self.parent():
            self.setGeometry(self.parent().rect())
        super().resizeEvent(e)


class LoginDialog(QDialog):
    """Login con el mismo look (usa styles.qss) y sin cerrar la app en fallo."""

    def __init__(self, parent=None):
        super().__init__(parent)

        # --- [MODIFICACIÓN FRAMELESS] ---
        # Ocultar la barra de título y activar fondo transparente
        self.setWindowFlags(Qt.FramelessWindowHint | Qt.Dialog)
        self.setAttribute(Qt.WA_TranslucentBackground)
        # --- [FIN MODIFICACIÓN] ---

        self.setWindowTitle("Acceso a FactuNabo") # Aunque no se vea, es bueno mantenerlo
        self.setModal(True)
        self.setMinimumWidth(420) # Ancho adecuado
        self.setMinimumHeight(480) # Altura mínima para evitar corte visual
        self.settings = QSettings("FactuNabo", "Login")
        self._ensure_users_file()
        self._init_ui()

    def _ensure_users_file(self):
        if not os.path.exists(USERS_PATH):
            default = {
                "users": [
                    {
                        "username": "admin",
                        "password_hash": hashlib.sha256("admin".encode("utf-8")).hexdigest(),
                    }
                ]
            }
            with open(USERS_PATH, "w", encoding="utf-8") as f:
                json.dump(default, f, indent=2, ensure_ascii=False)

    def _load_users(self):
        try:
            with open(USERS_PATH, "r", encoding="utf-8") as f:
                return json.load(f).get("users", [])
        except Exception:
            return []

    # [MODIFICADO] _init_ui para un look más moderno
    def _init_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)

        card = QFrame()
        card.setObjectName("LoginCard") # <-- ID para el QSS
        apply_shadow(card, blur=30, offset_y=5) # Sombra

        card_layout = QVBoxLayout(card)
        card_layout.setContentsMargins(32, 32, 32, 36) # Más padding inferior para evitar corte
        card_layout.setSpacing(18) # Más espaciado entre elementos

        title = QLabel("FactuNabo – Inicio de sesión")
        title.setAlignment(Qt.AlignCenter)
        title.setStyleSheet("font-size: 20px; font-weight: 700;")
        card_layout.addWidget(title)

        subtitle = QLabel("Introduce tus credenciales para continuar")
        subtitle.setAlignment(Qt.AlignCenter)
        subtitle.setStyleSheet("color:#8E8E93;")
        card_layout.addWidget(subtitle)

        card_layout.addSpacing(10) # Espacio antes de los inputs

        # --- [MODIFICACIÓN] Quitar QFormLayout por un QVBoxLayout ---
        input_layout = QVBoxLayout()
        input_layout.setSpacing(10) # Espacio entre inputs

        self.username = QLineEdit()
        self.username.setPlaceholderText("Usuario")
        self.username.returnPressed.connect(lambda: (self.password.setFocus(), self.password.selectAll()))
        input_layout.addWidget(self.username)

        self.password = QLineEdit()
        self.password.setPlaceholderText("Contraseña")
        self.password.setEchoMode(QLineEdit.Password)
        self.password.returnPressed.connect(self.do_login)
        input_layout.addWidget(self.password)

        card_layout.addLayout(input_layout)
        # --- [FIN MODIFICACIÓN] ---

        self.remember = QCheckBox("Recordarme")
        card_layout.addWidget(self.remember)

        self.error_label = QLabel("")
        self.error_label.setStyleSheet("color:#FF3B30; font-weight:600;")
        self.error_label.setWordWrap(True)
        self.error_label.setVisible(False)
        card_layout.addWidget(self.error_label)

        last_user = self.settings.value("last_user", "")
        if last_user:
            self.username.setText(last_user)
            self.remember.setChecked(True)
            QTimer.singleShot(0, self.password.setFocus)

        card_layout.addSpacing(10) # Espacio antes de los botones

        # --- [MODIFICACIÓN] Botones apilados, "Entrar" al 100% ---
        btns_layout = QVBoxLayout()
        btns_layout.setSpacing(8)

        btn_login = AnimatedButton("Entrar")
        btn_login.setStyleSheet("padding: 8px 20px; min-height: 32px; font-size: 13px;")
        btn_login.clicked.connect(self.do_login)
        btns_layout.addWidget(btn_login)

        btn_cancel = QPushButton("Cancelar")
        btn_cancel.clicked.connect(self.reject) # self.reject() cierra el diálogo
        # Hacemos que el botón "Cancelar" parezca un enlace
        btn_cancel.setProperty("class", "LinkButton")
        btns_layout.addWidget(btn_cancel, alignment=Qt.AlignCenter)

        card_layout.addLayout(btns_layout)
        # --- [FIN MODIFICACIÓN] ---

        root.addStretch()
        root.addWidget(card, alignment=Qt.AlignCenter)
        root.addStretch()


    def do_login(self):
        u = (self.username.text() or "").strip()
        p = self.password.text() or ""
        if not u or not p:
            self._show_error("Indica usuario y contraseña.")
            return
        users = self._load_users()
        ok = any(
            (u == item.get("username") and hashlib.sha256(p.encode("utf-8")).hexdigest() == item.get("password_hash"))
            for item in users
        )
        if not ok:
            self._show_error("Usuario o contraseña incorrectos.")
            self.password.clear()
            self.password.setFocus()
            return
        if self.remember.isChecked():
            self.settings.setValue("last_user", u)
        else:
            self.settings.remove("last_user")
        os.environ["FACTUNABO_USER"] = u
        self.accept()

    def _show_error(self, msg: str):
        self.error_label.setText(msg)
        self.error_label.setVisible(True)


# --- Config API Dialog ---
# [MODIFICADO] ConfigDialog ahora es Frameless y centrado
class ConfigDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        
        # --- 1. Añadir Banderas Frameless ---
        self.setWindowFlags(Qt.FramelessWindowHint | Qt.Dialog)
        self.setAttribute(Qt.WA_TranslucentBackground)
        
        self.setMinimumWidth(750) # Ancho aumentado para que se vea completa la URL
        self.settings = QSettings("FactuNabo", "APIConfig")
        self.init_ui()
        self.load_settings()

    # --- 2. Reestructurar UI para que sea una "tarjeta" ---
    def init_ui(self):
        # Layout raíz transparente
        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)

        # Tarjeta contenedora (usamos "LoginCard" para reusar el estilo QSS)
        card = QFrame()
        card.setObjectName("LoginCard") 
        apply_shadow(card, blur=25, offset_y=4) 

        card_layout = QVBoxLayout(card)
        card_layout.setContentsMargins(32, 32, 32, 36) # Más padding inferior para evitar corte
        card_layout.setSpacing(18) # Más espaciado entre elementos

        # Título
        title = QLabel("Configuración de API")
        title.setAlignment(Qt.AlignCenter)
        title.setStyleSheet("font-size: 20px; font-weight: 700;")
        card_layout.addWidget(title)
        card_layout.addSpacing(10)

        # Formulario
        form = QFormLayout()
        form.setSpacing(12) # Espaciado entre filas
        form.setLabelAlignment(Qt.AlignRight | Qt.AlignVCenter)
        
        self.url_input = QLineEdit()
        self.url_input.setPlaceholderText("https://api.example.com")
        self.url_input.setMinimumWidth(450) # Ancho aumentado para que se vea completa la URL
        form.addRow("URL de API:", self.url_input)
        
        self.token_input = QLineEdit()
        self.token_input.setEchoMode(QLineEdit.Password)
        self.token_input.setPlaceholderText("Token de autenticación")
        self.token_input.setMinimumWidth(450)
        form.addRow("Token:", self.token_input)
        
        self.user_input = QLineEdit()
        self.user_input.setPlaceholderText("Usuario")
        self.user_input.setMinimumWidth(450)
        form.addRow("Usuario:", self.user_input)
        
        self.timeout_input = QLineEdit()
        self.timeout_input.setPlaceholderText("30")
        self.timeout_input.setText("30")
        self.timeout_input.setMinimumWidth(450)
        form.addRow("Timeout (seg):", self.timeout_input)
        card_layout.addLayout(form)
        
        card_layout.addSpacing(10) # Espacio antes de botones

        # Botones
        btn_layout = QHBoxLayout()
        btn_layout.setSpacing(12) # Espacio entre botones para permitir crecimiento
        btn_save = AnimatedButton("Guardar") # Botón primario animado
        btn_save.setStyleSheet("padding: 8px 20px; min-height: 32px; font-size: 13px;")
        btn_save.clicked.connect(self.save_settings)
        btn_cancel = QPushButton("Cancelar") # Botón secundario normal
        btn_cancel.clicked.connect(self.reject)
        
        # Estilo para el botón Cancelar sin subrayado y con espacio suficiente
        btn_cancel.setStyleSheet("""
            QPushButton {
                background-color: transparent;
                color: #8E8E93;
                border: none;
                padding: 8px 16px;
                font-size: 15px;
                text-decoration: none;
            }
            QPushButton:hover {
                color: #007AFF;
                text-decoration: none;
                background-color: transparent;
            }
        """)
        
        btn_layout.addStretch()
        btn_layout.addWidget(btn_cancel)
        btn_layout.addWidget(btn_save)
        card_layout.addLayout(btn_layout)

        # Centrar la tarjeta en el layout raíz
        root.addStretch()
        root.addWidget(card, alignment=Qt.AlignCenter)
        root.addStretch()

    def load_settings(self):
        self.url_input.setText(self.settings.value("api_url", ""))
        self.token_input.setText(self.settings.value("api_token", ""))
        self.user_input.setText(self.settings.value("api_user", ""))
        self.timeout_input.setText(self.settings.value("api_timeout", "30"))

    def save_settings(self):
        self.settings.setValue("api_url", self.url_input.text())
        self.settings.setValue("api_token", self.token_input.text())
        self.settings.setValue("api_user", self.user_input.text())
        self.settings.setValue("api_timeout", self.timeout_input.text())
        self.accept()


# --- Ventana Principal ---
class MainWindow(QMainWindow):

    # --- PDF helpers (abrir local o URL) + botón con icono SVG ---
    def _open_invoice_pdf(self, invoice_id: str, fallback_url: str = "", *, local_path: str = None, cliente: str = None, importe: str = None):
        """Intenta abrir el PDF guardado (según nº de factura o metadatos). Si no lo encuentra, abre la URL."""
        try:
            # 1) Priorizar ruta proporcionada explícitamente (si existe)
            if local_path:
                candidate = local_path
                if not os.path.isabs(candidate):
                    base_dir = ""
                    dest_widget = getattr(self, "txt_pdf_dest", None)
                    if dest_widget:
                        try:
                            base_dir = dest_widget.text().strip()
                        except Exception:
                            base_dir = ""
                    candidate = os.path.abspath(os.path.join(base_dir, candidate))
                candidate = os.path.normpath(candidate)
                if os.path.exists(candidate) and os.path.getsize(candidate) > 0:
                    QDesktopServices.openUrl(QUrl.fromLocalFile(candidate))
                    return
            
            # Directorio configurado en la UI (si existe ese campo)
            dest = getattr(self, "txt_pdf_dest", None).text() if hasattr(self, "txt_pdf_dest") else ""
            if dest and os.path.isdir(dest):
                # Patrones habituales: "Nº - Cliente - Importe.pdf" u otros
                safe_tokens = []
                invoice_token = str(invoice_id or "").strip()
                if invoice_token:
                    safe_tokens.append(invoice_token)
                if cliente:
                    safe_tokens.append(re.sub(r"\s+", " ", str(cliente)))
                if importe:
                    safe_tokens.append(str(importe).replace("€", "").strip())

                patterns = []
                for token in safe_tokens:
                    token = token.strip()
                    if not token:
                        continue
                    patterns.extend([
                        f"{token}*.pdf",
                        f"*{token}*.pdf"
                    ])
                if not patterns:
                    patterns = ["*.pdf"]

                matches = []
                for pat in patterns:
                    matches.extend(glob.glob(os.path.join(dest, pat)))
                matches = [p for p in matches if os.path.getsize(p) > 0]
                if matches:
                    best = max(matches, key=os.path.getmtime)
                    QDesktopServices.openUrl(QUrl.fromLocalFile(best))
                    return
            # Fallback a la URL si no hay PDF local
            if fallback_url:
                webbrowser.open(fallback_url)
            else:
                if hasattr(self, "show_toast"):
                    self.show_toast("⚠️ No se encontró el PDF guardado ni hay URL disponible.")
        except Exception as e:
            if hasattr(self, "show_error"):
                self.show_error(f"No se pudo abrir el PDF: {e}")

    def _make_pdf_button(self, invoice_id: str, pdf_url: str, *, local_path: str = None, cliente: str = None, importe: str = None, svg_path: str = None):
        btn = QToolButton()
        btn.setToolTip("Abrir PDF")
        btn.setMinimumHeight(28)
        btn.setCursor(Qt.PointingHandCursor)
        btn.clicked.connect(
            lambda checked=False, _id=invoice_id, _url=pdf_url, _path=local_path, _cliente=cliente, _importe=importe: self._open_invoice_pdf(
                _id, _url, local_path=_path, cliente=_cliente, importe=_importe
            )
        )

        # Intentar cargar icono SVG
        icon = None
        try:
            if svg_path and os.path.exists(svg_path):
                icon = QIcon(svg_path)
            else:
                default_svg = os.path.join(RESOURCE_DIR, "ver.pdf.svg")
                if os.path.exists(default_svg):
                    icon = QIcon(default_svg)
        except Exception:
            icon = None

        if icon and not icon.isNull():
            btn.setIcon(icon)
            btn.setIconSize(QSize(20, 20))
            btn.setToolButtonStyle(Qt.ToolButtonIconOnly)
            btn.setText("")
        else:
            btn.setText("Ver")  # fallback textual

        return btn

    def select_pdf_destination(self):
        if not hasattr(self, "txt_pdf_dest"):
            return
        start_dir = self.txt_pdf_dest.text().strip()
        if not start_dir or not os.path.isdir(start_dir):
            start_dir = os.path.expanduser("~")
        directory = QFileDialog.getExistingDirectory(
            self,
            "Seleccionar carpeta destino de los PDFs",
            start_dir
        )
        if directory:
            self.txt_pdf_dest.setText(directory)

    def __init__(self):
        super().__init__()
        self.setAttribute(Qt.WA_TranslucentBackground) # Habilitar fondo translúcido
        self.setWindowTitle("FactuNabo")
        self.setMinimumSize(1200, 700)
        self.showMaximized() # <-- ABRIR MAXIMIZADO
        self.theme = "light"
        self.setProperty("theme", self.theme)
        self.loaded_invoice_count = 0  # KPI: facturas cargadas
        # Activar backdrop del sistema (Windows 11): Mica/Acrylic
        try:
            enable_windows_backdrop(self.winId(), dark_mode=(self.theme == "dark"))
        except Exception:
            pass

        icon_path = os.path.join(RESOURCE_DIR, "logo.ico")
        if os.path.exists(icon_path):
            self.setWindowIcon(QIcon(icon_path))

        init_database()
        
        # Cargar configuraciones guardadas
        global COLOR_PRIMARY
        self.settings = QSettings("FactuNabo", "FactuNabo")
        saved_color = self.settings.value("accent_color", COLOR_PRIMARY)
        COLOR_PRIMARY = saved_color
        self.browser_code, self.browser_path = detect_available_browser()
        browser_names = {"chrome": "Google Chrome", "edge": "Microsoft Edge"}
        self.browser_display_name = browser_names.get(self.browser_code, self.browser_code.title())
        
        # Cargar tamaño de fuente guardado
        app = QApplication.instance()
        if app:
            saved_font_size = self.settings.value("font_size", "Normal (15px)")
            size_map = {
                "Pequeño (13px)": 13,
                "Normal (15px)": 15,
                "Grande (17px)": 17,
                "Muy Grande (19px)": 19
            }
            font_size = size_map.get(saved_font_size, 15)
            font = QFont("Segoe UI Variable", font_size)
            font.setStyleStrategy(QFont.StyleStrategy.PreferQuality)
            app.setFont(font)

        # Cargar espaciado guardado (por defecto "Compacto")
        saved_spacing = self.settings.value("spacing", "Compacto")
        self.setProperty("spacing", saved_spacing.lower())
        
        # Carga de estilos
        style_sheet_content = self._get_themed_stylesheet()
        if style_sheet_content:
            # Aplicar a toda la app para herencia consistente (dialogs incluidos)
            app = QApplication.instance()
            if app:
                app.setStyleSheet(style_sheet_content)
            # Asegurar también en la ventana por si el QApplication no existiera aún
            self.setStyleSheet(style_sheet_content)

        # El worker y el thread se crearán bajo demanda en send_facturas
        self.worker = None
        self.thread = None

        self.sending_in_progress = False
        self.current_excel_path = None
        # --- [NUEVO] Atributos para almacenar los dataframes históricos ---
        self.df_factura_historico = None
        self.df_conceptos_historico = None
        # --- [FIN NUEVO] ---
        self.validation_errors = []
        self.post_action_mode = "MARK"
        self._overlay = Overlay(self)

        # --- [MODIFICADO] Referencias del Dashboard ---
        self.total_label = None
        self.success_label = None
        self.month_total_label = None
        self.month_count_label = None
        
        # --- [NUEVO] Referencias a las nuevas tablas y widgets de consulta ---
        self.table_last_errors = None
        self.dash_combo_empresas = None
        self.dash_combo_periodo = None
        self.dash_btn_consultar = None
        self.dash_label_resultado = None
        self.dash_table_resultados = None
        # --- [FIN NUEVO] ---
        
        # --- [NUEVO] Referencia al logo ---
        self.logo_label = None
        # --- [FIN NUEVO] ---

        self.init_ui()

        # Atajos de teclado
        self.shortcut_open = QShortcut(QKeySequence("Ctrl+O"), self)
        self.shortcut_open.activated.connect(self.select_excel)
        
        self.shortcut_send = QShortcut(QKeySequence("Ctrl+S"), self)
        self.shortcut_send.activated.connect(lambda: self.btn_send.click() if self.btn_send.isEnabled() else None)
        
        self.shortcut_search = QShortcut(QKeySequence("Ctrl+F"), self)
        self.shortcut_search.activated.connect(self.focus_search)
        
        self.shortcut_dashboard = QShortcut(QKeySequence("Ctrl+D"), self)
        self.shortcut_dashboard.activated.connect(lambda: self.nav_list.setCurrentRow(0))
        
        self.shortcut_history = QShortcut(QKeySequence("Ctrl+H"), self)
        self.shortcut_history.activated.connect(lambda: self.nav_list.setCurrentRow(3))
        
        self.shortcut_config = QShortcut(QKeySequence("Ctrl+,"), self)
        self.shortcut_config.activated.connect(lambda: self.nav_list.setCurrentRow(4))

        self.toast_timer = QTimer(self)
        self.toast_timer.timeout.connect(self.hide_toast)

        # Hilo temporal para descarga manual de PDFs
        self.dthread = None
        self.dworker = None

    def _get_themed_stylesheet(self):
        """Reads the QSS file and replaces placeholders with resource paths and colors."""
        qss_path = resource_path("styles.qss")
        if not os.path.exists(qss_path):
            return ""

        with open(qss_path, "r", encoding="utf-8") as f:
            qss = f.read()

        # Replace placeholders with actual paths
        logo_light_path = resource_path(os.path.join("resources", "logo_light.png")).replace("\\", "/")
        logo_dark_path = resource_path(os.path.join("resources", "logo_dark.png")).replace("\\", "/")

        qss = qss.replace("%%LOGO_LIGHT%%", logo_light_path)
        qss = qss.replace("%%LOGO_DARK%%", logo_dark_path)
        
        # Replace color placeholders with current accent color
        # Calculate darker shade for hover (reduce brightness by ~15%)
        from PySide6.QtGui import QColor
        base_color = QColor(COLOR_PRIMARY)
        darker_color = base_color.darker(115)  # 15% darker
        
        # Replace all instances of the default green color with the accent color
        # IMPORTANTE: Solo reemplazar en background-color, no en color (texto)
        # Usar regex para ser más preciso y evitar reemplazar el color del texto
        import re
        # Reemplazar solo en background-color
        qss = re.sub(r'background-color:\s*#A0BF6E', f'background-color: {COLOR_PRIMARY}', qss)
        qss = re.sub(r'background-color:\s*#a0bf6e', f'background-color: {COLOR_PRIMARY.lower()}', qss, flags=re.IGNORECASE)
        qss = re.sub(r'background-color:\s*#87a15D', f'background-color: {darker_color.name()}', qss)
        qss = re.sub(r'background-color:\s*#87a15d', f'background-color: {darker_color.name().lower()}', qss, flags=re.IGNORECASE)
        
        # También reemplazar en border-color si aparece
        qss = re.sub(r'border.*:\s*[^;]*#A0BF6E', lambda m: m.group(0).replace('#A0BF6E', COLOR_PRIMARY), qss)
        qss = re.sub(r'border.*:\s*[^;]*#87a15D', lambda m: m.group(0).replace('#87a15D', darker_color.name()), qss)

        return qss

    # ----- Login gating -----
    # [REVISADO] require_login con centrado y estilo más robustos
    def require_login(self):
        self._overlay.setGeometry(self.rect())
        self._overlay.show()
        dlg = LoginDialog(self)

        # --- [MODIFICADO: Aplicar estilo directamente] ---
        # Cargar el QSS general y aplicarlo al diálogo
        style_sheet_content = self._get_themed_stylesheet()
        if style_sheet_content:
            dlg.setStyleSheet(style_sheet_content)
            dlg.setProperty("theme", self.theme)

        # Forzar un repintado inicial del diálogo
        dlg.style().unpolish(dlg)
        dlg.style().polish(dlg)
        dlg.adjustSize() # Asegura que el tamaño se calcule con el nuevo estilo
        # --- [FIN MODIFICADO] ---

        # --- [Centrado respecto a la pantalla] ---
        screen = QApplication.primaryScreen()
        if screen: # Asegurarse de que tenemos una pantalla
            screen_geometry = screen.availableGeometry()
            # Usamos geometry() DESPUÉS de adjustSize() para obtener el tamaño correcto
            dlg_geometry = dlg.geometry()
            center_point = screen_geometry.center() - QPoint(dlg_geometry.width() // 2, dlg_geometry.height() // 2)
            dlg.move(center_point)
        # --- [FIN CENTRADO] ---

        res = dlg.exec()
        if res == QDialog.Accepted:
            self._overlay.hide()
            return True
        else:
            QTimer.singleShot(0, self.close)
            return False


    # ----- UI -----
    def init_ui(self):
        main_widget = QWidget()
        main_layout = QHBoxLayout(main_widget)
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)

        splitter = QSplitter(Qt.Horizontal)
        splitter.setHandleWidth(0) # <-- ASEGURADO que sea 0

        # --- [AÑADIDO ESTILO DIRECTO AL SPLITTER] ---
        # Forzar estilo invisible directamente en el splitter para eliminar la línea
        splitter.setStyleSheet("QSplitter::handle { background-color: transparent; border: none; width: 0px; image: none; }")
        # --- [FIN AÑADIDO] ---

        # Sidebar
        sidebar = QWidget()
        sidebar.setProperty("class", "Sidebar")
        sidebar.setFixedWidth(270)
        sidebar_layout = QVBoxLayout(sidebar)
        sidebar_layout.setContentsMargins(15, 20, 15, 20)
        sidebar_layout.setSpacing(10)

        logo_container = QHBoxLayout()
        
        # --- [MODIFICADO PARA LOGO TEMATIZADO] ---
        # Ya no cargamos QPixmap aquí, solo creamos el QLabel con un ID
        self.logo_label = QLabel() 
        self.logo_label.setObjectName("SidebarLogo") # ID para el QSS
        self.logo_label.setMinimumSize(100, 100) # Darle un tamaño fijo
        self.logo_label.setScaledContents(False) # Dejamos que QSS controle (con 'image:')
        self.logo_label.setAlignment(Qt.AlignCenter) 
        logo_container.addWidget(self.logo_label)
        # --- [FIN MODIFICACIÓN] ---
        
        title = QLabel("FactuNabo")
        title.setObjectName("sidebarTitle")
        logo_container.addWidget(title)
        logo_container.addStretch()
        sidebar_layout.addLayout(logo_container)

        sidebar_layout.addSpacing(20)

        self.nav_list = AnimatedNavList()

        # --- [Revertido a Emojis] ---
        menu_items = [
            ("📊 Dashboard", 0),
            ("📁 Cargar Excel", 1),
            ("🚀 Enviar Facturas", 2), # <-- El cohete
            ("📜 Histórico", 3),
            ("⚙️ Configuración", 4),
        ]

        for text, index in menu_items:
            item = QListWidgetItem(text) # Item solo con texto (incluye emoji)
            item.setData(Qt.UserRole, index)
            self.nav_list.addItem(item)
        # --- [Fin Revertido a Emojis] ---

        self.nav_list.setCurrentRow(0)
        self.nav_list.currentRowChanged.connect(self.change_page)
        sidebar_layout.addWidget(self.nav_list)

        sidebar_layout.addStretch()

        self.dark_toggle = QCheckBox("Modo Oscuro")
        self.dark_toggle.toggled.connect(self.toggle_theme)
        sidebar_layout.addWidget(self.dark_toggle)

        splitter.addWidget(sidebar)

        # Content
        self.content_stack = QStackedWidget()
        self.content_stack.setProperty("class", "ContentStack")

        self.dashboard_page = self.create_dashboard_page()
        self.content_stack.addWidget(self.dashboard_page)

        self.excel_page = self.create_excel_page()
        self.content_stack.addWidget(self.excel_page)

        self.send_page = self.create_send_page()
        self.content_stack.addWidget(self.send_page)

        self.history_page = self.create_history_page()
        self.content_stack.addWidget(self.history_page)

        self.config_page = self.create_config_page()
        self.content_stack.addWidget(self.config_page)

        splitter.addWidget(self.content_stack)
        splitter.setStretchFactor(1, 1)

        main_layout.addWidget(splitter)
        self.setCentralWidget(main_widget)

        # Toast
        self.toast = QLabel(self)
        self.toast.setObjectName("toast")
        self.toast.setAlignment(Qt.AlignCenter)
        self.toast.hide()
        self.toast_anim = QPropertyAnimation(self.toast, b"geometry")
        self.toast_anim.setEasingCurve(QEasingCurve.OutCubic)

    # [MODIFICADO] create_dashboard_page ahora es mucho más completo
    def create_dashboard_page(self):
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setContentsMargins(40, 40, 40, 40)
        layout.setSpacing(20) # Reducimos el espaciado

        # --- TÍTULO ---
        header = QLabel("DASHBOARD")
        header.setFont(QFont(QApplication.font().family(), 28, QFont.Bold)) 
        header.setStyleSheet("margin-bottom: 5px;")
        layout.addWidget(header)

        # --- Subtítulo de Estadísticas ---
        stats_title = QLabel("Estadísticas Clave")
        stats_title.setFont(QFont(QApplication.font().family(), 18, QFont.Bold))
        layout.addWidget(stats_title)

        stats_layout = QHBoxLayout()
        # Las tarjetas se crean aquí, y guardan la referencia en self.X_label
        total_card = self.create_stat_card("Total Enviados", "0", COLOR_SECONDARY_TEXT)
        stats_layout.addWidget(total_card)
        success_card = self.create_stat_card("Éxitos (Total)", "0", COLOR_SECONDARY_TEXT)
        stats_layout.addWidget(success_card)
        # --- TARJETAS MENSUALES ---
        month_total_card = self.create_stat_card("Facturado (Mes)", "0,00€", COLOR_PRIMARY)
        stats_layout.addWidget(month_total_card)
        month_count_card = self.create_stat_card("Envíos (Mes)", "0", COLOR_SUCCESS)
        stats_layout.addWidget(month_count_card)
        # --- FIN TARJETAS MENSUALES ---
        layout.addLayout(stats_layout)

        layout.addSpacing(15)

        # --- [NUEVO] Sección de Paneles de Consulta ---
        panels_layout = QHBoxLayout()
        panels_layout.setSpacing(20)

        # --- Panel 1: Consulta por Emisor ---
        consulta_card = QFrame()
        consulta_card.setProperty("class", "ConfigGroup") # Reusamos el estilo de "tarjeta"
        apply_shadow(consulta_card, blur=20, offset_y=3)
        consulta_layout = QVBoxLayout(consulta_card)
        
        consulta_title = QLabel("Consulta por Emisor")
        consulta_title.setFont(QFont(QApplication.font().family(), 16, QFont.Bold))
        consulta_layout.addWidget(consulta_title)
        
        # Formulario de consulta
        form_layout = QFormLayout()
        form_layout.setSpacing(10)
        self.dash_combo_empresas = QComboBox()
        self.dash_combo_periodo = QComboBox()
        self.dash_combo_periodo.addItems([
            "1º Trimestre",
            "2º Trimestre",
            "3º Trimestre",
            "4º Trimestre",
            "Ejercicio Actual",
            "Total Histórico"
        ])
        form_layout.addRow("Empresa Emisora:", self.dash_combo_empresas)
        form_layout.addRow("Periodo:", self.dash_combo_periodo)
        consulta_layout.addLayout(form_layout)

        # Botón de consulta
        self.dash_btn_consultar = AnimatedButton("🔍 Consultar")
        self.dash_btn_consultar.setStyleSheet("padding: 8px 20px; min-height: 32px; font-size: 13px;")
        self.dash_btn_consultar.clicked.connect(self.run_dashboard_query)
        consulta_layout.addWidget(self.dash_btn_consultar)

        # Línea separadora
        line = QFrame()
        line.setFrameShape(QFrame.HLine)
        line.setProperty("class", "SeparatorLine")
        consulta_layout.addWidget(line)

        # --- [NUEVO] Barra de búsqueda ---
        self.dash_search_bar = QLineEdit()
        self.dash_search_bar.setPlaceholderText("Buscar en resultados...")
        self.dash_search_bar.textChanged.connect(self.filter_dashboard_table)
        consulta_layout.addWidget(self.dash_search_bar)

        # Resultados de consulta
        self.dash_label_resultado = QLabel("Selecciona filtros y pulsa 'Consultar'")
        self.dash_label_resultado.setObjectName("excelSummary") # Reusamos estilo
        self.dash_label_resultado.setAlignment(Qt.AlignCenter)
        consulta_layout.addWidget(self.dash_label_resultado)
        consulta_layout.addWidget(TableTools(self.dash_table_resultados))

        self.dash_table_resultados = ModernTable(0, 6)
        self.dash_table_resultados.setHorizontalHeaderLabels(["Fecha", "Factura", "Cliente", "Empresa Emisora", "Importe", "Ver Factura"])
        hdr_res = self.dash_table_resultados.horizontalHeader()
        hdr_res.setSectionResizeMode(0, QHeaderView.ResizeToContents) # Fecha
        hdr_res.setSectionResizeMode(1, QHeaderView.Stretch) # Factura
        hdr_res.setSectionResizeMode(2, QHeaderView.Stretch) # Cliente
        hdr_res.setSectionResizeMode(3, QHeaderView.Stretch) # Empresa Emisora
        hdr_res.setSectionResizeMode(4, QHeaderView.ResizeToContents) # Importe
        hdr_res.setSectionResizeMode(5, QHeaderView.ResizeToContents) # Ver Factura
        consulta_layout.addWidget(self.dash_table_resultados, 1)

        panels_layout.addWidget(consulta_card, 1) # '1' para que ocupe más espacio

        layout.addLayout(panels_layout, 1) # El '1' hace que se expanda verticalmente

        layout.addSpacing(15)
        # --- [FIN NUEVA SECCIÓN] ---

        actions_label = QLabel("Acciones Rápidas")
        actions_label.setFont(QFont(QApplication.font().family(), 18, QFont.Bold)) # Usa fuente global
        layout.addWidget(actions_label)
        actions_layout = QHBoxLayout()
        btn_load = AnimatedButton("📁 Cargar Excel")
        btn_load.setToolTip("Cargar archivo Excel con facturas (Ctrl+O)")
        btn_load.setStyleSheet("padding: 8px 20px; min-height: 32px; font-size: 13px;")
        btn_load.clicked.connect(lambda: self.nav_list.setCurrentRow(1))
        actions_layout.addWidget(btn_load)
        btn_send = AnimatedButton("🚀 Enviar Facturas")
        btn_send.setToolTip("Enviar facturas a la API de Facturantia")
        btn_send.setStyleSheet("padding: 8px 20px; min-height: 32px; font-size: 13px;")
        btn_send.clicked.connect(lambda: self.nav_list.setCurrentRow(2))
        actions_layout.addWidget(btn_send)
        btn_history = AnimatedButton("📜 Ver Histórico")
        btn_history.setToolTip("Ver historial de envíos realizados")
        btn_history.setStyleSheet("padding: 8px 20px; min-height: 32px; font-size: 13px;")
        btn_history.clicked.connect(lambda: self.nav_list.setCurrentRow(3))
        actions_layout.addWidget(btn_history)
        layout.addLayout(actions_layout)
        
        layout.addStretch(0) # Modificado a 0
        
        # Cargar datos iniciales
        QTimer.singleShot(100, self.update_dashboard_stats)
        QTimer.singleShot(150, self.populate_dashboard_filters) # Cargar filtros después
        return page

    # [MODIFICADO] create_stat_card para asignar nuevas referencias
    def create_stat_card(self, title, value, color):
        card = QFrame()
        card.setProperty("class", "StatCard")
        card_layout = QVBoxLayout(card)
        value_label = QLabel(value)
        # Usamos la fuente global pero ajustamos tamaño/peso
        value_label.setFont(QFont(QApplication.font().family(), 36, QFont.Bold))
        value_label.setStyleSheet(f"color: {color};")
        title_label = QLabel(title)
        title_label.setStyleSheet(f"color: {COLOR_SECONDARY_TEXT};") # Mantiene estilo secundario
        card_layout.addWidget(value_label)
        card_layout.addWidget(title_label)

        # Guardamos la referencia a la etiqueta del valor
        if "Total Enviados" in title:
            self.total_label = value_label
        elif "Éxitos" in title or "Éxitos (Total)" in title:
            self.success_label = value_label
        elif "Facturado (Mes)" in title:  # <-- [NUEVO]
            self.month_total_label = value_label
        elif "Envíos (Mes)" in title: # <-- [NUEVO]
            self.month_count_label = value_label

        apply_shadow(card, blur=25, offset_y=5)

        return card

    def create_excel_page(self):
        page = QWidget()
        page.setAcceptDrops(True)
        page.dragEnterEvent = self.dragEnterEvent
        page.dropEvent = self.dropEvent
        layout = QVBoxLayout(page)
        layout.setContentsMargins(40, 40, 40, 40)
        layout.setSpacing(20)
        header = QLabel("CARGAR ARCHIVO EXCEL")
        header.setFont(QFont(QApplication.font().family(), 32, QFont.Bold)) # Usa fuente global
        layout.addWidget(header)
        self.stepper = StepperWidget(["Cargar", "Validar", "Listo"])
        layout.addWidget(self.stepper)
        self.btn_select_excel = AnimatedButton("📁 Seleccionar Excel")
        self.btn_select_excel.setToolTip("Seleccionar archivo Excel con facturas para procesar (Ctrl+O)")
        self.btn_select_excel.setStyleSheet("padding: 8px 20px; min-height: 32px; font-size: 13px;")
        self.btn_select_excel.clicked.connect(self.select_excel)
        layout.addWidget(self.btn_select_excel)

        # --- Botón de limpieza manual ---
        btn_clear = AnimatedButton("🧹 Limpiar")
        btn_clear.setToolTip("Limpiar tabla y datos cargados")
        btn_clear.setStyleSheet("padding: 8px 20px; min-height: 32px; font-size: 13px;")
        btn_clear.clicked.connect(self.clear_excel_table)
        layout.addWidget(btn_clear)

        hint = QLabel("o arrastra el archivo aquí")
        hint.setAlignment(Qt.AlignCenter)
        hint.setStyleSheet(f"color: {COLOR_SECONDARY_TEXT};")
        layout.addWidget(hint)
        self.validation_label = QLabel("")
        self.validation_label.setWordWrap(True)
        layout.addWidget(self.validation_label)
        # KPI contador + total base
        self.loaded_summary = QLabel("")
        self.loaded_summary.setObjectName("excelSummary")
        self.loaded_summary.setVisible(False)
        layout.addWidget(self.loaded_summary)

        # Contenedor redondeado para la tabla de carga de Excel
        excel_card = QFrame()
        excel_card.setProperty("class", "TableCard")
        apply_shadow(excel_card, blur=20, offset_y=3)
        excel_layout = QVBoxLayout(excel_card)
        excel_layout.setContentsMargins(0, 0, 0, 0)
        excel_layout.setSpacing(0)

        self.table_excel = ModernTable(0, 8)
        self.table_excel.setHorizontalHeaderLabels([
            "Factura", "Empresa Emisora", "Cliente", "Base Imponible",
            "Cantidad IVA", "Retención", "Importe Total", "Fecha"
        ])

        # --- Ajuste de columnas ---
        hdr = self.table_excel.horizontalHeader()
        hdr.setSectionResizeMode(0, QHeaderView.ResizeToContents)    # Factura
        hdr.setSectionResizeMode(1, QHeaderView.Stretch)            # Empresa Emisora
        hdr.setSectionResizeMode(2, QHeaderView.Stretch)            # Cliente
        hdr.setSectionResizeMode(3, QHeaderView.ResizeToContents)    # Base Imponible
        hdr.setSectionResizeMode(4, QHeaderView.ResizeToContents)    # Cantidad IVA
        hdr.setSectionResizeMode(5, QHeaderView.ResizeToContents)    # Retención
        hdr.setSectionResizeMode(6, QHeaderView.ResizeToContents)    # Importe Total
        hdr.setSectionResizeMode(7, QHeaderView.ResizeToContents)    # Fecha

        excel_layout.addWidget(TableTools(self.table_excel))
        excel_layout.addWidget(self.table_excel)
        layout.addWidget(excel_card)
        layout.addStretch()
        return page

    # [MODIFICADO] create_send_page con layout más compacto
    def create_send_page(self):
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setContentsMargins(36, 16, 36, 16) # Reducir márgenes
        layout.setSpacing(8) # Espaciado más compacto

        header = QLabel("ENVIAR FACTURAS")
        header.setFont(QFont(QApplication.font().family(), 32, QFont.Bold)) # Usa fuente global
        layout.addWidget(header)

        self.send_stepper = StepperWidget(["Preparar", "Enviar", "Generar PDF", "Completado"])
        self.send_stepper.layout().setContentsMargins(20, 2, 20, 2) # Menos margen vertical en stepper
        self.send_stepper.setFixedHeight(64)
        layout.addWidget(self.send_stepper)

        # --- Grupo 1: Opciones de Envío ---
        group1_frame = QFrame()
        group1_frame.setProperty("class", "ConfigGroup")
        group1_frame.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        group1_frame.setMaximumHeight(150)
        apply_shadow(group1_frame, blur=20, offset_y=3)
        group1_layout = QVBoxLayout(group1_frame)
        group1_layout.setContentsMargins(20, 14, 20, 14)
        group1_layout.setSpacing(8) # Reducir espaciado interno del grupo
        g1_title = QLabel("Acción Post-Envío")
        g1_title.setFont(QFont(QApplication.font().family(), 16, QFont.Bold)) # Usa fuente global
        group1_layout.addWidget(g1_title)
        post_row = QHBoxLayout()
        post_row.addWidget(QLabel("Acción sobre la Macro:"))
        self.post_action_combo = QComboBox()
        self.post_action_combo.addItem("Marcar estado en Macro (col. AC)", userData="MARK")
        self.post_action_combo.addItem("Borrar filas enviadas OK/Duplicadas", userData="DELETE_OK")
        self.post_action_combo.currentIndexChanged.connect(self._on_post_action_changed)
        post_row.addWidget(self.post_action_combo)
        post_row.addStretch()
        group1_layout.addLayout(post_row)
        layout.addWidget(group1_frame)


        # --- Grupo 2: Opciones de Descarga PDF ---
        group2_frame = QFrame()
        group2_frame.setProperty("class", "ConfigGroup")
        group2_frame.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        group2_frame.setMaximumHeight(190)
        apply_shadow(group2_frame, blur=20, offset_y=3)
        group2_layout = QVBoxLayout(group2_frame)
        group2_layout.setContentsMargins(20, 14, 20, 14)
        group2_layout.setSpacing(8) # Reducir espaciado interno del grupo
        g2_title = QLabel("Descarga de PDFs")
        g2_title.setFont(QFont(QApplication.font().family(), 16, QFont.Bold)) # Usa fuente global
        group2_layout.addWidget(g2_title)
        self.chk_auto_download = QCheckBox("Descargar PDFs automáticamente al terminar")
        self.chk_auto_download.setChecked(False)
        group2_layout.addWidget(self.chk_auto_download)
        dl_row = QFormLayout()
        dl_row.setSpacing(10)

        path_widget = QWidget()
        path_layout = QHBoxLayout(path_widget)
        path_layout.setContentsMargins(0, 0, 0, 0)
        path_layout.setSpacing(6)
        self.txt_pdf_dest = QLineEdit(
            r"C:\Users\administracionaba\Desktop\FACTURANTIA\FactuNabo EXE\Facturas PDF" # Considera usar una ruta relativa o configurable
        )
        self.txt_pdf_dest.setPlaceholderText("Carpeta donde guardar los PDFs generados")
        path_layout.addWidget(self.txt_pdf_dest, 1)
        self.btn_pdf_dest = QToolButton()
        self.btn_pdf_dest.setText("...")
        self.btn_pdf_dest.setToolTip("Seleccionar carpeta de destino")
        self.btn_pdf_dest.setCursor(Qt.PointingHandCursor)
        self.btn_pdf_dest.setFixedWidth(36)
        self.btn_pdf_dest.clicked.connect(self.select_pdf_destination)
        path_layout.addWidget(self.btn_pdf_dest, 0)
        dl_row.addRow("Carpeta destino:", path_widget)

        self.lbl_browser_detected = QLabel(f"Navegador detectado: {self.browser_display_name}")
        self.lbl_browser_detected.setProperty("descriptionLabel", True)
        dl_row.addRow("Navegador:", self.lbl_browser_detected)
        group2_layout.addLayout(dl_row)
        layout.addWidget(group2_frame)


        # --- Botones de Acción Principales ---
        action_layout = QHBoxLayout()
        action_layout.setSpacing(12) # Espacio entre botones para permitir crecimiento
        self.btn_send = AnimatedButton("🚀 Iniciar Envío")
        self.btn_send.setToolTip("Iniciar el envío de facturas a la API (Ctrl+S)")
        self.btn_send.setStyleSheet("padding: 6px 18px; min-height: 28px; font-size: 13px;")
        self.btn_send.clicked.connect(self.send_facturas)
        self.btn_send.setEnabled(False)
        action_layout.addWidget(self.btn_send)
        self.btn_download_pdfs = AnimatedButton("📥 Guardar PDFs")
        self.btn_download_pdfs.setToolTip("Descargar los PDFs de las facturas del último envío")
        self.btn_download_pdfs.setStyleSheet("padding: 6px 18px; min-height: 28px; font-size: 13px;")
        self.btn_download_pdfs.setEnabled(False)
        self.btn_download_pdfs.clicked.connect(self.download_pdfs_clicked)
        action_layout.addWidget(self.btn_download_pdfs)
        layout.addLayout(action_layout)


        # --- Barra de Progreso ---
        self.progress = QProgressBar()
        self.progress.setTextVisible(True)
        self.progress.setFixedHeight(18)
        self.progress_anim = QPropertyAnimation(self.progress, b"value")
        self.progress_anim.setEasingCurve(QEasingCurve.InOutSine)
        layout.addWidget(self.progress)


        # --- [NUEVO] Tabla de Previsualización ---
        self.preview_title = QLabel("Previsualización de Facturas a Enviar")
        self.preview_title.setFont(QFont(QApplication.font().family(), 16, QFont.Bold))
        layout.addWidget(self.preview_title)

        # Contenedor redondeado para la tabla de previsualización
        self.preview_card = QFrame()
        self.preview_card.setProperty("class", "TableCard")
        self.preview_card.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        apply_shadow(self.preview_card, blur=20, offset_y=3)
        preview_layout = QVBoxLayout(self.preview_card)
        preview_layout.setContentsMargins(0, 0, 0, 0)
        preview_layout.setSpacing(0)

        self.table_preview = ModernTable(0, 5)
        self.table_preview.setHorizontalHeaderLabels(["Factura", "Empresa Emisora", "Cliente", "Importe Total", "Fecha"])
        hp = self.table_preview.horizontalHeader()
        hp.setSectionResizeMode(0, QHeaderView.ResizeToContents)
        hp.setSectionResizeMode(1, QHeaderView.Stretch)
        hp.setSectionResizeMode(2, QHeaderView.Stretch)
        hp.setSectionResizeMode(3, QHeaderView.ResizeToContents)
        hp.setSectionResizeMode(4, QHeaderView.ResizeToContents)
        preview_layout.addWidget(TableTools(self.table_preview))
        preview_layout.addWidget(self.table_preview)
        layout.addWidget(self.preview_card, 1)


        # --- [MODIFICADO] Grupo 3 ahora es un atributo de clase (self.results_group) y está oculto ---
        self.results_group = QFrame()
        self.results_group.setProperty("class", "ConfigGroup")
        apply_shadow(self.results_group, blur=20, offset_y=3)
        group3_layout = QVBoxLayout(self.results_group)
        group3_layout.setSpacing(10) # Reducir espaciado interno
        g3_title = QLabel("Resultados del Envío")
        g3_title.setFont(QFont(QApplication.font().family(), 16, QFont.Bold)) # Usa fuente global
        group3_layout.addWidget(g3_title)
        filter_layout = QHBoxLayout()
        self.search_bar = QLineEdit()
        self.search_bar.setPlaceholderText("Buscar por factura o empresa...")
        
        # ######################################################################
        # INICIO DE LA CORRECCIÓN DE ERROR
        # La línea de abajo era la que fallaba. Ahora 'self.apply_search' existe.
        # ######################################################################
        self.search_bar.textChanged.connect(self.apply_search)
        
        filter_layout.addWidget(self.search_bar)
        self.filters = {}
        for status, _color in [("ÉXITO", COLOR_SUCCESS), ("DUPLICADO", COLOR_WARNING), ("ERROR", COLOR_ERROR)]:
            btn = QPushButton(status)
            btn.setCheckable(True)
            btn.setProperty("filter", "true")
            btn.setProperty("status", status)
            
            # ######################################################################
            # INICIO DE LA CORRECCIÓN DE ERROR
            # La línea de abajo fallaría después. Ahora 'self.apply_filter' existe.
            # ######################################################################
            btn.clicked.connect(self.apply_filter)
            
            self.filters[status] = btn
            filter_layout.addWidget(btn)
        group3_layout.addLayout(filter_layout)
        self.table_envio = ModernTable(0, 7)
        self.table_envio.setHorizontalHeaderLabels(["Factura", "Empresa Emisora", "Cliente", "Importe", "Estado", "Detalles", "Ver Factura"])
        henv = self.table_envio.horizontalHeader()
        self.table_envio.verticalHeader().setDefaultSectionSize(36)
        henv.setStretchLastSection(False)
        henv.setSectionResizeMode(0, QHeaderView.ResizeToContents)
        henv.setSectionResizeMode(1, QHeaderView.Stretch)
        henv.setSectionResizeMode(2, QHeaderView.Stretch)
        henv.setSectionResizeMode(3, QHeaderView.ResizeToContents)
        henv.setSectionResizeMode(4, QHeaderView.ResizeToContents)
        henv.setSectionResizeMode(5, QHeaderView.Stretch)
        henv.setSectionResizeMode(6, QHeaderView.ResizeToContents)
        group3_layout.addWidget(self.table_envio)
        self.log_area = QTextEdit()
        self.log_area.setReadOnly(True)
        self.log_area.setMaximumHeight(120) # Reducir altura máx log
        group3_layout.addWidget(self.log_area)
        self.btn_export = AnimatedButton("📤 Exportar Resultados")
        self.btn_export.setStyleSheet("padding: 8px 20px; min-height: 32px; font-size: 13px;")
        self.btn_export.clicked.connect(self.export_results)
        group3_layout.addWidget(self.btn_export)

        # Añadimos el results_group con factor de estiramiento y lo ocultamos
        layout.addWidget(self.results_group, 3) # Mayor peso para aprovechar altura
        self.results_group.setVisible(False)

        # Distribuir alturas: priorizar tabla de resultados
        layout.setStretchFactor(group1_frame, 0)
        layout.setStretchFactor(group2_frame, 0)
        layout.setStretchFactor(action_layout, 0)
        layout.setStretchFactor(self.progress, 0)
        layout.setStretchFactor(self.preview_card, 1)

        return page

    def create_history_page(self):
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setContentsMargins(40, 40, 40, 40)
        layout.setSpacing(20)
        header_layout = QHBoxLayout()
        header = QLabel("HISTÓRICO DE ENVÍOS")
        header.setFont(QFont(QApplication.font().family(), 32, QFont.Bold)) # Usa fuente global
        header_layout.addWidget(header)
        header_layout.addStretch()
        
        # Botón de exportar
        btn_export = AnimatedButton("📥 Exportar")
        btn_export.setToolTip("Exportar historial a Excel o CSV")
        btn_export.setStyleSheet("padding: 8px 20px; min-height: 32px; font-size: 13px;")
        btn_export.clicked.connect(self.export_history)
        header_layout.addWidget(btn_export)
        
        btn_refresh = AnimatedButton("🔄 Actualizar")
        btn_refresh.setToolTip("Actualizar lista de envíos")
        btn_refresh.setStyleSheet("padding: 8px 20px; min-height: 32px; font-size: 13px;")
        btn_refresh.clicked.connect(self.load_history)
        header_layout.addWidget(btn_refresh)
        layout.addLayout(header_layout)
        
        # [NUEVO] Filtros de búsqueda avanzada
        filters_card = QFrame()
        filters_card.setProperty("class", "ConfigGroup")
        filters_layout = QVBoxLayout(filters_card)
        filters_layout.setContentsMargins(20, 15, 20, 15)
        filters_layout.setSpacing(12)
        
        filters_title = QLabel("Filtros de Búsqueda")
        filters_title.setFont(QFont(QApplication.font().family(), 14, QFont.Bold))
        filters_layout.addWidget(filters_title)
        
        filters_row1 = QHBoxLayout()
        filters_row1.addWidget(QLabel("Empresa:"))
        self.history_filter_empresa = QComboBox()
        self.history_filter_empresa.addItem("Todas las Empresas")
        self.history_filter_empresa.setToolTip("Filtrar por empresa emisora")
        filters_row1.addWidget(self.history_filter_empresa)
        
        filters_row1.addWidget(QLabel("Estado:"))
        self.history_filter_estado = QComboBox()
        self.history_filter_estado.addItems(["Todos", "ÉXITO", "DUPLICADO", "ERROR"])
        self.history_filter_estado.setToolTip("Filtrar por estado de envío")
        filters_row1.addWidget(self.history_filter_estado)
        
        filters_row1.addWidget(QLabel("Período:"))
        self.history_filter_periodo = QComboBox()
        self.history_filter_periodo.addItems(["Todos", "1º Trimestre", "2º Trimestre", "3º Trimestre", "4º Trimestre", "Este mes", "Mes anterior"])
        self.history_filter_periodo.setToolTip("Filtrar por período de tiempo")
        filters_row1.addWidget(self.history_filter_periodo)
        filters_row1.addStretch()
        filters_layout.addLayout(filters_row1)
        
        # Conectar cambios de filtros
        self.history_filter_empresa.currentTextChanged.connect(self.apply_history_filters)
        self.history_filter_estado.currentTextChanged.connect(self.apply_history_filters)
        self.history_filter_periodo.currentTextChanged.connect(self.apply_history_filters)
        
        filters_row2 = QHBoxLayout()
        filters_row2.addWidget(QLabel("Buscar:"))
        self.history_search = QLineEdit()
        self.history_search.setPlaceholderText("Buscar por factura, cliente, empresa...")
        self.history_search.setToolTip("Búsqueda de texto en facturas, clientes y empresas")
        self.history_search.textChanged.connect(self.apply_history_filters)
        filters_row2.addWidget(self.history_search)
        
        btn_clear_filters = QPushButton("Limpiar Filtros")
        btn_clear_filters.setToolTip("Limpiar todos los filtros aplicados")
        btn_clear_filters.clicked.connect(self.clear_history_filters)
        filters_row2.addWidget(btn_clear_filters)
        filters_layout.addLayout(filters_row2)
        
        layout.addWidget(filters_card)

        # [MODIFICADO] Contenedor redondeado para el histórico
        history_card = QFrame()
        history_card.setProperty("class", "TableCard")
        apply_shadow(history_card, blur=20, offset_y=3)
        history_layout = QVBoxLayout(history_card)
        history_layout.setContentsMargins(0, 0, 0, 0)
        history_layout.setSpacing(0)

        # Histórico ahora muestra 8 columnas (incluye importe)
        self.table_history = ModernTable(0, 9)
        self.table_history.setHorizontalHeaderLabels(["ID", "Fecha", "Factura", "Empresa Emisora", "Cliente", "Importe", "Estado", "Detalles", "PDF"])
        # Ajuste de columnas histórico
        hh = self.table_history.horizontalHeader()
        hh.setStretchLastSection(False)
        hh.setSectionResizeMode(0, QHeaderView.ResizeToContents)  # ID
        hh.setSectionResizeMode(1, QHeaderView.ResizeToContents)  # Fecha
        hh.setSectionResizeMode(2, QHeaderView.ResizeToContents)  # Factura
        hh.setSectionResizeMode(3, QHeaderView.Stretch)  # Empresa
        hh.setSectionResizeMode(4, QHeaderView.Stretch)  # Cliente
        hh.setSectionResizeMode(5, QHeaderView.ResizeToContents)  # Importe
        hh.setSectionResizeMode(6, QHeaderView.ResizeToContents)  # Estado
        hh.setSectionResizeMode(7, QHeaderView.Stretch)  # Detalles
        hh.setSectionResizeMode(8, QHeaderView.ResizeToContents)  # PDF

        history_layout.addWidget(TableTools(self.table_history))
        history_layout.addWidget(self.table_history)
        layout.addWidget(history_card)
        self.load_history()
        return page

    # [MODIFICADO] create_config_page ahora usa ConfigGroup y tiene botón Borrar Histórico
    def create_config_page(self):
        page = QWidget()
        layout = QVBoxLayout(page)
        layout.setContentsMargins(40, 40, 40, 40)
        layout.setSpacing(20)

        header = QLabel("CONFIGURACIÓN")
        header.setFont(QFont(QApplication.font().family(), 32, QFont.Bold)) # Usa fuente global
        layout.addWidget(header)

        # --- Grupo 0: Personalización ---
        group0_frame = QFrame()
        group0_frame.setProperty("class", "ConfigGroup")
        apply_shadow(group0_frame, blur=20, offset_y=3)
        group0_layout = QVBoxLayout(group0_frame)
        group0_layout.setSpacing(10)  # Reducido de 15 a 10

        g0_title = QLabel("Personalización")
        g0_title.setFont(QFont(QApplication.font().family(), 18, QFont.Bold))
        group0_layout.addWidget(g0_title)

        # Selector de color de acento
        color_row = QHBoxLayout()
        color_row.addWidget(QLabel("Color de Acento:"))
        self.color_preview = QLabel()
        self.color_preview.setFixedSize(40, 40)
        self.color_preview.setStyleSheet(f"background-color: {COLOR_PRIMARY}; border-radius: 8px; border: 2px solid {COLOR_BORDER};")
        color_row.addWidget(self.color_preview)
        
        btn_color = AnimatedButton("Cambiar Color")
        btn_color.setToolTip("Seleccionar un nuevo color corporativo para la interfaz")
        # Botón en gris como "Borrar Histórico"
        if self.theme != "dark":
            btn_color.setStyleSheet("background-color: #8E8E93; color: white !important; padding: 6px 12px; min-height: 28px; font-size: 13px;")
        else:
            btn_color.setStyleSheet("background-color: #636366; color: white !important; padding: 6px 12px; min-height: 28px; font-size: 13px;")
        btn_color.setMinimumWidth(120)
        btn_color.setMaximumWidth(160)
        btn_color.clicked.connect(self.select_accent_color)
        color_row.addWidget(btn_color)
        
        btn_reset_color = AnimatedButton("Restaurar")
        btn_reset_color.setToolTip("Restaurar el color corporativo original (#A0BF6E)")
        # Botón en gris como "Borrar Histórico"
        if self.theme != "dark":
            btn_reset_color.setStyleSheet("background-color: #8E8E93; color: white !important; padding: 6px 12px; min-height: 28px; font-size: 13px;")
        else:
            btn_reset_color.setStyleSheet("background-color: #636366; color: white !important; padding: 6px 12px; min-height: 28px; font-size: 13px;")
        btn_reset_color.setMinimumWidth(120)
        btn_reset_color.setMaximumWidth(160)
        btn_reset_color.clicked.connect(self.reset_accent_color)
        color_row.addWidget(btn_reset_color)
        color_row.addStretch()
        group0_layout.addLayout(color_row)
        
        info_color = QLabel("El color de acento se aplica a botones, elementos seleccionados y resaltes.")
        info_color.setWordWrap(True)
        info_color.setStyleSheet(f"color: {COLOR_SECONDARY_TEXT};")
        group0_layout.addWidget(info_color)
        
        group0_layout.addSpacing(10)  # Reducido de 15 a 10
        
        # Botones en layout horizontal con ancho limitado y tamaño más compacto
        buttons_row = QHBoxLayout()
        buttons_row.setSpacing(12)
        
        # Botón para generar plantilla Excel
        btn_template = AnimatedButton("📋 Generar Plantilla Excel")
        btn_template.setToolTip("Generar un archivo Excel de ejemplo con la estructura correcta")
        btn_template.setProperty("configButton", "true")
        btn_template.setStyleSheet("padding: 8px 20px; min-height: 32px; max-width: 220px; font-size: 13px;")
        btn_template.clicked.connect(self.generate_excel_template)
        buttons_row.addWidget(btn_template)
        
        # Botón para comprimir logs antiguos
        btn_compress_logs = AnimatedButton("🗜️ Comprimir Logs")
        btn_compress_logs.setToolTip("Comprimir logs y XMLs más antiguos de 30 días para ahorrar espacio")
        btn_compress_logs.setProperty("configButton", "true")
        btn_compress_logs.setStyleSheet("padding: 8px 20px; min-height: 32px; max-width: 180px; font-size: 13px;")
        btn_compress_logs.clicked.connect(self.compress_old_logs)
        buttons_row.addWidget(btn_compress_logs)
        
        buttons_row.addStretch()
        group0_layout.addLayout(buttons_row)
        
        group0_layout.addSpacing(10)  # Reducido de 15 a 10
        
        # Modo offline y botón procesar en una fila
        offline_row = QHBoxLayout()
        offline_row.setSpacing(15)
        
        # Checkbox para modo offline
        self.chk_offline_mode = QCheckBox("Modo Offline (Cola de Envíos)")
        self.chk_offline_mode.setToolTip("Si no hay conexión, guardar facturas en cola para enviar después")
        saved_offline = self.settings.value("offline_mode", "0") == "1"
        self.chk_offline_mode.setChecked(saved_offline)
        self.chk_offline_mode.toggled.connect(self.toggle_offline_mode)
        offline_row.addWidget(self.chk_offline_mode)
        
        offline_row.addStretch()
        
        # Botón para procesar cola offline (a la derecha) - más compacto
        btn_process_queue = AnimatedButton("📤 Procesar Cola Offline")
        btn_process_queue.setToolTip("Intentar enviar facturas pendientes de la cola offline")
        btn_process_queue.setProperty("configButton", "true")
        btn_process_queue.setStyleSheet("padding: 8px 20px; min-height: 32px; max-width: 200px; font-size: 13px;")
        btn_process_queue.clicked.connect(self.process_offline_queue)
        offline_row.addWidget(btn_process_queue)
        
        group0_layout.addLayout(offline_row)
        
        layout.addWidget(group0_frame)

        # --- Grupo 1: API ---
        group1_frame = QFrame()
        group1_frame.setProperty("class", "ConfigGroup")
        apply_shadow(group1_frame, blur=20, offset_y=3)
        group1_layout = QVBoxLayout(group1_frame)
        group1_layout.setSpacing(15)

        g1_title = QLabel("Conexión API")
        g1_title.setFont(QFont(QApplication.font().family(), 18, QFont.Bold)) # Usa fuente global
        group1_layout.addWidget(g1_title)

        btn_config = AnimatedButton("⚙️ Configurar Parámetros API")
        btn_config.setToolTip("Configurar URL, Token y Usuario de la API de Facturantia")
        btn_config.setStyleSheet("padding: 8px 20px; min-height: 32px; font-size: 13px;")
        btn_config.clicked.connect(self.open_config_dialog)
        group1_layout.addWidget(btn_config)
        
        group1_layout.addSpacing(8)
        
        # Botón para consultar certificados
        btn_certificates = AnimatedButton("🔐 Consultar Certificados Digitales")
        btn_certificates.setToolTip("Consultar información sobre los certificados digitales y sus fechas de caducidad")
        btn_certificates.setStyleSheet("padding: 8px 20px; min-height: 32px; font-size: 13px;")
        btn_certificates.clicked.connect(self.check_certificates)
        group1_layout.addWidget(btn_certificates)

        info = QLabel("Configura la URL, Token y Usuario para la conexión con Facturantia.")
        info.setWordWrap(True)
        info.setStyleSheet(f"color: {COLOR_SECONDARY_TEXT};")
        group1_layout.addWidget(info)
        layout.addWidget(group1_frame)

        # --- Grupo 2: Gestión de Datos --- # Modificado título
        group2_frame = QFrame()
        group2_frame.setProperty("class", "ConfigGroup")
        apply_shadow(group2_frame, blur=20, offset_y=3)
        group2_layout = QVBoxLayout(group2_frame)
        group2_layout.setSpacing(15)

        sec_title = QLabel("Usuarios e Historial") # Modificado título
        sec_title.setFont(QFont(QApplication.font().family(), 18, QFont.Bold)) # Usa fuente global
        group2_layout.addWidget(sec_title)

        group2_layout.addSpacing(6)

        # Tabla de usuarios eliminada - los usuarios se gestionan directamente desde users.json
        users_title = QLabel("Gestión de Usuarios")
        users_title.setFont(QFont(QApplication.font().family(), 16, QFont.Medium)) # Usa fuente global
        group2_layout.addWidget(users_title)

        group2_layout.addSpacing(8)

        # Botonera Usuarios - botones más compactos
        user_row = QHBoxLayout()
        user_row.setSpacing(8) # Espacio entre botones
        btn_add = AnimatedButton("➕ Añadir usuario")
        btn_add.setToolTip("Añadir un nuevo usuario al sistema")
        btn_reset = AnimatedButton("🔑 Cambiar contraseña")
        btn_reset.setToolTip("Cambiar la contraseña de un usuario existente")
        btn_del = AnimatedButton("🗑️ Eliminar usuario")
        btn_del.setToolTip("Eliminar un usuario del sistema")
        # Botones aún más compactos para no tapar la tabla
        for btn in [btn_add, btn_reset, btn_del]:
            btn.setStyleSheet("color: white !important; padding: 6px 12px; min-height: 28px; font-size: 13px;")
            btn.setMinimumWidth(120)
            btn.setMaximumWidth(160)
        user_row.addWidget(btn_add)
        user_row.addWidget(btn_reset)
        user_row.addWidget(btn_del)
        user_row.addStretch()
        group2_layout.addLayout(user_row)

        group2_layout.addSpacing(25) # Separador visual más grande

        # --- [AÑADIDO BORRAR HISTORIAL] ---
        history_title = QLabel("Gestión de Historial")
        history_title.setFont(QFont(QApplication.font().family(), 16, QFont.Medium)) # Usa fuente global
        group2_layout.addWidget(history_title)

        group2_layout.addSpacing(4)

        history_row = QHBoxLayout()
        btn_clear_history = AnimatedButton("🧹 Borrar Histórico")
        # Guardar referencia para actualizar cuando cambie el tema
        self.btn_clear_history = btn_clear_history
        # Botón en gris manteniendo la sombra (AnimatedButton ya tiene sombra)
        # Usar color gris en lugar de naranja
        btn_clear_history.setStyleSheet("background-color: #8E8E93; color: white !important; padding: 6px 12px; min-height: 28px; font-size: 13px;")
        # Mismo tamaño que los otros botones
        btn_clear_history.setMinimumWidth(120)
        btn_clear_history.setMaximumWidth(160)
        history_row.addWidget(btn_clear_history)
        history_row.addStretch()
        group2_layout.addLayout(history_row)
        # --- [FIN AÑADIDO BORRAR HISTORIAL] ---

        layout.addWidget(group2_frame) # Añadir el grupo 2 al layout principal

        # Conectar señales
        btn_add.clicked.connect(self.cfg_add_user)
        btn_reset.clicked.connect(self.cfg_reset_password)
        btn_del.clicked.connect(self.cfg_delete_user)
        btn_clear_history.clicked.connect(self.clear_history_confirmation) # <-- Conectar señal nueva

        layout.addStretch()
        return page

    # --- Gestión de usuarios ---
    def _users_file(self):
        return USERS_PATH

    def _read_users(self):
        path = self._users_file()
        if not os.path.exists(path):
            return []
        try:
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)
            return data.get("users", [])
        except Exception:
            return []

    def _write_users(self, users_list):
        data = {"users": users_list}
        with open(self._users_file(), "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, ensure_ascii=False)


    def cfg_add_user(self):
        username, ok = QInputDialog.getText(self, "Nuevo usuario", "Usuario:")
        if not ok or not username.strip():
            return
        password, ok = QInputDialog.getText(self, "Nuevo usuario", "Contraseña:", QLineEdit.Password)
        if not ok or not password:
            return
        users = self._read_users()
        if any(u.get("username") == username for u in users):
            QMessageBox.warning(self, "Usuarios", f"El usuario '{username}' ya existe.")
            return
        users.append({"username": username, "password_hash": hashlib.sha256(password.encode("utf-8")).hexdigest()})
        self._write_users(users)
        self.show_toast(f"✅ Usuario '{username}' creado")

    def cfg_selected_username(self):
        """Solicita al usuario que ingrese el nombre de usuario."""
        users = self._read_users()
        if not users:
            QMessageBox.information(self, "Usuarios", "No hay usuarios en el sistema.")
            return None
        
        # Crear lista de nombres de usuario
        usernames = [u.get("username", "") for u in users if u.get("username")]
        if not usernames:
            QMessageBox.information(self, "Usuarios", "No hay usuarios en el sistema.")
            return None
        
        # Si solo hay un usuario, usarlo directamente
        if len(usernames) == 1:
            return usernames[0]
        
        # Si hay varios, mostrar diálogo para seleccionar
        username, ok = QInputDialog.getItem(
            self,
            "Seleccionar Usuario",
            "Selecciona el usuario:",
            usernames,
            0,
            False
        )
        return username if ok and username else None

    def cfg_reset_password(self):
        username = self.cfg_selected_username()
        if not username:
            return
        password, ok = QInputDialog.getText(
            self, "Cambiar contraseña", f"Nueva contraseña para '{username}':", QLineEdit.Password
        )
        if not ok or not password:
            return
        users = self._read_users()
        for u in users:
            if u.get("username") == username:
                u["password_hash"] = hashlib.sha256(password.encode("utf-8")).hexdigest()
                break
        self._write_users(users)
        self.show_toast("✅ Contraseña actualizada")

    def cfg_delete_user(self):
        username = self.cfg_selected_username()
        if not username:
            return
        if username.lower() == "admin":
            QMessageBox.warning(self, "Usuarios", "No se permite eliminar el usuario 'admin'.")
            return
        ret = QMessageBox.question(self, "Eliminar usuario", f"¿Eliminar el usuario '{username}'?")
        if ret != QMessageBox.Yes:
            return
        users = [u for u in self._read_users() if u.get("username") != username]
        self._write_users(users)
        self.show_toast(f"🗑️ Usuario '{username}' eliminado")

    # --- [NUEVAS FUNCIONES BORRAR HISTORIAL] ---
    def clear_history_confirmation(self):
        reply = QMessageBox.question(self, 'Confirmar Borrado',
                                     "¿Estás seguro de que quieres borrar TODO el historial de envíos?\n"
                                     "Esta acción NO se puede deshacer y pondrá el Dashboard a cero.",
                                     QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                                     QMessageBox.StandardButton.No)

        if reply == QMessageBox.StandardButton.Yes:
            self.clear_history_execute()

    def clear_history_execute(self):
        try:
            conn = sqlite3.connect(DB_PATH)
            cursor = conn.cursor()
            cursor.execute("DELETE FROM envios") # Borra todos los registros
            conn.commit()
            # Opcional: Limpiar espacio no usado en la DB
            cursor.execute("VACUUM")
            conn.commit()
            conn.close()

            self.show_toast("✅ Historial de envíos borrado.")
            # Recargar la tabla del histórico (que ahora estará vacía)
            self.load_history()
            # update_dashboard_stats() ya se llama dentro de load_history()

        except Exception as e:
            self.show_error(f"Error al borrar el historial: {e}")
            print(f"Error borrando historial: {e}")
    # --- [FIN NUEVAS FUNCIONES] ---


    # --- Navegación/tema ---
    def change_page(self, index):
        # Animación de transición tipo iOS (fade out/in)
        if index == self.content_stack.currentIndex():
            return
        # Tomar snapshot de la vista actual
        try:
            prev_pix = self.content_stack.grab()
        except Exception:
            prev_pix = None
        # Cambiar inmediatamente al destino para preparar fade-in del nuevo contenido
        self.content_stack.setCurrentIndex(index)
        # Si no hay snapshot, no animamos
        if prev_pix is None:
            return
        # Crear overlay con el snapshot anterior y desvanecerlo
        overlay = QLabel(self.content_stack)
        overlay.setPixmap(prev_pix)
        overlay.setGeometry(self.content_stack.rect())
        overlay.raise_()
        effect = QGraphicsOpacityEffect(overlay)
        overlay.setGraphicsEffect(effect)
        effect.setOpacity(1.0)
        anim = QPropertyAnimation(effect, b"opacity", self)
        anim.setDuration(220)
        anim.setStartValue(1.0)
        anim.setEndValue(0.0)
        anim.setEasingCurve(QEasingCurve.OutCubic)
        # Mantener referencias hasta finalizar
        if not hasattr(self, "_page_overlays"):
            self._page_overlays = []
        self._page_overlays.append((overlay, anim))
        def _cleanup():
            overlay.deleteLater()
            try:
                self._page_overlays.remove((overlay, anim))
            except Exception:
                pass
        anim.finished.connect(_cleanup)
        overlay.show()
        anim.start()
        if index == 0: # Índice del Dashboard
            self.update_dashboard_stats() # Recarga stats de tarjetas y errores
            self.populate_dashboard_filters() # Recarga el combo de empresas
        elif index == 3: # Índice del Histórico
            self.load_history() # Recarga el histórico al cambiar a esa página

    def toggle_theme(self, checked):
        self.theme = "dark" if checked else "light"
        self.setProperty("theme", self.theme)
        app = QApplication.instance()
        style_sheet_content = self._get_themed_stylesheet()
        if style_sheet_content:
            app.setStyleSheet(style_sheet_content)

        # Re-aplicar estilos (necesario para que todo se repinte)
        # Usamos un temporizador para asegurar que se haga después del evento actual
        QTimer.singleShot(0, self._refresh_styles)

    # [CORREGIDO] _refresh_styles ahora es más simple y evita el TypeError
    def _refresh_styles(self):
        app = QApplication.instance()
        for w in app.allWidgets():
            try:
                # Estas dos líneas suelen ser suficientes para refrescar el estilo
                w.style().unpolish(w)
                w.style().polish(w)
                # Eliminamos w.update() ya que causaba TypeErrors en varios widgets
            except RuntimeError:
                # Mantenemos esto por si el widget se elimina mientras iteramos
                pass
            # Eliminamos el bloque 'except TypeError' ya que quitamos la causa
        
        # Actualizar botón "Borrar Histórico" según el tema (gris en ambos modos)
        if hasattr(self, 'btn_clear_history'):
            if self.theme != "dark":
                self.btn_clear_history.setStyleSheet("background-color: #8E8E93; color: white !important; padding: 6px 12px; min-height: 28px; font-size: 13px;")
            else:
                # En modo oscuro también gris, pero más claro
                self.btn_clear_history.setStyleSheet("background-color: #636366; color: white !important; padding: 6px 12px; min-height: 28px; font-size: 13px;")


    def _on_post_action_changed(self, idx):
        self.post_action_mode = self.post_action_combo.itemData(idx) or "MARK"
        self.append_log(f"Acción post-envío seleccionada: {self.post_action_mode}")

    # --- Drag & drop Excel ---
    def dragEnterEvent(self, event):
        if event.mimeData().hasUrls():
            event.accept()
        else:
            event.ignore()

    def dropEvent(self, event):
        file_path = event.mimeData().urls()[0].toLocalFile()
        if file_path.lower().endswith((".xlsx", ".xlsm", ".xls")):
            self.process_dropped_excel(file_path)
        event.accept()

    def process_dropped_excel(self, path):
        self.select_excel(path)

    # --- Limpieza integral de la página "Cargar Excel" ---
    def clear_excel_table(self):
        if hasattr(self, "table_excel"):
            self.table_excel.setRowCount(0)
        if hasattr(self, "validation_label"):
            self.validation_label.clear()
        if hasattr(self, "stepper"):
            self.stepper.set_step(0)
        if hasattr(self, "btn_send"):
            self.btn_send.setEnabled(False)
        self.current_excel_path = None
        self.loaded_invoice_count = 0
        self._update_send_badge()

        # Limpiar también la página de envío
        self.clear_send_page()

    def select_excel(self, path=None):
        if not path:
            path, _ = QFileDialog.getOpenFileName(
                self,
                "Seleccionar Excel",
                "",
                "Hojas de cálculo (*.xlsx *.XLSX *.xlsm *.XLSM *.xls *.XLS);;Todos los archivos (*.*)",
            )
        if not path:
            return

        self.clear_excel_table()

        self.current_excel_path = path
        self.stepper.set_step(0)
        self.append_log(f"📁 Excel seleccionado: {path}")
        if self.validate_excel(path):
            self.stepper.set_step(2)
            self.btn_send.setEnabled(True)
            self.validation_label.setText("✅ Archivo validado correctamente")
            self.validation_label.setStyleSheet(f"color: {COLOR_SUCCESS}; font-weight: bold;")
        else:
            self.stepper.set_step(1)
            self.btn_send.setEnabled(False)

            def _fmt_err(e):
                if isinstance(e, str):
                    return e
                if isinstance(e, (list, tuple)):
                    # FIX: 'y' → 'and' para Python
                    return " - ".join(str(x) for x in e if x is not None and str(x).strip() != "")
                try:
                    return str(e)
                except Exception:
                    return repr(e)

            errs_str = [_fmt_err(e) for e in self.validation_errors]
            error_msg = "❌ Errores de validación:\n" + "\n".join(errs_str)
            self.validation_label.setText(error_msg)
            self.validation_label.setStyleSheet(f"color: {COLOR_ERROR};")
            return

        # --- [INICIO DE LA CORRECCIÓN] ---
        # Llamar a adapt_from_macro con UN solo argumento, como en tus archivos.
        try:
            # --- [MODIFICADO] Capturar y almacenar los 6 dataframes ---
            (
                df_factura, df_conceptos, df_forma_pago, df_txt,
                self.df_factura_historico, self.df_conceptos_historico
            ) = macro_adapter.adapt_from_macro(path)
        # --- [FIN MODIFICADO] ---
        except Exception as e:
            self.append_log(f"❌ Error leyendo Excel (Macro): {e}")
            self.show_error(f"Error procesando el Excel: {e}")
            return

        # --- Normalizamos NumFactura en ambos DF y usamos el mismo ID para todo ---
        if "NumFactura" in df_factura.columns:
            df_factura["__id_norm__"] = df_factura["NumFactura"].map(_normalize_invoice_id)
        else:
            df_factura["__id_norm__"] = ""

        if "NumFactura" in df_conceptos.columns:
            df_conceptos["__id_norm__"] = df_conceptos["NumFactura"].map(_normalize_invoice_id)
        else:
            df_conceptos["__id_norm__"] = ""

        self.table_excel.setRowCount(0)
        for i, row in df_factura.iterrows():
            self.table_excel.insertRow(i)
            has_error = i in [err[0] for err in self.validation_errors if isinstance(err, tuple)]

            # [MODIFICADO] Color de error más sutil
            err_color = QColor(COLOR_ERROR)
            err_color.setAlpha(40) # 40/255 de opacidad
            bg_color = err_color if has_error else QColor(COLOR_CARD)

            # Color de texto normal
            text_color = QColor(COLOR_TEXT)
            if self.theme == "dark":
                bg_color = err_color if has_error else QColor(COLOR_DARK_CARD)
                text_color = QColor(COLOR_DARK_TEXT)

            inv_id = row.get("__id_norm__", "")
            item_factura = QTableWidgetItem(inv_id)
            item_factura.setBackground(bg_color)
            item_factura.setForeground(text_color)
            self.table_excel.setItem(i, 0, item_factura)

            item_empresa = QTableWidgetItem(str(row.get("empresa_emisora", "")))
            item_empresa.setBackground(bg_color)
            item_empresa.setForeground(text_color)
            self.table_excel.setItem(i, 1, item_empresa)

            # Cliente
            item_cliente = QTableWidgetItem(str(row.get("cliente_nombre", "")))
            item_cliente.setBackground(bg_color)
            item_cliente.setForeground(text_color)
            self.table_excel.setItem(i, 2, item_cliente)

            # Cálculos de importes
            base_sum = 0.0
            iva_sum = 0.0
            ret_sum = 0.0
            if inv_id:
                # Añadir filtro por empresa emisora
                em_act = row.get("empresa_emisora", "")
                conceptos_factura = df_conceptos[
                    (df_conceptos["__id_norm__"] == inv_id) &
                    (df_conceptos["empresa_emisora"] == em_act)
                ]
                if not conceptos_factura.empty:
                    base_sum = float(conceptos_factura["base_unidad"].sum() or 0.0)

                    # Calcular IVA
                    iva_sum = (conceptos_factura["base_unidad"] * (conceptos_factura["porcentaje"] / 100.0)).sum() if not conceptos_factura["base_unidad"].empty else 0.0


                    # Calcular Retención
                    ret_sum = (conceptos_factura["base_unidad"] * (conceptos_factura["porcentaje_retenido"] / 100.0)).sum() if not conceptos_factura["base_unidad"].empty else 0.0

            total_sum = base_sum + iva_sum - ret_sum

            # Base Imponible
            item_base = QTableWidgetItem(format_eur(base_sum))
            item_base.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            item_base.setBackground(bg_color)
            item_base.setForeground(text_color)
            self.table_excel.setItem(i, 3, item_base)

            # Cantidad IVA
            item_iva = QTableWidgetItem(format_eur(iva_sum))
            item_iva.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            item_iva.setBackground(bg_color)
            item_iva.setForeground(text_color)
            self.table_excel.setItem(i, 4, item_iva)

            # Retención
            item_ret = QTableWidgetItem(format_eur(ret_sum))
            item_ret.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            item_ret.setBackground(bg_color)
            item_ret.setForeground(text_color)
            self.table_excel.setItem(i, 5, item_ret)

            # Importe Total
            item_total = QTableWidgetItem(format_eur(total_sum))
            item_total.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            item_total.setBackground(bg_color)
            item_total.setForeground(text_color)
            self.table_excel.setItem(i, 6, item_total)

            fecha = row.get("fecha_emision", "")
            fecha_str = pd.to_datetime(fecha).strftime("%d/%m/%Y") if pd.notna(fecha) else ""
            item_fecha = QTableWidgetItem(fecha_str)
            item_fecha.setBackground(bg_color)
            item_fecha.setForeground(text_color)
            self.table_excel.setItem(i, 7, item_fecha)

        # Reajuste final de columnas después de cargar
        for col in [0, 3, 4, 5, 6, 7]:
            self.table_excel.resizeColumnToContents(col)

        # --- [NUEVO] Poblar la tabla de previsualización en la página de envío ---
        try:
            self.table_preview.setRowCount(0)
            for i in range(self.table_excel.rowCount()):
                row_idx = self.table_preview.rowCount()
                self.table_preview.insertRow(row_idx)
                # Copiamos los datos relevantes de la tabla excel a la de previsualización
                # Factura (col 0), Empresa (col 1), Cliente (col 2), Total (col 6), Fecha (col 7)
                self.table_preview.setItem(row_idx, 0, QTableWidgetItem(self.table_excel.item(i, 0).text()))
                self.table_preview.setItem(row_idx, 1, QTableWidgetItem(self.table_excel.item(i, 1).text()))
                self.table_preview.setItem(row_idx, 2, QTableWidgetItem(self.table_excel.item(i, 2).text()))

                item_total = QTableWidgetItem(self.table_excel.item(i, 6).text())
                item_total.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                self.table_preview.setItem(row_idx, 3, item_total)

                self.table_preview.setItem(row_idx, 4, QTableWidgetItem(self.table_excel.item(i, 7).text()))

            # Ajustar columnas de la tabla de previsualización
            for col_prev in [0, 3, 4]:
                self.table_preview.resizeColumnToContents(col_prev)
        except Exception as e:
            print(f"Error al poblar la tabla de previsualización: {e}")
        # --- [FIN NUEVO] ---

        # KPI: contador y total base
        try:
            count = len(df_factura)
            total_base = float(df_conceptos["base_unidad"].sum() or 0.0)
            msg = f"📥 {count} facturas cargadas · Total base: {format_eur(total_base)}"
            self.loaded_summary.setText(msg)
            self.loaded_summary.setVisible(True)
            self.loaded_invoice_count = count
            self._update_send_badge()
            self.show_toast(msg)
        except Exception:
            pass

        if hasattr(self.worker, "set_excel_path"):
            self.worker.set_excel_path(path)

    def validate_excel(self, path):
        self.validation_errors = []
        
        # --- [INICIO DE LA CORRECCIÓN] ---
        # Llamar a adapt_from_macro con UN solo argumento
        try:
            # --- [MODIFICADO] Capturar los 6 dataframes, aunque no se usen todos aquí ---
            (
                df_factura, df_conceptos, df_forma_pago, df_txt,
                _, _ # Ignoramos los históricos en la validación simple
            ) = macro_adapter.adapt_from_macro(path)
        # --- [FIN MODIFICADO] ---
        except Exception as e:
            self.validation_errors.append(f"Error leyendo archivo (Macro): {str(e)}")
            return False

        required_cols = ["NumFactura", "empresa_emisora", "fecha_emision"]
        missing_cols = [col for col in required_cols if col not in df_factura.columns]
        if missing_cols:
            self.validation_errors.append(f"Faltan columnas: {', '.join(missing_cols)}")
            return False

        bases = df_conceptos.groupby("NumFactura")["base_unidad"].sum().to_dict()
        for i, row in df_factura.reset_index(drop=True).iterrows():
            row_errors = []
            if pd.isna(row.get("NumFactura")) or str(row.get("NumFactura")).strip() == "":
                row_errors.append("Número de factura vacío")
            if pd.isna(row.get("empresa_emisora")) or str(row.get("empresa_emisora")).strip() == "":
                row_errors.append("Empresa emisora vacía")
            if pd.isna(row.get("fecha_emision")):
                row_errors.append("Fecha de emisión vacía")
            if bases.get(row.get("NumFactura"), 0) <= 0:
                row_errors.append("Importe inválido (base_unidad <= 0)")
            if row_errors:
                self.validation_errors.append((i, f"Fila {i+2}: {', '.join(row_errors)}"))
        return len(self.validation_errors) == 0

    def send_facturas(self):
        if not self.current_excel_path:
            self.show_toast("❌ No hay archivo Excel cargado")
            return

        # --- [NUEVO] Ocultar previsualización y mostrar resultados ---
        self.table_preview.setVisible(False)
        if hasattr(self, "preview_card"):
            self.preview_card.setVisible(False)
        if hasattr(self, "preview_title"):
            self.preview_title.setVisible(False)
        self.results_group.setVisible(True)
        # --- [FIN NUEVO] ---

        # Limpiar tabla de envío actual y pasos
        self.table_envio.setRowCount(0)
        self.send_stepper.set_step(0)

        # Progreso indeterminado mientras dura el envío (más honesto)
        self.progress.setRange(0, 0)

        self.sending_in_progress = True
        self.btn_send.setEnabled(False)
        self.btn_download_pdfs.setEnabled(False)

        # Evitar reseguir resultados antiguos: respaldar summary.json si existe
        try:
            summary_path = os.path.join("responses", "summary.json")
            if os.path.exists(summary_path):
                ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                backup = os.path.join("responses", f"summary_{ts}.json")
                os.replace(summary_path, backup)
        except Exception:
            pass

        # EXISTENTE: acción post-macro
        os.environ["POST_MACRO_ACTION"] = self.post_action_mode
        
        # [NUEVO] Configurar modo offline
        offline_enabled = self.settings.value("offline_mode", "0") == "1"
        os.environ["USE_OFFLINE_QUEUE"] = "1" if offline_enabled else "0"

        # EXISTENTE: opciones en entorno para compatibilidad
        os.environ["AUTO_DOWNLOAD_PDFS"] = "1" if getattr(self, "chk_auto_download", None) and self.chk_auto_download.isChecked() else "0"
        os.environ["PDF_DEST_DIR"] = (
            self.txt_pdf_dest.text()
            if hasattr(self, "txt_pdf_dest")
            else r"C:\Users\administracionaba\Desktop\FACTURANTIA\FactuNabo EXE\Facturas PDF"
        )
        os.environ["PDF_BROWSER"] = self.browser_code
        os.environ["PDF_HEADLESS"] = "1"

        # --- [FIX] Crear un nuevo worker y thread para cada envío ---
        self.thread = QThread()
        self.worker = Worker()

        # Mover worker al thread
        self.worker.moveToThread(self.thread)

        # Conectar señales
        self.worker.log_signal.connect(self.append_log, Qt.QueuedConnection)
        self.worker.finished.connect(self.on_finished, Qt.QueuedConnection)
        self.thread.started.connect(self.worker.process)

        # Configurar el worker con los datos actuales
        self.worker.set_excel_path(self.current_excel_path)

        # --- [NUEVO] Pasar los dataframes históricos al worker ---
        self.worker.set_historical_data(
            self.df_factura_historico,
            self.df_conceptos_historico
        )
        # --- [FIN NUEVO] ---

        self.worker.set_post_macro_action(self.post_action_mode)
        self.worker.set_download_options(
            auto=(self.chk_auto_download.isChecked() if hasattr(self, "chk_auto_download") else False),
            dest=(self.txt_pdf_dest.text() if hasattr(self, "txt_pdf_dest") else ""),
            browser=self.browser_code,
            headless=True,
        )

        QTimer.singleShot(500, lambda: self.send_stepper.set_step(1))
        self.thread.start()

    # --- FIX: Slot GUI para limpiar hilo de descargas de forma segura ---
    def on_downloads_done_gui(self):
        """Siempre ejecutado en el hilo de la GUI."""
        try:
            self.progress.setRange(0, 100)
            self.progress.setValue(100)
            self.show_toast("✅ Descarga de PDFs terminada")
        finally:
            # Cierre/limpieza del hilo de descargas SOLO desde la GUI
            try:
                if getattr(self, "dthread", None) and self.dthread.isRunning():
                    self.dthread.quit()
                    self.dthread.wait(3000)
            except Exception:
                pass
            try:
                if getattr(self, "dworker", None):
                    self.dworker.deleteLater()
                if getattr(self, "dthread", None):
                    self.dthread.deleteLater()
            except Exception:
                pass
            self.dthread = None
            self.dworker = None
            if hasattr(self, "btn_download_pdfs") and self.btn_download_pdfs:
                self.btn_download_pdfs.setEnabled(True)
            # Sincronizar rutas locales con el historial
            try:
                summary_path = os.path.join("responses", "summary.json")
                if os.path.exists(summary_path):
                    with open(summary_path, "r", encoding="utf-8") as f:
                        data = json.load(f)
                    if isinstance(data, list):
                        self._update_pdf_paths_in_history(data)
            except Exception as sync_err:
                print(f"Error sincronizando rutas PDF tras descarga: {sync_err}")

    def download_pdfs_clicked(self):
        # protección: necesita summary.json
        summary_path = os.path.join("responses", "summary.json")
        if not os.path.exists(summary_path):
            self.show_toast("⚠️ No hay resumen disponible. Primero ejecuta un envío.")
            return

        # Barra indeterminada mientras descarga
        self.progress.setRange(0, 0)
        self.append_log("📥 Iniciando descarga manual de PDFs...")
        self.btn_download_pdfs.setEnabled(False)

        # Hilo temporal con un Worker “solo descarga”
        self.dthread = QThread(self)
        self.dworker = Worker()
        try:
            self.dworker.set_download_options(
                auto=True,  # habilita opciones internas; se llama download_pdfs() directamente
                dest=(self.txt_pdf_dest.text() if hasattr(self, "txt_pdf_dest") else ""),
                browser=self.browser_code,
                headless=True,
            )
        except Exception:
            pass

        self.dworker.moveToThread(self.dthread)
        # FIX(Queued): logs del worker de descargas también encolados a GUI
        self.dworker.log_signal.connect(self.append_log, Qt.QueuedConnection)
        self.dthread.started.connect(self.dworker.download_pdfs)

        # --- FIX(Queued): conectar señal a GUI con entrega encolada
        if hasattr(self.dworker, "downloads_done"):
            self.dworker.downloads_done.connect(self.on_downloads_done_gui, Qt.QueuedConnection)
        else:
            self.show_error("El Worker no expone la señal 'downloads_done'. Actualiza worker.py.")

        self.dthread.start()

    # ######################################################################
    # INICIO DEL BLOQUE AÑADIDO (FUNCIONES DE FILTRADO QUE FALTABAN)
    # ######################################################################

    def _apply_envio_filters(self):
        """Función central que aplica BÚSQUEDA y FILTROS de estado a la tabla de envío."""
        if not hasattr(self, 'table_envio'):
            return

        # 1. Obtener el texto de búsqueda
        search_text = ""
        if hasattr(self, 'search_bar'):
            search_text = self.search_bar.text().strip().lower()

        # 2. Obtener los filtros de estado activos
        active_statuses = []
        if hasattr(self, 'filters'):
            for status, btn in self.filters.items():
                if btn.isChecked():
                    # Usamos el nombre del botón (ÉXITO, DUPLICADO, ERROR)
                    active_statuses.append(status.lower())

        # 3. Iterar y aplicar
        for row in range(self.table_envio.rowCount()):
            # Obtener datos de la fila
            item_factura = self.table_envio.item(row, 0)
            item_empresa = self.table_envio.item(row, 1)
            widget_estado = self.table_envio.cellWidget(row, 2) # Es un StatusChip

            text_factura = item_factura.text().lower() if item_factura else ""
            text_empresa = item_empresa.text().lower() if item_empresa else ""
            
            # El StatusChip tiene el texto (ÉXITO, DUPLICADO, NABO!/ERROR)
            text_estado = ""
            if isinstance(widget_estado, StatusChip):
                text_estado = widget_estado.text().lower()
            elif isinstance(widget_estado, QLabel): # Fallback
                text_estado = widget_estado.text().lower()
            
            # Mapear "NABO!" a "error" si es necesario
            if "nabo" in text_estado:
                text_estado = "error"

            # --- Aplicar Lógica ---
            
            # 1. Comprobar filtro de texto
            match_text = True # Asumir que coincide si no hay texto
            if search_text:
                match_text = (search_text in text_factura) or (search_text in text_empresa)

            # 2. Comprobar filtro de estado
            match_status = True # Asumir que coincide si no hay filtros activos
            if active_statuses:
                match_status = (text_estado in active_statuses)

            # 3. Decisión final
            self.table_envio.setRowHidden(row, not (match_text and match_status))

    def apply_search(self, text):
        """Slot para la barra de búsqueda. Llama al filtro principal."""
        self._apply_envio_filters()

    def apply_filter(self):
        """Slot para los botones de filtro. Llama al filtro principal."""
        self._apply_envio_filters()
        
    # ######################################################################
    # FIN DEL BLOQUE AÑADIDO
    # ######################################################################

    # --- FIX(GUI-guard): asegurar que append_log corre en el hilo GUI
    def append_log(self, msg):
        # Si esta función entra desde un hilo que no es el de la GUI, reencola de forma segura
        if QThread.currentThread() != QApplication.instance().thread():
            # Usamos Qt.QueuedConnection para encolar el mensaje de forma segura
            self.worker.log_signal.emit(str(msg))
            return

        # Actualización del Stepper si el log lo indica
        msg_str = str(msg)
        if "Generando PDF" in msg_str or "descarga" in msg_str.lower():
            if self.sending_in_progress and self.send_stepper.current_step < 2:
                self.send_stepper.set_step(2)

        # Asegurarse de que log_area existe antes de usarla
        if hasattr(self, 'log_area') and self.log_area:
             self.log_area.append(msg_str)
             self.log_area.verticalScrollBar().setValue(self.log_area.verticalScrollBar().maximum())

        self.show_toast(msg_str)

        # Habilita el botón de PDFs en cuanto exista el summary
        try:
            if os.path.exists(os.path.join("responses", "summary.json")):
                 if hasattr(self, 'btn_download_pdfs') and self.btn_download_pdfs:
                      self.btn_download_pdfs.setEnabled(True)
        except Exception:
            pass

        # Si el proceso principal no está en marcha, no intentes actualizar la tabla de envío
        if not self.sending_in_progress or not hasattr(self, 'table_envio'):
            return

        # --- [INICIO DE LA MODIFICACIÓN] ---
        # La lógica de guardado en BBDD se ha movido a on_finished.
        # Esta función ahora solo actualiza la tabla de la UI.
        summary_path = os.path.join("responses", "summary.json")
        if not os.path.exists(summary_path):
            return
        try:
            with open(summary_path, "r", encoding="utf-8") as f:
                try:
                    data = json.load(f)
                except json.JSONDecodeError:
                    return
        except Exception:
            return

        if not isinstance(data, list):
             return

        self.table_envio.setRowCount(0)

        for i, item in enumerate(data):
            self.table_envio.insertRow(i)
            if not isinstance(item, dict):
                 continue
            try:
                self.table_envio.setItem(i, 0, QTableWidgetItem(str(item.get("id", ""))))
                self.table_envio.setItem(i, 1, QTableWidgetItem(str(item.get("empresa", ""))))
                self.table_envio.setItem(i, 2, QTableWidgetItem(str(item.get("cliente", ""))))

                item_importe = QTableWidgetItem(format_eur(item.get("importe", 0.0)))
                item_importe.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                self.table_envio.setItem(i, 3, item_importe)

                estado = str(item.get("status", "")).upper()
                if estado == "DUPLICATE": estado = "DUPLICADO"
                elif estado == "ATENCION": estado = "DUPLICADO"
                chip = StatusChip(estado)
                self.table_envio.setCellWidget(i, 4, chip)
                self.table_envio.setItem(i, 5, QTableWidgetItem(str(item.get("details", ""))))
                pdf_url = item.get("pdf_url", "")
                local_path = item.get("pdf_local_path")
                cliente = item.get("cliente")
                importe_val = item.get("importe") or item.get("importe_total") or item.get("total")
                invoice_id = str(item.get("id") or item.get("NumFactura") or item.get("num_factura") or "").strip()
                btn_open = self._make_pdf_button(
                    invoice_id,
                    pdf_url,
                    local_path=local_path,
                    cliente=cliente,
                    importe=format_eur(importe_val) if importe_val else None
                )
                self.table_envio.setCellWidget(i, 6, btn_open)
                # LA LLAMADA A self.save_to_history(item) HA SIDO ELIMINADA DE AQUÍ
            except Exception as e:
                 print(f"Error procesando item {i} del summary para la UI: {e} - Item: {item}")
                 error_item = QTableWidgetItem("Error procesando")
                 error_item.setForeground(QColor(COLOR_ERROR))
                 for col_err in range(self.table_envio.columnCount()):
                      if not self.table_envio.item(i, col_err):
                           self.table_envio.setItem(i, col_err, error_item.clone())
        # --- [FIN DE LA MODIFICACIÓN] ---


    # ######################################################################
    # INICIO DEL BLOQUE CORREGIDO (INDENTACIÓN AÑADIDA)
    # Todas las siguientes funciones ahora están DENTRO de MainWindow
    # ######################################################################

    def on_finished(self, *args, **kwargs):
        """
        Slot llamado cuando el worker termina.
        1. Restaura botones y estado de la UI.
        2. Llama a la función que guarda el resumen en la BBDD.
        """
        try:
            self.sending_in_progress = False
            if hasattr(self, 'progress'):
                self.progress.setRange(0, 100)
                self.progress.setValue(100)
            if hasattr(self, 'send_stepper'):
                self.send_stepper.set_step(3)

            # --- [NUEVO] Guardar el resumen en la BBDD ---
            summary_path = os.path.join("responses", "summary.json")
            if os.path.exists(summary_path):
                try:
                    with open(summary_path, "r", encoding="utf-8") as f:
                        summary_data = json.load(f)
                    if isinstance(summary_data, list):
                        self.save_summary_to_history(summary_data)
                except Exception as e:
                    self.show_error(f"Error al procesar el resumen para el historial: {e}")
            # --- [FIN NUEVO] ---

        except Exception as e:
            print(f"Error en on_finished: {e}")
        finally:
            if hasattr(self, "btn_send"):
                self.btn_send.setEnabled(True)
            if hasattr(self, "btn_download_pdfs"):
                self.btn_download_pdfs.setEnabled(True)
            if hasattr(self, "_overlay"):
                self._overlay.hide()

            # --- [FIX] Limpieza del thread y worker ---
            if self.thread and self.thread.isRunning():
                self.thread.quit()
                self.thread.wait() # Esperar a que termine limpiamente

            # Marcar para eliminación segura
            if self.worker:
                self.worker.deleteLater()
            if self.thread:
                self.thread.deleteLater()

            self.worker = None
            self.thread = None


    def show_error(self, message):
        """Muestra un error de forma centralizada (para conexiones del worker)."""
        try:
            if hasattr(self, "statusbar"):
                self.statusbar.showMessage(str(message), 8000)
            QMessageBox.critical(self, "Error", str(message))
        except Exception:
            # Fallback: imprimir en consola
            print(f"[ERROR] {message}")

    # [NUEVA FUNCIÓN]
    def _parse_and_sum_amount(self, raw_amount):
        """
        Parsea y convierte un importe a float, manejando tanto números como strings
        en formato español ('1.234,56').
        """
        if isinstance(raw_amount, (int, float)):
            return float(raw_amount)

        if not raw_amount or not str(raw_amount).strip():
            return 0.0

        try:
            # Si es un string, aplicamos la normalización
            amount_str = str(raw_amount).replace('.', '').replace(',', '.')
            return float(amount_str)
        except (ValueError, TypeError):
            return 0.0

    # [REDISEÑADO] save_summary_to_history para guardar facturas individuales
    def save_summary_to_history(self, summary_data: list):
        """
        Procesa el resultado de un envío (summary.json) y guarda una entrada
        por cada factura individual en la BBDD.
        """
        if not summary_data:
            return

        records_to_insert = []
        for item in summary_data:
            if not isinstance(item, dict):
                continue

            # Extraer datos de cada factura
            num_factura = item.get("id") or item.get("NumFactura") or "N/A"
            empresa = item.get("empresa") or "Desconocida"
            cliente = item.get("cliente") or ""
            raw_amount = (item.get("amount") or item.get("importe") or
                          item.get("importe_total") or item.get("total") or 0.0)
            importe = self._parse_and_sum_amount(raw_amount)

            status = str(item.get("status", "ERROR")).upper()
            if status in ("OK", "SUCCESS"):
                status = "ÉXITO"
            elif status in ("DUPLICATE", "ATENCION"):
                status = "DUPLICADO"

            detalles = item.get("details", "")
            if isinstance(detalles, dict):
                detalles = json.dumps(detalles, ensure_ascii=False)

            pdf_url = item.get("pdf_url", "")
            if isinstance(pdf_url, dict):
                pdf_url = json.dumps(pdf_url, ensure_ascii=False)

            local_path = item.get("pdf_local_path") or ""
            # Preparar la tupla para la inserción
            records_to_insert.append((
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                num_factura,
                empresa,
                status,
                detalles,
                pdf_url,
                self.current_excel_path or "",
                local_path,
                importe,
                cliente
            ))

        # Insertar todos los registros en una única transacción
        if not records_to_insert:
            return

        try:
            conn = sqlite3.connect(DB_PATH)
            cursor = conn.cursor()
            cursor.executemany(
                """INSERT INTO envios (
                    fecha_envio, num_factura, empresa, estado,
                    detalles, pdf_url, excel_path, pdf_local_path,
                    importe, cliente
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                records_to_insert
            )
            conn.commit()
            conn.close()
            self.show_toast(f"✅ {len(records_to_insert)} facturas guardadas en el historial.")
            self._update_pdf_paths_in_history(summary_data)
        except Exception as e:
            print(f"Error guardando resumen en BBDD: {e}")
            self.show_error(f"Error al guardar en el historial: {e}")

    def _update_pdf_paths_in_history(self, summary_data: list):
        """Sincroniza las rutas locales de PDF en la base de datos a partir del summary."""
        if not summary_data:
            return
        updates = []
        for item in summary_data:
            if not isinstance(item, dict):
                continue
            local_path = item.get("pdf_local_path")
            if not local_path:
                continue
            num_factura = item.get("id") or item.get("NumFactura") or item.get("num_factura")
            empresa = item.get("empresa")
            if not num_factura or not empresa:
                continue
            updates.append((local_path, str(num_factura), str(empresa)))

        if not updates:
            return
        try:
            conn = sqlite3.connect(DB_PATH)
            cursor = conn.cursor()
            cursor.executemany(
                """
                UPDATE envios
                SET pdf_local_path = ?
                WHERE num_factura = ? AND empresa = ?
                """,
                updates
            )
            conn.commit()
            conn.close()
        except Exception as e:
            print(f"Error actualizando rutas de PDF en historial: {e}")

    # ######################################################################
    # INICIO DEL BLOQUE CORREGIDO (AÑADIDA INDENTACIÓN)
    # ######################################################################

    def load_history(self, apply_filters=True):
        try:
            conn = sqlite3.connect(DB_PATH)
            cursor = conn.cursor()
            
            # [NUEVO] Construir query con filtros
            query = "SELECT id, fecha_envio, num_factura, empresa, estado, detalles, pdf_url, pdf_local_path, importe, cliente FROM envios WHERE 1=1"
            params = []
            
            if apply_filters and hasattr(self, 'history_filter_empresa'):
                empresa = self.history_filter_empresa.currentText()
                if empresa != "Todas las Empresas":
                    query += " AND empresa = ?"
                    params.append(empresa)
                
                estado = self.history_filter_estado.currentText()
                if estado != "Todos":
                    query += " AND estado = ?"
                    params.append(estado)
                
                periodo = self.history_filter_periodo.currentText()
                if periodo != "Todos":
                    now = datetime.now()
                    if periodo == "1º Trimestre":
                        query += " AND strftime('%m', fecha_envio) IN ('01', '02', '03')"
                    elif periodo == "2º Trimestre":
                        query += " AND strftime('%m', fecha_envio) IN ('04', '05', '06')"
                    elif periodo == "3º Trimestre":
                        query += " AND strftime('%m', fecha_envio) IN ('07', '08', '09')"
                    elif periodo == "4º Trimestre":
                        query += " AND strftime('%m', fecha_envio) IN ('10', '11', '12')"
                    elif periodo == "Este mes":
                        query += " AND strftime('%Y-%m', fecha_envio) = strftime('%Y-%m', 'now')"
                    elif periodo == "Mes anterior":
                        last_month = (now.replace(day=1) - timedelta(days=1))
                        query += f" AND strftime('%Y-%m', fecha_envio) = '{last_month.strftime('%Y-%m')}'"
                
                search_text = self.history_search.text().strip()
                if search_text:
                    query += " AND (num_factura LIKE ? OR cliente LIKE ? OR empresa LIKE ?)"
                    search_param = f"%{search_text}%"
                    params.extend([search_param, search_param, search_param])
            
            query += " ORDER BY fecha_envio DESC LIMIT 1000"  # Aumentado límite para exportación
            
            cursor.execute(query, params)
            rows = cursor.fetchall()
            conn.close()
            
            self.table_history.setRowCount(0)
            
            for row in rows:
                (db_id, fecha, num_factura, empresa, estado, detalles, pdf_url, pdf_local_path, importe, cliente) = row
                row_index = self.table_history.rowCount()
                self.table_history.insertRow(row_index)
                
                self.table_history.setItem(row_index, 0, QTableWidgetItem(str(db_id)))
                self.table_history.setItem(row_index, 1, QTableWidgetItem(fecha))
                self.table_history.setItem(row_index, 2, QTableWidgetItem(num_factura))
                self.table_history.setItem(row_index, 3, QTableWidgetItem(empresa))
                self.table_history.setItem(row_index, 4, QTableWidgetItem(cliente))
                
                item_importe = QTableWidgetItem(format_eur(importe))
                item_importe.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                self.table_history.setItem(row_index, 5, item_importe)
                
                chip = StatusChip(estado)
                self.table_history.setCellWidget(row_index, 6, chip)
                
                self.table_history.setItem(row_index, 7, QTableWidgetItem(detalles))
                
                # Botón PDF (ahora en la columna 7)
                btn_pdf = self._make_pdf_button(
                    num_factura,
                    pdf_url,
                    local_path=pdf_local_path,
                    cliente=cliente,
                    importe=format_eur(importe)
                )
                if btn_pdf:
                    self.table_history.setCellWidget(row_index, 8, btn_pdf)

            # Ajustar columnas
            self.table_history.resizeColumnToContents(0)
            self.table_history.resizeColumnToContents(1)
            self.table_history.resizeColumnToContents(2)
            self.table_history.resizeColumnToContents(4)
            self.table_history.resizeColumnToContents(5)
            self.table_history.resizeColumnToContents(7)

            # [NUEVO] Actualizar combo de empresas con empresas únicas
            if hasattr(self, 'history_filter_empresa'):
                empresas_actuales = [self.history_filter_empresa.itemText(i) for i in range(self.history_filter_empresa.count())]
                empresas_en_bd = set()
                for row in rows:
                    if row[3]:  # empresa
                        empresas_en_bd.add(row[3])
                for emp in sorted(empresas_en_bd):
                    if emp not in empresas_actuales:
                        self.history_filter_empresa.addItem(emp)
            
            self.update_dashboard_stats() # Llama a la actualización después de cargar
        except Exception as e:
            self.show_toast(f"Error cargando histórico: {str(e)}")
            print(f"Error cargando histórico: {e}") # Debug
    
    def apply_history_filters(self):
        """Aplica los filtros de búsqueda al histórico."""
        self.load_history(apply_filters=True)
    
    def clear_history_filters(self):
        """Limpia todos los filtros del histórico."""
        if hasattr(self, 'history_filter_empresa'):
            self.history_filter_empresa.setCurrentIndex(0)
        if hasattr(self, 'history_filter_estado'):
            self.history_filter_estado.setCurrentIndex(0)
        if hasattr(self, 'history_filter_periodo'):
            self.history_filter_periodo.setCurrentIndex(0)
        if hasattr(self, 'history_search'):
            self.history_search.clear()
        self.load_history(apply_filters=False)
    
    def export_history(self):
        """Exporta el historial a Excel o CSV."""
        if not hasattr(self, 'table_history') or self.table_history.rowCount() == 0:
            self.show_toast("⚠️ No hay datos en el historial para exportar.")
            return
        
        # Diálogo para elegir formato
        path, selected_filter = QFileDialog.getSaveFileName(
            self, 
            "Exportar Histórico", 
            f"historico_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            "Excel (*.xlsx);;CSV (*.csv)"
        )
        if not path:
            return
        
        try:
            # Obtener datos de la base de datos con los mismos filtros
            conn = sqlite3.connect(DB_PATH)
            cursor = conn.cursor()
            
            query = "SELECT fecha_envio, num_factura, empresa, cliente, importe, estado, detalles, pdf_url FROM envios WHERE 1=1"
            params = []
            
            if hasattr(self, 'history_filter_empresa'):
                empresa = self.history_filter_empresa.currentText()
                if empresa != "Todas las Empresas":
                    query += " AND empresa = ?"
                    params.append(empresa)
                
                estado = self.history_filter_estado.currentText()
                if estado != "Todos":
                    query += " AND estado = ?"
                    params.append(estado)
                
                periodo = self.history_filter_periodo.currentText()
                if periodo != "Todos":
                    now = datetime.now()
                    if periodo == "1º Trimestre":
                        query += " AND strftime('%m', fecha_envio) IN ('01', '02', '03')"
                    elif periodo == "2º Trimestre":
                        query += " AND strftime('%m', fecha_envio) IN ('04', '05', '06')"
                    elif periodo == "3º Trimestre":
                        query += " AND strftime('%m', fecha_envio) IN ('07', '08', '09')"
                    elif periodo == "4º Trimestre":
                        query += " AND strftime('%m', fecha_envio) IN ('10', '11', '12')"
                    elif periodo == "Este mes":
                        query += " AND strftime('%Y-%m', fecha_envio) = strftime('%Y-%m', 'now')"
                    elif periodo == "Mes anterior":
                        last_month = (now.replace(day=1) - timedelta(days=1))
                        query += f" AND strftime('%Y-%m', fecha_envio) = '{last_month.strftime('%Y-%m')}'"
                
                search_text = self.history_search.text().strip()
                if search_text:
                    query += " AND (num_factura LIKE ? OR cliente LIKE ? OR empresa LIKE ?)"
                    search_param = f"%{search_text}%"
                    params.extend([search_param, search_param, search_param])
            
            query += " ORDER BY fecha_envio DESC"
            cursor.execute(query, params)
            rows = cursor.fetchall()
            conn.close()
            
            if path.endswith('.csv'):
                # Exportar a CSV
                import csv
                with open(path, 'w', newline='', encoding='utf-8-sig') as f:
                    writer = csv.writer(f, delimiter=';')
                    writer.writerow(["Fecha", "Factura", "Empresa", "Cliente", "Importe", "Estado", "Detalles", "PDF URL"])
                    for row in rows:
                        writer.writerow(row)
            else:
                # Exportar a Excel
                import pandas as pd
                df = pd.DataFrame(rows, columns=["Fecha", "Factura", "Empresa", "Cliente", "Importe", "Estado", "Detalles", "PDF URL"])
                df.to_excel(path, index=False, engine='openpyxl')
            
            self.show_toast(f"✅ Historial exportado a {path}")
            QDesktopServices.openUrl(QUrl.fromLocalFile(os.path.dirname(path)))
            
        except Exception as e:
            self.show_error(f"Error al exportar historial: {e}")
    
    def focus_search(self):
        """Enfoca el campo de búsqueda según la página actual."""
        current_index = self.nav_list.currentRow()
        if current_index == 0:  # Dashboard
            if hasattr(self, 'dash_search_bar'):
                self.dash_search_bar.setFocus()
        elif current_index == 3:  # Histórico
            if hasattr(self, 'history_search'):
                self.history_search.setFocus()
        elif current_index == 1:  # Cargar Excel
            if hasattr(self, 'table_excel'):
                # Buscar TableTools en la tabla
                for i in range(self.table_excel.parent().layout().count()):
                    widget = self.table_excel.parent().layout().itemAt(i).widget()
                    if isinstance(widget, TableTools):
                        widget.search.setFocus()
                        break
    
    def select_accent_color(self):
        """Abre el selector de color y actualiza el color de acento."""
        global COLOR_PRIMARY
        current_color = QColor(COLOR_PRIMARY)
        color = QColorDialog.getColor(current_color, self, "Seleccionar Color de Acento")
        if color.isValid():
            COLOR_PRIMARY = color.name()
            self.settings.setValue("accent_color", COLOR_PRIMARY)
            
            # Actualizar preview
            if hasattr(self, 'color_preview'):
                self.color_preview.setStyleSheet(f"background-color: {COLOR_PRIMARY}; border-radius: 8px; border: 2px solid {COLOR_BORDER};")
            
            # Recargar estilos completamente
            app = QApplication.instance()
            style_sheet_content = self._get_themed_stylesheet()
            if style_sheet_content:
                app.setStyleSheet(style_sheet_content)
            
            # Forzar actualización de todos los widgets
            self._refresh_styles()
            self.show_toast(f"✅ Color de acento actualizado a {COLOR_PRIMARY}")
    
    def reset_accent_color(self):
        """Restaura el color de acento al valor por defecto."""
        global COLOR_PRIMARY
        COLOR_PRIMARY = "#A0BF6E"  # Color corporativo original
        self.settings.setValue("accent_color", COLOR_PRIMARY)
        
        # Actualizar preview
        if hasattr(self, 'color_preview'):
            self.color_preview.setStyleSheet(f"background-color: {COLOR_PRIMARY}; border-radius: 8px; border: 2px solid {COLOR_BORDER};")
        
        # Recargar estilos completamente
        app = QApplication.instance()
        style_sheet_content = self._get_themed_stylesheet()
        if style_sheet_content:
            app.setStyleSheet(style_sheet_content)
        
        # Forzar actualización de todos los widgets
        self._refresh_styles()
        self.show_toast("✅ Color de acento restaurado al valor por defecto (#A0BF6E)")
    
    def apply_font_size(self, size_text):
        """Aplica el tamaño de fuente seleccionado."""
        size_map = {
            "Pequeño (13px)": 13,
            "Normal (15px)": 15,
            "Grande (17px)": 17,
            "Muy Grande (19px)": 19
        }
        font_size = size_map.get(size_text, 15)
        self.settings.setValue("font_size", size_text)
        
        # Actualizar fuente global
        app = QApplication.instance()
        if app:
            font = QFont("Segoe UI Variable", font_size)
            font.setStyleStrategy(QFont.StyleStrategy.PreferQuality)
            app.setFont(font)
        
        # Recargar estilos
        style_sheet_content = self._get_themed_stylesheet()
        if style_sheet_content:
            app.setStyleSheet(style_sheet_content)
        self._refresh_styles()
        self.show_toast(f"✅ Tamaño de fuente actualizado a {size_text}")
    
    def apply_spacing(self, spacing_text):
        """Aplica el espaciado seleccionado."""
        self.settings.setValue("spacing", spacing_text)
        self.setProperty("spacing", spacing_text.lower())
        
        # Recargar estilos para aplicar espaciado
        app = QApplication.instance()
        style_sheet_content = self._get_themed_stylesheet()
        if style_sheet_content and app:
            app.setStyleSheet(style_sheet_content)
        self._refresh_styles()
        self.show_toast(f"✅ Espaciado actualizado a {spacing_text}")
    
    def generate_excel_template(self):
        """Genera un archivo Excel de plantilla con la estructura correcta."""
        path, _ = QFileDialog.getSaveFileName(
            self,
            "Guardar Plantilla Excel",
            "Plantilla_Facturas.xlsx",
            "Excel (*.xlsx)"
        )
        if not path:
            return
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Macro"
            
            # Encabezados según macro_adapter.py
            headers = [
                "A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M",
                "N", "O", "P", "Q", "R", "S", "T", "U", "V", "W", "X", "Y", "Z", "AA", "AB", "AC"
            ]
            
            # Mapeo de columnas según EXCEL_COLS en macro_adapter.py
            column_map = {
                "A": "num_factura",
                "B": "fecha_factura",
                "C": "tipo_factura",
                "D": "cliente_nombre",
                "E": "cliente_nif",
                "F": "cliente_direccion",
                "G": "cliente_cp",
                "H": "cliente_poblacion",
                "I": "cliente_provincia",
                "J": "cliente_pais",
                "K": "base_imponible",
                "L": "iva_porcentaje",
                "M": "iva_importe",
                "N": "total",
                "O": "forma_pago",
                "P": "vencimiento",
                "Q": "observaciones",
                "R": "empresa_cif",
                "S": "empresa_nombre",
                "T": "empresa_direccion",
                "U": "empresa_cp",
                "V": "empresa_poblacion",
                "W": "empresa_provincia",
                "X": "empresa_pais",
                "Y": "concepto_descripcion",
                "Z": "concepto_cantidad",
                "AA": "concepto_precio",
                "AB": "concepto_importe",
                "AC": "estado_envio"
            }
            
            # Escribir encabezados
            header_row = 1
            for col_idx, col_letter in enumerate(headers, start=1):
                cell = ws.cell(row=header_row, column=col_idx)
                cell.value = column_map.get(col_letter, col_letter)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Añadir fila de ejemplo
            example_row = 2
            example_data = [
                "F001",  # A: num_factura
                "2024-01-15",  # B: fecha_factura
                "F1",  # C: tipo_factura
                "Cliente Ejemplo SL",  # D: cliente_nombre
                "B12345678",  # E: cliente_nif
                "Calle Ejemplo 123",  # F: cliente_direccion
                "28001",  # G: cliente_cp
                "Madrid",  # H: cliente_poblacion
                "Madrid",  # I: cliente_provincia
                "España",  # J: cliente_pais
                "1000.00",  # K: base_imponible
                "21",  # L: iva_porcentaje
                "210.00",  # M: iva_importe
                "1210.00",  # N: total
                "Transferencia",  # O: forma_pago
                "2024-02-15",  # P: vencimiento
                "Factura de ejemplo",  # Q: observaciones
                "A12345678",  # R: empresa_cif
                "Mi Empresa SL",  # S: empresa_nombre
                "Calle Empresa 456",  # T: empresa_direccion
                "28002",  # U: empresa_cp
                "Madrid",  # V: empresa_poblacion
                "Madrid",  # W: empresa_provincia
                "España",  # X: empresa_pais
                "Servicio de ejemplo",  # Y: concepto_descripcion
                "1",  # Z: concepto_cantidad
                "1000.00",  # AA: concepto_precio
                "1000.00",  # AB: concepto_importe
                ""  # AC: estado_envio (vacío)
            ]
            
            for col_idx, value in enumerate(example_data, start=1):
                ws.cell(row=example_row, column=col_idx, value=value)
            
            # Ajustar ancho de columnas
            for col in ws.columns:
                max_length = 0
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)
                ws.column_dimensions[col_letter].width = adjusted_width
            
            # Crear hoja CLIENTES
            ws_clientes = wb.create_sheet("CLIENTES")
            clientes_headers = ["CIF/NIF", "Nombre", "Dirección", "CP", "Población", "Provincia", "País"]
            for col_idx, header in enumerate(clientes_headers, start=1):
                cell = ws_clientes.cell(row=1, column=col_idx)
                cell.value = header
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Ejemplo de cliente
            example_cliente = ["B12345678", "Cliente Ejemplo SL", "Calle Ejemplo 123", "28001", "Madrid", "Madrid", "España"]
            for col_idx, value in enumerate(example_cliente, start=1):
                ws_clientes.cell(row=2, column=col_idx, value=value)
            
            wb.save(path)
            self.show_toast(f"✅ Plantilla Excel generada en {path}")
            QDesktopServices.openUrl(QUrl.fromLocalFile(os.path.dirname(path)))
            
        except Exception as e:
            self.show_error(f"Error al generar plantilla: {e}")
    
    def compress_old_logs(self):
        """Comprime logs antiguos para ahorrar espacio."""
        try:
            import log_compressor
            logs_count = log_compressor.compress_old_logs()
            xmls_count = log_compressor.compress_old_xmls()
            total = logs_count + xmls_count
            if total > 0:
                self.show_toast(f"✅ Comprimidos {total} archivos ({logs_count} logs, {xmls_count} XMLs)")
            else:
                self.show_toast("ℹ️ No hay archivos antiguos para comprimir")
        except Exception as e:
            self.show_error(f"Error al comprimir logs: {e}")
    
    def toggle_offline_mode(self, checked):
        """Habilita/deshabilita el modo offline."""
        self.settings.setValue("offline_mode", "1" if checked else "0")
        os.environ["USE_OFFLINE_QUEUE"] = "1" if checked else "0"
        status = "habilitado" if checked else "deshabilitado"
        self.show_toast(f"✅ Modo offline {status}")
    
    def process_offline_queue(self):
        """Procesa la cola de envíos offline."""
        try:
            import offline_queue
            import prueba
            import requests
            
            # Verificar conexión
            try:
                requests.get("https://www.facturantia.com", timeout=5)
            except:
                self.show_error("❌ No hay conexión a internet. No se puede procesar la cola.")
                return
            
            items = offline_queue.get_pending_items(limit=50)
            if not items:
                self.show_toast("ℹ️ No hay facturas pendientes en la cola")
                return
            
            self.show_toast(f"📤 Procesando {len(items)} facturas de la cola...")
            
            # Procesar cada item
            success_count = 0
            fail_count = 0
            
            for item in items:
                try:
                    result = prueba.send_proforma(
                        item["xml_content"],
                        item["api_key"],
                        item["num_factura"],
                        item["empresa"],
                        item["ejercicio"],
                        item["cliente_doc"],
                        use_offline_queue=False  # No volver a añadir a la cola
                    )
                    
                    if result.get("status") in ["ÉXITO", "DUPLICADO"]:
                        offline_queue.mark_as_sent(item["id"])
                        success_count += 1
                    else:
                        offline_queue.mark_as_failed(item["id"], result.get("details", "Error desconocido"))
                        fail_count += 1
                except Exception as e:
                    offline_queue.mark_as_failed(item["id"], str(e))
                    fail_count += 1
            
            self.show_toast(f"✅ Cola procesada: {success_count} exitosos, {fail_count} fallidos")
            
        except Exception as e:
            self.show_error(f"Error procesando cola offline: {e}")

    # [MODIFICADO] update_dashboard_stats ahora es más simple (sin Top 5)
    def update_dashboard_stats(self):
        """Actualiza las 4 tarjetas principales del Dashboard desde la DB."""
        if not all([self.total_label, self.success_label, self.month_total_label, self.month_count_label]):
            print("Warning: Faltan referencias a las etiquetas del Dashboard.")
            return
        
        try:
            conn = sqlite3.connect(DB_PATH)
            cursor = conn.cursor()

            # 1. Totales Históricos
            cursor.execute("SELECT COUNT(*) FROM envios WHERE estado = 'ÉXITO'")
            total_exitos = cursor.fetchone()[0] or 0

            # 2. Totales Mensuales (Mes actual)
            mes_actual = datetime.now().strftime("%Y-%m")
            cursor.execute(
                "SELECT COUNT(*), SUM(importe) FROM envios WHERE (estado LIKE 'ÉXITO%' OR estado = 'OK' OR estado = 'SUCCESS') AND strftime('%Y-%m', fecha_envio) = ?",
                (mes_actual,)
            )
            mes_count, mes_total = cursor.fetchone() or (0, 0.0)
            
            conn.close()

            # 3. Asignar valores
            self.total_label.setText(str(total_exitos))
            self.success_label.setText(str(total_exitos))
            self.month_count_label.setText(str(mes_count))
            self.month_total_label.setText(format_eur(mes_total or 0.0))

        except Exception as e:
            self.show_toast(f"Error actualizando dashboard: {str(e)}")
            print(f"Error actualizando dashboard: {e}") # Debug


    def populate_dashboard_filters(self):
        """Carga el combo de empresas desde la DB (tabla envios)."""
        self.dash_combo_empresas.clear()
        self.dash_combo_empresas.addItem("Todas las Empresas", userData="ALL")
        try:
            conn = sqlite3.connect(DB_PATH)
            cur = conn.cursor()
            rows = cur.execute("SELECT DISTINCT empresa FROM envios WHERE empresa IS NOT NULL AND empresa != '' ORDER BY empresa").fetchall()
            conn.close()
            for (e,) in rows:
                self.dash_combo_empresas.addItem(str(e), userData=str(e))
        except Exception as e:
            print(f"Error cargando empresas: {e}")
    
    def run_dashboard_query(self):
        """Consulta la DB por empresa y periodo y muestra resultados en la tabla del Dashboard."""
        if not self.dash_combo_empresas or not self.dash_table_resultados:
            return

        emisor = self.dash_combo_empresas.currentData()
        periodo = self.dash_combo_periodo.currentText()

        def _month_bounds(d=None):
            d = d or date.today()
            start = d.replace(day=1).strftime("%Y-%m-%d")
            if d.month == 12:
                end = d.replace(year=d.year+1, month=1, day=1).strftime("%Y-%m-%d")
            else:
                end = d.replace(month=d.month+1, day=1).strftime("%Y-%m-%d")
            return start, end

        def _period_bounds(label: str):
            today = date.today()
            y, m = today.year, today.month
            if label in ("Este Año", "Ejercicio Actual", "Todo el año"):
                return f"{y}-01-01", f"{y+1}-01-01"
            if label == "Año Anterior":
                return f"{y-1}-01-01", f"{y}-01-01"
            if label == "1º Trimestre":
                return f"{y}-01-01", f"{y}-04-01"
            if label == "2º Trimestre":
                return f"{y}-04-01", f"{y}-07-01"
            if label == "3º Trimestre":
                return f"{y}-07-01", f"{y}-10-01"
            if label == "4º Trimestre":
                return f"{y}-10-01", f"{y+1}-01-01"
            if label == "Total Histórico":
                return None, None
            # fallback: mes actual
            return _month_bounds(today)

        dfrom, dto = _period_bounds(periodo)

        where = ["estado = 'ÉXITO'"]
        params = []
        if dfrom:
            where.append("substr(fecha_envio,1,10) >= ?"); params.append(dfrom)
        if dto:
            where.append("substr(fecha_envio,1,10) < ?"); params.append(dto)
        if emisor and emisor != "ALL":
            where.append("empresa = ?"); params.append(emisor)

        sql = "SELECT fecha_envio, num_factura, IFNULL(importe,0.0), cliente, empresa, pdf_url, pdf_local_path FROM envios"
        if where:
            sql += " WHERE " + " AND ".join(where)
        sql += " ORDER BY fecha_envio DESC, id DESC"

        try:
            conn = sqlite3.connect(DB_PATH)
            cur = conn.cursor()
            rows = cur.execute(sql, params).fetchall()
            conn.close()
        except Exception as e:
            self.dash_label_resultado.setText(f"Error consultando DB: {e}")
            return

        # volcar en tabla
        self.dash_table_resultados.setRowCount(0)
        total = 0.0
        for i, (fecha, factura, importe, cliente, empresa, pdf_url, pdf_local_path) in enumerate(rows):
            self.dash_table_resultados.insertRow(i)
            self.dash_table_resultados.setItem(i, 0, QTableWidgetItem(str(fecha)[:10]))
            self.dash_table_resultados.setItem(i, 1, QTableWidgetItem(str(factura or "")))
            self.dash_table_resultados.setItem(i, 2, QTableWidgetItem(str(cliente or "")))
            self.dash_table_resultados.setItem(i, 3, QTableWidgetItem(str(empresa or "")))
            item_imp = QTableWidgetItem(format_eur(importe or 0.0))
            item_imp.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.dash_table_resultados.setItem(i, 4, item_imp)

            # Botón Ver PDF
            btn_pdf = self._make_pdf_button(
                factura,
                pdf_url,
                local_path=pdf_local_path,
                cliente=cliente,
                importe=format_eur(importe or 0.0)
            )
            if btn_pdf:
                self.dash_table_resultados.setCellWidget(i, 5, btn_pdf)

            total += float(importe or 0.0)

        self.dash_label_resultado.setText(f"Total: {format_eur(total)} en {len(rows)} facturas")
        header = self.dash_table_resultados.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(1, QHeaderView.Stretch)
        header.setSectionResizeMode(2, QHeaderView.Stretch)
        header.setSectionResizeMode(3, QHeaderView.Stretch)
        header.setSectionResizeMode(4, QHeaderView.ResizeToContents)
        header.setSectionResizeMode(5, QHeaderView.ResizeToContents)

    def filter_dashboard_table(self):
        """Filtra la tabla de resultados del dashboard según el texto de búsqueda."""
        search_text = self.dash_search_bar.text().strip().lower()
        for i in range(self.dash_table_resultados.rowCount()):
            factura_item = self.dash_table_resultados.item(i, 1)
            cliente_item = self.dash_table_resultados.item(i, 2)
            empresa_item = self.dash_table_resultados.item(i, 3)

            factura_text = factura_item.text().lower() if factura_item else ""
            cliente_text = cliente_item.text().lower() if cliente_item else ""
            empresa_text = empresa_item.text().lower() if empresa_item else ""

            # La fila es visible si el texto de búsqueda está en cualquiera de los campos relevantes
            is_match = (search_text in factura_text or
                        search_text in cliente_text or
                        search_text in empresa_text)

            self.dash_table_resultados.setRowHidden(i, not is_match)


    # ######################################################################
    # MÁS FUNCIONES QUE FALTABAN (HELPER UI)
    # ######################################################################

    def _update_send_badge(self):
        """Actualiza el texto del botón 'Enviar Facturas' en el sidebar."""
        try:
            # Buscar el item 2 (Enviar Facturas)
            item = self.nav_list.item(2) # Asumiendo que "Enviar" es el índice 2
            if not item:
                return

            count = self.loaded_invoice_count
            base_text = "🚀 Enviar Facturas"
            
            if count > 0:
                item.setText(f"{base_text} ({count})")
            else:
                item.setText(base_text)
        except Exception as e:
            print(f"Error actualizando badge: {e}")

    def clear_send_page(self):
        """Resetea la UI de la página de envío a su estado inicial."""
        # --- [MODIFICADO] Limpia también la tabla de previsualización y restaura la visibilidad ---
        if hasattr(self, "table_preview"):
            self.table_preview.setRowCount(0)
            self.table_preview.setVisible(True) # Aseguramos que sea visible
        if hasattr(self, "preview_card"):
            self.preview_card.setVisible(True)
        if hasattr(self, "preview_title"):
            self.preview_title.setVisible(True)

        if hasattr(self, "results_group"):
            self.results_group.setVisible(False) # Ocultamos los resultados

        if hasattr(self, "table_envio"):
            self.table_envio.setRowCount(0)
        if hasattr(self, "log_area"):
            self.log_area.clear()
        if hasattr(self, "progress"):
            self.progress.setValue(0)
            self.progress.setRange(0, 100) # Reset range in case it was indeterminate
        if hasattr(self, "send_stepper"):
            self.send_stepper.set_step(0)
        if hasattr(self, "search_bar"):
            self.search_bar.clear()
        if hasattr(self, "filters"):
            for status, btn in self.filters.items():
                btn.setChecked(False)
        if hasattr(self, "btn_download_pdfs"):
            self.btn_download_pdfs.setEnabled(False)
            
    def open_config_dialog(self):
        """Abre el diálogo de configuración de API."""
        dlg = ConfigDialog(self)
        # Aplicar el QSS actual
        style_sheet_content = self._get_themed_stylesheet()
        if style_sheet_content:
            dlg.setStyleSheet(style_sheet_content)
            dlg.setProperty("theme", self.theme)
        
        # Forzar repintado y centrado (similar al login)
        dlg.style().unpolish(dlg)
        dlg.style().polish(dlg)
        dlg.adjustSize()
        
        screen = QApplication.primaryScreen()
        if screen:
            screen_geometry = screen.availableGeometry()
            dlg_geometry = dlg.geometry()
            center_point = screen_geometry.center() - QPoint(dlg_geometry.width() // 2, dlg_geometry.height() // 2)
            dlg.move(center_point)
        
        dlg.exec()
    
    def check_certificates(self):
        """Consulta la API de Facturantia para obtener información sobre certificados digitales."""
        import requests
        import urllib.parse
        
        # Obtener credenciales de la API
        api_settings = QSettings("FactuNabo", "APIConfig")
        api_url = api_settings.value("api_url", "https://www.facturantia.com/API/proformas_receptor.php")
        api_token = api_settings.value("api_token", "")
        api_user = api_settings.value("api_user", "")
        
        if not api_token or not api_user:
            self.show_error("❌ Debes configurar el Token y Usuario de la API primero.\nVe a 'Configurar Parámetros API'.")
            return
        
        # Intentar diferentes endpoints posibles para consultar certificados
        # Basado en el patrón de la API, podría ser un endpoint de consulta
        base_url = api_url.replace("/proformas_receptor.php", "").replace("/API/proformas_receptor.php", "")
        possible_endpoints = [
            f"{base_url}/API/certificados.php",
            f"{base_url}/API/consultar_certificados.php",
            f"{base_url}/API/empresas_certificados.php",
            f"{base_url}/API/info_certificados.php",
        ]
        
        # También intentar con el mismo endpoint pero con acción diferente
        headers = {
            "X-Usuario-Email": api_user,
            "X-Token": api_token,
            "X-Accion": "consultar_certificados",
        }
        
        self.show_toast("🔍 Consultando certificados...")
        
        # Intentar con el endpoint principal con acción diferente
        try:
            resp = requests.get(api_url, headers=headers, params={"accion": "certificados"}, timeout=30)
            if resp.status_code == 200:
                try:
                    data = resp.json()
                    self._show_certificates_dialog(data)
                    return
                except:
                    pass
        except:
            pass
        
        # Si no funciona, intentar otros endpoints
        for endpoint in possible_endpoints:
            try:
                resp = requests.get(endpoint, headers=headers, timeout=30)
                if resp.status_code == 200:
                    try:
                        data = resp.json()
                        self._show_certificates_dialog(data)
                        return
                    except:
                        # Si no es JSON, mostrar el texto de respuesta
                        self.show_error(f"Respuesta del servidor:\n{resp.text[:500]}")
                        return
            except requests.exceptions.RequestException as e:
                continue
        
        # Si ningún endpoint funcionó, mostrar mensaje informativo
        self.show_error(
            "⚠️ No se pudo consultar los certificados.\n\n"
            "Posibles causas:\n"
            "• El endpoint de certificados no está disponible en esta versión de la API\n"
            "• Verifica que las credenciales de la API sean correctas\n"
            "• Contacta con Facturantia para confirmar el endpoint correcto"
        )
    
    def _show_certificates_dialog(self, data):
        """Muestra un diálogo con la información de los certificados."""
        dialog = QDialog(self)
        dialog.setWindowTitle("Certificados Digitales")
        dialog.setMinimumSize(800, 500)
        
        # Aplicar estilos
        style_sheet_content = self._get_themed_stylesheet()
        if style_sheet_content:
            dialog.setStyleSheet(style_sheet_content)
            dialog.setProperty("theme", self.theme)
        
        layout = QVBoxLayout(dialog)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(15)
        
        title = QLabel("Información de Certificados Digitales")
        title.setFont(QFont(QApplication.font().family(), 16, QFont.Bold))
        layout.addWidget(title)
        
        # Crear tabla para mostrar certificados
        table = QTableWidget()
        table.setColumnCount(4)
        table.setHorizontalHeaderLabels(["Empresa", "Ejercicio", "Fecha Caducidad", "Estado"])
        table.horizontalHeader().setStretchLastSection(True)
        table.setAlternatingRowColors(True)
        
        # Procesar datos según la estructura que devuelva la API
        if isinstance(data, dict):
            if "certificados" in data:
                certs = data["certificados"]
            elif "empresas" in data:
                certs = data["empresas"]
            elif "data" in data:
                certs = data["data"]
            else:
                # Intentar interpretar el diccionario directamente
                certs = [data] if data else []
        elif isinstance(data, list):
            certs = data
        else:
            certs = []
        
        table.setRowCount(len(certs))
        for row, cert in enumerate(certs):
            if isinstance(cert, dict):
                empresa = cert.get("empresa", cert.get("nombre_empresa", cert.get("empresa_nombre", "N/A")))
                ejercicio = cert.get("ejercicio", cert.get("año", cert.get("year", "N/A")))
                fecha_cad = cert.get("fecha_caducidad", cert.get("caducidad", cert.get("expira", cert.get("fecha_expira", "N/A"))))
                estado = cert.get("estado", cert.get("status", "N/A"))
                
                # Determinar color según fecha de caducidad
                try:
                    if fecha_cad and fecha_cad != "N/A":
                        # Intentar diferentes formatos de fecha
                        fecha_obj = None
                        for fmt in ["%Y-%m-%d", "%d/%m/%Y", "%Y-%m-%d %H:%M:%S", "%d-%m-%Y"]:
                            try:
                                fecha_obj = datetime.strptime(str(fecha_cad), fmt)
                                break
                            except:
                                continue
                        
                        if fecha_obj:
                            dias_restantes = (fecha_obj - datetime.now()).days
                            if dias_restantes < 0:
                                estado = f"❌ Caducado ({abs(dias_restantes)} días)"
                            elif dias_restantes < 30:
                                estado = f"⚠️ Caduca en {dias_restantes} días"
                            elif dias_restantes < 90:
                                estado = f"⚠️ Caduca en {dias_restantes} días"
                            else:
                                estado = f"✅ Válido ({dias_restantes} días)"
                except Exception as e:
                    pass
                
                table.setItem(row, 0, QTableWidgetItem(str(empresa)))
                table.setItem(row, 1, QTableWidgetItem(str(ejercicio)))
                table.setItem(row, 2, QTableWidgetItem(str(fecha_cad)))
                table.setItem(row, 3, QTableWidgetItem(str(estado)))
        
        if len(certs) == 0:
            # Si no hay datos estructurados, mostrar el JSON completo
            text_area = QTextEdit()
            text_area.setReadOnly(True)
            text_area.setPlainText(json.dumps(data, indent=2, ensure_ascii=False))
            layout.addWidget(text_area)
        else:
            layout.addWidget(table)
        
        # Botón cerrar
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        btn_close = QPushButton("Cerrar")
        btn_close.setStyleSheet("padding: 8px 20px; min-height: 32px; font-size: 13px;")
        btn_close.clicked.connect(dialog.accept)
        btn_layout.addWidget(btn_close)
        layout.addLayout(btn_layout)
        
        dialog.exec()


    def export_results(self):
        """Exporta la tabla de resultados de envío (self.table_envio) a un CSV."""
        if not hasattr(self, 'table_envio') or self.table_envio.rowCount() == 0:
            self.show_toast("⚠️ No hay resultados para exportar.")
            return

        path, _ = QFileDialog.getSaveFileName(self, "Guardar resultados", "resultados_envio.csv", "CSV (*.csv)")
        if not path:
            return

        try:
            data = []
            headers = ["Factura", "Empresa Emisora", "Estado", "Detalles", "PDF_URL"]
            data.append(headers)

            for row in range(self.table_envio.rowCount()):
                factura = self.table_envio.item(row, 0).text() if self.table_envio.item(row, 0) else ""
                empresa = self.table_envio.item(row, 1).text() if self.table_envio.item(row, 1) else ""
                
                # Estado desde el Chip
                estado_widget = self.table_envio.cellWidget(row, 2)
                estado = ""
                if isinstance(estado_widget, StatusChip):
                    estado = estado_widget.text()
                
                detalles = self.table_envio.item(row, 3).text() if self.table_envio.item(row, 3) else ""
                
                # PDF URL (esto es más complejo, está en el botón)
                # Por simplicidad, lo dejaremos vacío. Si fuera crucial, necesitaríamos
                # almacenar la URL en el 'data' del item de la tabla.
                pdf_url = "" # Simplificación
                
                data.append([factura, empresa, estado, detalles, pdf_url])

            with open(path, 'w', newline='', encoding='utf-8-sig') as f:
                import csv
                writer = csv.writer(f, delimiter=';') # Usar ; para Excel en español
                writer.writerows(data)
                
            self.show_toast(f"✅ Resultados exportados a {path}")
            QDesktopServices.openUrl(QUrl.fromLocalFile(os.path.dirname(path))) # Abrir carpeta

        except Exception as e:
            self.show_error(f"Error al exportar CSV: {e}")

    # --- Toast (Notificaciones) ---
    def show_toast(self, message, duration=3000, color_class="info"):
        """Muestra una notificación toast."""
        if not hasattr(self, 'toast'):
            print(f"TOAST: {message}")
            return
            
        self.toast.setText(message)
        self.toast.setProperty("class", color_class) # info, success, warning, error
        self.toast.style().unpolish(self.toast)
        self.toast.style().polish(self.toast)
        
        self.toast.adjustSize()
        width = self.toast.width() + 40 # Añadir padding horizontal
        self.toast.setFixedWidth(width)
        
        start_y = self.height()
        end_y = self.height() - self.toast.height() - 20
        
        start_x = (self.width() - width) // 2
        
        start_rect = QRect(start_x, start_y, width, self.toast.height())
        end_rect = QRect(start_x, end_y, width, self.toast.height())
        
        self.toast.setGeometry(start_rect)
        self.toast.show()
        
        self.toast_anim.stop()
        self.toast_anim.setDuration(400)
        self.toast_anim.setStartValue(start_rect)
        self.toast_anim.setEndValue(end_rect)
        self.toast_anim.start()
        
        self.toast_timer.stop()
        self.toast_timer.start(duration)

    def hide_toast(self):
        """Oculta el toast con animación."""
        start_rect = self.toast.geometry()
        end_rect = QRect(start_rect.x(), self.height(), start_rect.width(), start_rect.height())
        
        self.toast_anim.stop()
        self.toast_anim.setDuration(300)
        self.toast_anim.setStartValue(start_rect)
        self.toast_anim.setEndValue(end_rect)
        self.toast_anim.start()
        
        # Ocultar realmente después de la animación
        QTimer.singleShot(300, self.toast.hide)
        
    # ######################################################################
    # FIN DEL BLOQUE CORREGIDO
    # ######################################################################


def main():
    # [DEPRECATED] Los atributos AA_EnableHighDpiScaling y AA_UseHighDpiPixmaps están deprecados en PySide6 6.5+
    # PySide6 ahora maneja HiDPI automáticamente, no es necesario configurarlos manualmente
    # if hasattr(Qt, 'AA_EnableHighDpiScaling'):
    #     QApplication.setAttribute(Qt.AA_EnableHighDpiScaling, True)
    # if hasattr(Qt, 'AA_UseHighDpiPixmaps'):
    #     QApplication.setAttribute(Qt.AA_UseHighDpiPixmaps, True)

    app = QApplication(sys.argv)
    
    # Configurar traducciones al español para diálogos nativos de Qt
    translator = QTranslator()
    locale = QLocale(QLocale.Spanish, QLocale.Spain)
    QLocale.setDefault(locale)
    
    # Intentar cargar traducciones de Qt (pueden estar en diferentes ubicaciones)
    qt_translations = [
        f"qtbase_es",  # Nombre común para traducciones de Qt base
        f"qt_es",      # Alternativa
    ]
    
    # Buscar traducciones en ubicaciones comunes
    import PySide6
    pyside6_path = os.path.dirname(PySide6.__file__)
    translations_paths = [
        os.path.join(pyside6_path, "translations"),
        os.path.join(sys.prefix, "share", "PySide6", "translations"),
        os.path.join(os.path.dirname(sys.executable), "translations"),
        resource_path("translations"),
    ]
    
    loaded = False
    for trans_name in qt_translations:
        for trans_path in translations_paths:
            trans_file = os.path.join(trans_path, f"{trans_name}.qm")
            if os.path.exists(trans_file):
                if translator.load(trans_file):
                    app.installTranslator(translator)
                    loaded = True
                    break
        if loaded:
            break
    
    # Aplicar la fuente global ANTES de crear la ventana
    font = QFont("Segoe UI Variable", 15) # <-- Fuente y tamaño base
    
    # --- [CORRECCIÓN AttributeError] ---
    # El valor correcto es QFont.StyleStrategy.PreferQuality
    font.setStyleStrategy(QFont.StyleStrategy.PreferQuality)
    # --- [FIN CORRECCIÓN] ---
    
    app.setFont(font)
    
    # --- [NUEVO] Importar el shim ANTES de crear la ventana ---
    # Esto reemplaza QMessageBox.critical, .question, etc.
    try:
        import dialog_shim
    except ImportError:
        print("Advertencia: no se encontró 'dialog_shim.py' o 'modern_dialogs.py'. Se usarán diálogos nativos.")
    # --- [FIN NUEVO] ---

    window = MainWindow()

    # [NUEVO] Forzar el repintado inicial después de mostrar la ventana
    window.show()
    # Usar QTimer.singleShot para asegurar que _refresh_styles se ejecute después de que la ventana sea visible
    QTimer.singleShot(100, window._refresh_styles) # Aumentado ligero delay

    # El login se ejecuta después de mostrar la ventana principal
    if window.require_login():
        sys.exit(app.exec())
    else:
        sys.exit(0) # Salir si el login es cancelado


if __name__ == "__main__":
    main()