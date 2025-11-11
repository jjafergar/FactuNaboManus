# -*- coding: utf-8 -*-
import pandas as pd, xml.etree.ElementTree as ET
from xml.dom import minidom
import requests, os, json, logging, numpy as np, urllib.parse, re
import unicodedata
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import macro_adapter
import re
import xmlschema

def quitar_tildes_empresa(nombre):
    """Quita acentos SOLO para nombres de empresa emisora"""
    if not isinstance(nombre, str):
        return nombre
    nfkd = unicodedata.normalize("NFKD", nombre)
    sin_tildes = "".join(c for c in nfkd if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", sin_tildes).strip()



def _norm_invoice_id(x: object) -> str:
    s = str(x).strip()
    if re.fullmatch(r"\d+(?:\.0+)?", s):
        try:
            return str(int(float(s)))
        except Exception:
            return s
    return s


LOG_DIR = "logs"
RESPONSE_DIR = "responses"
os.makedirs(LOG_DIR, exist_ok=True)
os.makedirs(RESPONSE_DIR, exist_ok=True)
log_filename = os.path.join(LOG_DIR, f"proforma_import_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log")
for h in logging.root.handlers[:]:
    logging.root.removeHandler(h)
logging.basicConfig(level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    handlers=[logging.FileHandler(log_filename, encoding="utf-8"), logging.StreamHandler()],
)

GUI_LOGGER_CALLBACK = None
def set_gui_logger(cb):
    global GUI_LOGGER_CALLBACK
    GUI_LOGGER_CALLBACK = cb

def log(msg):
    logging.info(msg)
    if GUI_LOGGER_CALLBACK:
        try:
            GUI_LOGGER_CALLBACK(str(msg))
        except Exception:
            pass

# --- utilidades excel macro (marcar/borrar filas) ---
MACRO_SHEET_NAMES = ["Macro", "MACRO", "Hoja1", "Resumen"]
COL_NUM_FACTURA = "A"   # según layout actual
COL_ESTADO      = "AC"  # EXCEL_COLS["estado"] = "AC"

def _find_macro_sheet(wb):
    for name in MACRO_SHEET_NAMES:
        if name in wb.sheetnames:
            return wb[name]
    return None # Return None if no matching sheet is found

def _build_row_map(ws, col_letter):
    row_map = {}
    for r in range(2, ws.max_row + 1):
        value = ws[f"{col_letter}{r}"].value
        if value is None:
            continue
        key = _norm_invoice_id(value)
        if key:
            if key not in row_map:
                row_map[key] = []
            row_map[key].append(r)
    return row_map

def _safe_timestamp():
    return datetime.now().strftime("%Y-%m-%d %H:%M")

def mark_rows_in_macro(excel_path, results, estado_col=COL_ESTADO, keep_vba=True):
    wb = load_workbook(excel_path, keep_vba=keep_vba)
    ws = _find_macro_sheet(wb)
    if ws is None:
        log("⚠️ No se encontró la hoja 'Macro' en el Excel. No se marcarán filas.")
        wb.close()
        return
    row_map = _build_row_map(ws, COL_NUM_FACTURA)
    processed_rows = {}  # Tracks which row index to use for a given invoice number
    col = estado_col
    for item in results:
        num = _norm_invoice_id(item.get("id", ""))
        if not num or num not in row_map:
            continue

        # Get the current index for this invoice number, defaulting to 0
        idx = processed_rows.get(num, 0)

        if idx < len(row_map[num]):
            r = row_map[num][idx]
            status = str(item.get("status", "")).upper()
            details = item.get("details", "")
            if status in ("ÉXITO", "OK", "SUCCESS"):
                ws[f"{col}{r}"].value = f"ENVIADA OK ({_safe_timestamp()})"
            elif status in ("DUPLICATE", "DUPLICADO"):
                ws[f"{col}{r}"].value = f"DUPLICADA ({_safe_timestamp()})"
            else:
                short = (str(details) or "").strip()
                if len(short) > 200:
                    short = short[:200] + "…"
                ws[f"{col}{r}"].value = f"ERROR: {short} ({_safe_timestamp()})"

            # Increment the index for the next time we see this invoice number
            processed_rows[num] = idx + 1

    wb.save(excel_path)

def delete_ok_rows_in_macro(excel_path, results, keep_vba=True):
    wb = load_workbook(excel_path, keep_vba=keep_vba)
    ws = _find_macro_sheet(wb)
    if ws is None:
        log("⚠️ No se encontró la hoja 'Macro' en el Excel. No se eliminarán filas.")
        wb.close()
        return
    row_map = _build_row_map(ws, COL_NUM_FACTURA)
    processed_rows = {}  # Tracks which row index to use for a given invoice number
    to_delete = []
    for item in results:
        num = _norm_invoice_id(item.get("id", ""))
        if not num or num not in row_map:
            continue

        # Get the current index for this invoice number, defaulting to 0
        idx = processed_rows.get(num, 0)

        if idx < len(row_map[num]):
            status = str(item.get("status", "")).upper()
            if status in ("ÉXITO", "OK", "SUCCESS", "DUPLICATE", "DUPLICADO"):
                to_delete.append(row_map[num][idx])

            # Increment the index for the next time we see this invoice number
            processed_rows[num] = idx + 1

    if not to_delete:
        wb.close()
        return

    # Sort rows in descending order to avoid shifting issues when deleting
    for r in sorted(to_delete, reverse=True):
        ws.delete_rows(r, 1)

    wb.save(excel_path)

# --- helpers XML/envío ---
def excel_date_to_datetime(excel_date):
    try:
        if isinstance(excel_date, (int, float)):
            return pd.Timestamp("1899-12-30") + pd.Timedelta(days=excel_date)
        return pd.to_datetime(excel_date)
    except Exception as e:
        log(f"Error al convertir fecha {excel_date}: {e}")
        raise ValueError(f"Formato de fecha inválido: {excel_date}")

def create_sub_element(parent, tag, text, default=None):
    if text is not None and pd.notna(text) and str(text).strip() != "":
        text_str = str(text)
        if isinstance(text, datetime):
            text_str = text.strftime("%Y-%m-%d")
        elif isinstance(text, (int, float, np.number)):
            if tag == "unidad_medida":
                try:
                    text_str = f"{int(float(text)):02d}"
                except (ValueError, TypeError):
                    text_str = str(text)
            else:
                try:
                    f = float(text)
                    text_str = str(int(f)) if f == int(f) else f"{f:.2f}"
                except Exception:
                    text_str = str(text)
    elif default is not None:
        text_str = str(default)
    else:
        text_str = ""
    el = ET.SubElement(parent, tag)
    el.text = text_str
    return el

def prettify(elem):
    return minidom.parseString(ET.tostring(elem, "utf-8")).toprettyxml(indent="    ", encoding="utf-8")

# --- Validación XSD (cache de esquema y resultados) ---
_XSD_SCHEMA = None
_VALIDATION_CACHE = {}  # Cache de validaciones: hash(xml_bytes) -> (is_valid, error_msg)

def _get_schema():
    global _XSD_SCHEMA
    if _XSD_SCHEMA is None:
        xsd_path = os.path.join(os.path.dirname(__file__), "EsquemaProformas.xsd")
        _XSD_SCHEMA = xmlschema.XMLSchema(xsd_path)
    return _XSD_SCHEMA

def _get_xml_hash(xml_bytes: bytes) -> str:
    """Genera un hash del XML para usar como clave de caché."""
    import hashlib
    return hashlib.md5(xml_bytes).hexdigest()

def validate_xml_against_xsd(xml_bytes: bytes, use_cache: bool = True) -> None:
    """Lanza excepción si el XML no cumple el XSD. Usa caché si está habilitado."""
    if use_cache:
        xml_hash = _get_xml_hash(xml_bytes)
        if xml_hash in _VALIDATION_CACHE:
            is_valid, error_msg = _VALIDATION_CACHE[xml_hash]
            if not is_valid:
                raise ValueError(error_msg)
            return  # Validación exitosa desde caché
    
    # Validación real
    try:
        schema = _get_schema()
        schema.validate(xml_bytes)  # admite bytes/str/etree
        # Guardar en caché si está habilitado
        if use_cache:
            xml_hash = _get_xml_hash(xml_bytes)
            _VALIDATION_CACHE[xml_hash] = (True, None)
    except Exception as e:
        error_msg = f"Validación XSD fallida: {e}"
        # Guardar error en caché si está habilitado
        if use_cache:
            xml_hash = _get_xml_hash(xml_bytes)
            _VALIDATION_CACHE[xml_hash] = (False, error_msg)
        raise ValueError(error_msg)

def clear_validation_cache():
    """Limpia el caché de validaciones."""
    global _VALIDATION_CACHE
    _VALIDATION_CACHE.clear()

def _safe_num(x, d=0.0):
    try:
        if x is None:
            return d
        if isinstance(x, str) and x.strip() == "":
            return d
        v = float(x)
        if np.isnan(v):
            return d
        return v
    except Exception:
        return d

def _build_descripcion_general(factura_row, df_conc):
    import re as _re
    dg = factura_row.get("descripcion_general", "")
    if dg is None or str(dg).strip() == "":
        titles = df_conc["descripcion"].dropna().astype(str).str.strip()
        dg = " | ".join(titles.tolist()[:4]) if not titles.empty else "Servicios profesionales"
    dg = _re.sub(r"\s+", " ", str(dg)).strip()
    return dg[:250]

def create_xml_from_data(df_factura_single, df_conceptos_single, df_forma_pago_single, df_conceptos_texto_single=None, df_factura_all=None, df_conceptos_all=None):
    root = ET.Element("proformas")
    for _, factura_row in df_factura_single.iterrows():
        proforma = ET.SubElement(root, "proforma")
        num_factura_actual = factura_row["NumFactura"]
        empresa_actual = factura_row["empresa_emisora"]

        df_conc = df_conceptos_single[(df_conceptos_single["NumFactura"] == num_factura_actual) &
                               (df_conceptos_single["empresa_emisora"] == empresa_actual)].copy()
        if df_conc.empty:
            raise ValueError(f"No se encontraron conceptos para la factura {num_factura_actual} ({empresa_actual}).")
        required_cols = ["NumFactura","empresa_emisora","descripcion","cuenta_contable",
                         "unidad_medida","unidades","base_unidad","tipo_impuesto","porcentaje"]
        missing = [c for c in required_cols if c not in df_conc.columns]
        if missing:
            raise ValueError(f"Faltan columnas requeridas en Conceptos: {', '.join(missing)}")
        invalid = df_conc[df_conc["base_unidad"].fillna(0) <= 0]
        if not invalid.empty:
            raise ValueError(f"Conceptos con base_unidad <= 0: {invalid['descripcion'].tolist()}")

        # Cálculos por línea
        df_conc["base_imponible_cal"] = df_conc["unidades"].fillna(0) * df_conc["base_unidad"].fillna(0)
        df_conc["importe_iva_cal"]    = df_conc["base_imponible_cal"] * (df_conc["porcentaje"].fillna(0) / 100.0)

        def calc_ret(r):
            try:
                if str(r.get("tipo_impuesto_retenido","")).strip() == "":
                    return 0.0
                p = float(r.get("porcentaje_retenido", 0.0))
                if p <= 0:
                    return 0.0
                b = float(r["base_imponible_cal"])
                return round(b * (p/100.0), 2)
            except Exception:
                return 0.0
        df_conc["importe_retencion_cal"] = df_conc.apply(calc_ret, axis=1)
        df_conc["importe_total_cal"] = df_conc["base_imponible_cal"] + df_conc["importe_iva_cal"] - df_conc["importe_retencion_cal"]

        # Totales
        total_base = _safe_num(df_conc["base_imponible_cal"].sum())
        total_iva  = _safe_num(df_conc["importe_iva_cal"].sum())
        total_ret  = _safe_num(df_conc["importe_retencion_cal"].sum())
        total_fact = _safe_num(df_conc["importe_total_cal"].sum())

        # Cabecera
        create_sub_element(proforma, "external_id", _norm_invoice_id(factura_row["NumFactura"]))

        # Serie (derivada si viene vacía)
        serie_opt = (factura_row.get("serie_factura") or "").strip()
        if not serie_opt:
            nf_upper = str(
                factura_row.get("NumFactura")
                or factura_row.get("external_id")
                or factura_row.get("id")
                or ""
            ).strip().upper()
            if nf_upper.startswith("INT25_"):
                serie_opt = "Int25_"
            elif nf_upper.startswith("A25"):
                serie_opt = "A25"
            elif nf_upper.startswith("AB25"):
                serie_opt = "AB25"
            else:
                serie_opt = "25"
        create_sub_element(proforma, "serie_factura", serie_opt)

        create_sub_element(proforma, "fecha_emision", excel_date_to_datetime(factura_row["fecha_emision"]))
        create_sub_element(proforma, "fecha_vencimiento", excel_date_to_datetime(factura_row["fecha_vencimiento"]))
        create_sub_element(proforma, "descripcion_general", _build_descripcion_general(factura_row, df_conc))
        create_sub_element(proforma, "permitir_email_y_afc", factura_row.get("permitir_email_y_afc"), default="1")

        # Literales legales en función de serie cuando IVA total es 0
        serie_hdr = str(factura_row.get("serie_factura","") or "").upper()
        if not re.match(r"^(A25|INT25|NC)", serie_hdr):
            serie_hdr = str(factura_row.get("NumFactura") or factura_row.get("external_id") or "").upper()
        literals = (factura_row.get("literales_legales") or "").strip()
        if total_iva <= 0.0000001:
            if serie_hdr.startswith("A25"):
                add_msg = "Operación exenta de IVA (art. 25 LIVA – entrega intracomunitaria)."
            elif serie_hdr.startswith("INT25"):
                add_msg = "Operación exenta de IVA (art. 20.Uno.18 Ley 37/1992)."
            elif serie_hdr.startswith("NC"):
                add_msg = "Operación no sujeta por reglas de localización (art. 69 LIVA – destinatario fuera de la UE)."
            else:
                add_msg = "Operación exenta de IVA."
            literals = (literals + ("\n" if literals else "") + add_msg)[:2500]
        create_sub_element(proforma, "literales_legales", literals, default="")
        create_sub_element(proforma, "referencia", factura_row.get("referencia", factura_row.get("NumFactura")))
        # Normalización de tipo_factura contra XSD {F1,F2,F3,R1..R5}
        allowed_tipos = {"F1","F2","F3","R1","R2","R3","R4","R5"}
        tf_raw = str(factura_row.get("tipo_factura","") or "").strip().upper()
        tipo_factura_norm = tf_raw if tf_raw in allowed_tipos else "F1"
        # Política pedida: rectificativa = R4 por defecto, salvo R1 cuando el motivo sea error de tipo/IVA.
        if tipo_factura_norm.startswith("R"):
            # Heurística de "error de IVA": si textos mencionan IVA/tipo/porcentaje o
            # si coexisten líneas con IVA > 0 y fiscalidad especial en la misma proforma (señal de cambio de tratamiento).
            textos = " ".join([
                str(factura_row.get("descripcion_general","") or ""),
                str(factura_row.get("texto_libre","") or ""),
                literals
            ]).lower()
            mentions_iva = ("iva" in textos) and (("tipo" in textos) or ("porcentaje" in textos) or ("rectific" in textos))
            has_iva_positive = bool(df_conc["porcentaje"].fillna(0).gt(0).any())
            has_all_zero = bool(df_conc["porcentaje"].fillna(0).eq(0).all())
            # Si hay mención explícita o hay indicios de cambio de gravado↔exento, usar R1
            if mentions_iva or (has_iva_positive and not has_all_zero):
                tipo_factura_norm = "R1"
            else:
                tipo_factura_norm = "R4"
        create_sub_element(proforma, "tipo_factura", tipo_factura_norm, default="F1")
        create_sub_element(proforma, "texto_libre", factura_row.get("texto_libre"), default="")
        create_sub_element(proforma, "plantilla_facturas_emitidas", factura_row.get("plantilla_facturas_emitidas"), default="")
        create_sub_element(proforma, "plantilla_facturas_proforma", factura_row.get("plantilla_facturas_proforma"), default="")
        create_sub_element(proforma, "ejercicio", factura_row["ejercicio"], default="2025")

        # Campos extra cuando es rectificativa (opcionales en XSD, pero recomendables)
        if tipo_factura_norm.startswith("R"):
            # booleano estilo '1'/'0'
            create_sub_element(proforma, "factura_rectificativa", "1")
            # Los siguientes solo si vienen; usamos defaults prudentes
            if pd.notna(factura_row.get("factura_rectificativa_numero", None)):
                create_sub_element(proforma, "factura_rectificativa_numero", factura_row.get("factura_rectificativa_numero"))
            if pd.notna(factura_row.get("factura_rectificativa_ejercicio", None)):
                create_sub_element(proforma, "factura_rectificativa_ejercicio", factura_row.get("factura_rectificativa_ejercicio"))
            if pd.notna(factura_row.get("factura_rectificativa_fecha_emision", None)):
                create_sub_element(proforma, "factura_rectificativa_fecha_emision", excel_date_to_datetime(factura_row.get("factura_rectificativa_fecha_emision")))
            # Campo renombrado: factura_rectificativa_codigo -> factura_rectificativa_motivo_codigo
            # Mantener compatibilidad con ambos nombres por si acaso
            motivo_codigo = factura_row.get("factura_rectificativa_motivo_codigo") or factura_row.get("factura_rectificativa_codigo")
            if pd.notna(motivo_codigo):
                create_sub_element(proforma, "factura_rectificativa_codigo", motivo_codigo)
            # Tipo de rectificativa: I (diferencias) o S (sustitución)
            rect_tipo = str(factura_row.get("factura_rectificativa_tipo","") or "").strip().upper() or "I"
            if rect_tipo not in {"I","S"}:
                rect_tipo = "I"
            create_sub_element(proforma, "factura_rectificativa_tipo", rect_tipo)

        # Cliente
        cliente = ET.SubElement(proforma, "cliente")
        create_sub_element(cliente, "tipo_persona", factura_row["cliente_tipo_persona"], default="J")
        create_sub_element(cliente, "nombre", factura_row["cliente_nombre"], default="Cliente")
        # Derivar doc/residencia según serie (A25 intracomunitaria -> NIF-IVA '02', residencia UE 'U'; NC -> extranjero 'E')
        doc_tipo_in = str(factura_row.get("cliente_tipo_documento","") or "nif")
        residencia_in = str(factura_row.get("cliente_tipo_residencia","") or "R")
        codigo_pais_in = str(factura_row.get("cliente_codigo_pais","") or "ESP")
        serie_for_res = str(factura_row.get("serie_factura","") or "").upper()
        if not re.match(r"^(A25|INT25|NC)", serie_for_res):
            serie_for_res = str(factura_row.get("NumFactura") or factura_row.get("external_id") or "").upper()
        # Reglas de residencia/documento (priorizar serie):
        # - A25 (intracomunitaria): UE 'U' y NIF-IVA '02'
        # - NC (no sujeta por localización): Extranjero 'E'
        # - Si no es A25/NC y país = España: Residente 'R'
        if serie_for_res.startswith("A25"):
            doc_tipo_in = "02"  # NIF-IVA
            residencia_in = "U" # Residente UE
        elif serie_for_res.startswith("NC"):
            residencia_in = "E" # Extranjero (no UE)
        elif codigo_pais_in.upper() in ("ESP","ES"):
            residencia_in = "R"
        create_sub_element(cliente, "tipo_documento", doc_tipo_in, default="nif")
        create_sub_element(cliente, "numero_documento", factura_row["cliente_numero_documento"], default="B06989537")
        create_sub_element(cliente, "cuenta_contable", factura_row["cliente_cuenta_contable"], default="4300010")
        create_sub_element(cliente, "observacion", factura_row["cliente_observacion"], default="")
        create_sub_element(cliente, "tipo_residencia", residencia_in, default="R")
        create_sub_element(cliente, "codigo_pais", codigo_pais_in, default="ESP")
        create_sub_element(cliente, "provincia", factura_row["cliente_provincia"][:20] if isinstance(factura_row["cliente_provincia"], str) else "", default="")
        create_sub_element(cliente, "poblacion", factura_row["cliente_poblacion"], default="")
        create_sub_element(cliente, "domicilio", factura_row["cliente_domicilio"], default="Calle Ejemplo 1")
        create_sub_element(cliente, "domicilio_2", factura_row.get("cliente_domicilio_2",""), default="")
        create_sub_element(cliente, "cp", factura_row["cliente_cp"], default="")
        create_sub_element(cliente, "telefono", factura_row["cliente_telefono"], default="")
        create_sub_element(cliente, "email", factura_row["cliente_email"], default="")

        # Conceptos
        conceptos = ET.SubElement(proforma, "conceptos")
        for _, c in df_conc.iterrows():
            concepto = ET.SubElement(conceptos, "concepto")
            desc = str(c["descripcion"]) if pd.notna(c["descripcion"]) else ""
            if len(desc) > 2500:
                desc = desc[:2500]
            create_sub_element(concepto, "descripcion", desc)
            create_sub_element(concepto, "categoria", c.get("categoria", "SERVICIOS"))
            create_sub_element(concepto, "cuenta_contable", c["cuenta_contable"])
            create_sub_element(concepto, "unidad_medida", c["unidad_medida"])
            create_sub_element(concepto, "descripcion_larga", c.get("descripcion_larga", ""))
            create_sub_element(concepto, "unidades", _safe_num(c["unidades"], 1.0))
            create_sub_element(concepto, "base_unidad", _safe_num(c["base_unidad"], 0.0))

            # IVA línea o fiscalidad especial
            porc_iva_linea = _safe_num(c.get("porcentaje", 0.0), 0.0)
            if porc_iva_linea > 0:
                imp_rep = ET.SubElement(concepto, "impuestos_repercutidos")
                ir = ET.SubElement(imp_rep, "impuesto_repercutido")
                create_sub_element(ir, "impuesto", str(c["tipo_impuesto"]).upper())
                create_sub_element(ir, "porcentaje", _safe_num(c["porcentaje"], 0.0))
                create_sub_element(ir, "importe", _safe_num(c["importe_iva_cal"], 0.0))
                create_sub_element(ir, "base_imponible", _safe_num(c["base_imponible_cal"], 0.0))
                create_sub_element(ir, "recargo_equivalencia_importe", 0.00)
                create_sub_element(ir, "importe_especial", 0.00)
                create_sub_element(ir, "importe_total", _safe_num(c["importe_iva_cal"], 0.0))
            else:
                fiscalidad = ET.SubElement(concepto, "fiscalidad_especial")
                # Detectar serie efectiva desde cabecera o número
                serie_fx = str(factura_row.get("serie_factura","") or "").upper()
                if not re.match(r"^(A25|INT25|NC)", serie_fx):
                    serie_fx = str(factura_row.get("NumFactura") or factura_row.get("external_id") or "").upper()
                if serie_fx.startswith("A25"):
                    # Intracomunitaria: exenta por artículo 25
                    create_sub_element(fiscalidad, "clave_regimen", "01")
                    create_sub_element(fiscalidad, "tipo", "E5")
                    create_sub_element(fiscalidad, "justificacion", "Operación exenta de IVA (art. 25 LIVA – entrega intracomunitaria).")
                elif serie_fx.startswith("INT25"):
                    # Intereses/servicios exentos art. 20.Uno.18
                    create_sub_element(fiscalidad, "clave_regimen", "01")
                    create_sub_element(fiscalidad, "tipo", "E1")
                    create_sub_element(fiscalidad, "justificacion", "Operación exenta de IVA (art. 20.Uno.18 LIVA).")
                elif serie_fx.startswith("NC"):
                    # No sujeta por reglas de localización
                    create_sub_element(fiscalidad, "clave_regimen", "01")
                    create_sub_element(fiscalidad, "tipo", "N2")
                    create_sub_element(fiscalidad, "justificacion", "No sujeta por reglas de localización (art. 69 LIVA).")
                else:
                    # Exenta genérica
                    create_sub_element(fiscalidad, "clave_regimen", "01")
                    create_sub_element(fiscalidad, "tipo", "E1")
                    create_sub_element(fiscalidad, "justificacion", "SIN IVA.")

            # Retención línea (con forzado 19% si serie INT25)
            ret_tipo = str(c.get("tipo_impuesto_retenido", "") or "").strip()
            ret_porc = float(c.get("porcentaje_retenido", 0.0) or 0.0)
            ret_imp  = float(c.get("importe_retencion_cal", 0.0) or 0.0)

            serie_val__ = str(factura_row.get("serie_factura","") or "").upper()
            if not re.match(r"^(A25|INT25|NC)", serie_val__):
                serie_val__ = str(factura_row.get("NumFactura") or factura_row.get("external_id") or "").upper()
            if serie_val__.startswith("INT25"):
                ret_tipo = "IRPF"
                ret_porc = 19.0
                try:
                    base_cal = float(_safe_num(c["base_imponible_cal"], 0.0))
                except Exception:
                    base_cal = 0.0
                ret_imp = round(base_cal * (ret_porc/100.0), 2)

            if ret_tipo != "" and ret_porc > 0 and ret_imp > 0:
                imp_ret = ET.SubElement(concepto, "impuestos_retenidos")
                ir2 = ET.SubElement(imp_ret, "impuesto_retenido")
                create_sub_element(ir2, "impuesto", "IRPF")
                create_sub_element(ir2, "porcentaje", ret_porc)
                create_sub_element(ir2, "base_imponible", _safe_num(c["base_imponible_cal"], 0.0))
                create_sub_element(ir2, "importe", ret_imp)
                create_sub_element(concepto, "total_impuestos_retenidos", ret_imp)
            else:
                create_sub_element(concepto, "total_impuestos_retenidos", 0.00)

            create_sub_element(concepto, "total_impuestos_repercutidos", _safe_num(c["importe_iva_cal"], 0.0))
            create_sub_element(concepto, "importe_bruto", _safe_num(c["base_imponible_cal"], 0.0))
            create_sub_element(concepto, "base_imponible", _safe_num(c["base_imponible_cal"], 0.0))
            create_sub_element(concepto, "importe_total", _safe_num(c["importe_total_cal"], 0.0))

        # Conceptos texto (respetando la posición del Excel si viene; si no, por índice estable)
        if df_conceptos_texto_single is not None and not df_conceptos_texto_single.empty:
            df_txt = df_conceptos_texto_single[
                (df_conceptos_texto_single["NumFactura"] == num_factura_actual) &
                (df_conceptos_texto_single["empresa_emisora"] == empresa_actual)
            ].copy()
            if not df_txt.empty:
                if "posicion" not in df_txt.columns:
                    df_txt = df_txt.reset_index(drop=True)
                    df_txt["posicion"] = df_txt.index
                df_txt = df_txt.sort_values(by="posicion", kind="stable")
                conceptos_textos = ET.SubElement(proforma, "conceptos_textos")
                for _, r in df_txt.iterrows():
                    ctxt = ET.SubElement(conceptos_textos, "concepto_texto")
                    ET.SubElement(ctxt, "posicion").text = str(int(r.get("posicion", 0)))
                    txt = str(r.get("descripcion", ""))[:2500]
                    ET.SubElement(ctxt, "descripcion").text = txt

        # --- Resumen de IVA (solo si porcentaje > 0) ---
        tiene_iva = df_conc["porcentaje"].fillna(0).gt(0).any()
        if tiene_iva:
            resumen = ET.SubElement(proforma, "impuestos_repercutidos")
            for porc, grp in df_conc.groupby(df_conc["porcentaje"].fillna(0.0)):
                base_sum = round(float(grp["base_imponible_cal"].sum()), 2)
                iva_sum  = round(float(grp["importe_iva_cal"].sum()), 2)
                if iva_sum <= 0:
                    continue
                ir = ET.SubElement(resumen, "impuesto_repercutido")
                create_sub_element(ir, "impuesto", "IVA")
                try:
                    porc_val = float(porc)
                    porc_out = int(round(porc_val)) if porc_val.is_integer() else porc_val
                except Exception:
                    porc_out = porc
                create_sub_element(ir, "porcentaje", porc_out)
                create_sub_element(ir, "importe", iva_sum)
                create_sub_element(ir, "base_imponible", base_sum)
                create_sub_element(ir, "recargo_equivalencia_importe", 0.00)
                create_sub_element(ir, "importe_especial", 0.00)
                create_sub_element(ir, "importe_total", iva_sum)

        # --- Resumen de RETENCIONES a nivel proforma (si hay) ---
        tiene_ret = df_conc["importe_retencion_cal"].fillna(0).gt(0).any()
        if tiene_ret:
            resumen_ret = ET.SubElement(proforma, "impuestos_retenidos")
            df_ret = df_conc[df_conc["importe_retencion_cal"].fillna(0) > 0].copy()
            for porc, grp in df_ret.groupby(df_ret["porcentaje_retenido"].fillna(0.0)):
                base_sum = round(float(grp["base_imponible_cal"].sum()), 2)
                ret_sum  = round(float(grp["importe_retencion_cal"].sum()), 2)
                ir2 = ET.SubElement(resumen_ret, "impuesto_retenido")
                tipo = str(grp["tipo_impuesto_retenido"].iloc[0] or "IRPF")
                create_sub_element(ir2, "impuesto", tipo)
                try:
                    porc_val = float(porc)
                    porc_out = int(round(porc_val)) if porc_val.is_integer() else porc_val
                except Exception:
                    porc_out = porc
                create_sub_element(ir2, "porcentaje", porc_out)
                create_sub_element(ir2, "importe", ret_sum)
                create_sub_element(ir2, "base_imponible", base_sum)

        # Totales
        suplidos = _safe_num(factura_row.get("total_suplidos", 0.0), 0.0)
        total_base = _safe_num(total_base, 0.0)
        total_iva  = _safe_num(total_iva, 0.0)
        total_ret  = _safe_num(total_ret, 0.0)
        total_fact = _safe_num(total_fact, 0.0)
        importe_total_con_suplidos = _safe_num(total_fact + suplidos, 0.0)

        create_sub_element(proforma, "importe_bruto", total_base)
        create_sub_element(proforma, "total_importe_bruto", total_base)
        create_sub_element(proforma, "total_impuestos_repercutidos", total_iva)
        create_sub_element(proforma, "total_impuestos_retenidos", total_ret)
        create_sub_element(proforma, "importe_total", importe_total_con_suplidos)
        create_sub_element(proforma, "total_subvenciones", factura_row.get("total_subvenciones", 0.00), default=0.00)
        create_sub_element(proforma, "total_anticipos", factura_row.get("total_anticipos", 0.00), default=0.00)
        create_sub_element(proforma, "total_a_pagar", _safe_num(importe_total_con_suplidos - _safe_num(factura_row.get("total_anticipos", 0.0), 0.0), 0.0))
        create_sub_element(proforma, "total_suplidos", suplidos)
        create_sub_element(proforma, "total_gastos_financieros", factura_row.get("total_gastos_financieros", 0.00), default=0.00)
        create_sub_element(proforma, "total_retencion", total_ret)
        create_sub_element(proforma, "total_pagos_especie", factura_row.get("total_pagos_especie", 0.00), default=0.00)
        create_sub_element(proforma, "total_a_ejecutar", importe_total_con_suplidos)

        # Forma de pago
        df_fp = df_forma_pago_single[(df_forma_pago_single["NumFactura"] == num_factura_actual) &
                              (df_forma_pago_single["empresa_emisora"] == empresa_actual)]
        if df_fp.empty:
            raise ValueError(f"No se encontró forma de pago para {num_factura_actual} ({empresa_actual}).")
        fp = df_fp.iloc[0]
        forma = ET.SubElement(proforma, "forma_pago")
        create_sub_element(forma, "metodo", fp["metodo"], default="transferencia")
        create_sub_element(forma, "transferencia_banco", fp["transferencia_banco"], default="ABANCA")
        create_sub_element(forma, "transferencia_beneficiario", fp["transferencia_beneficiario"], default=empresa_actual)
        create_sub_element(forma, "transferencia_concepto", fp["transferencia_concepto"], default="Pago Factura")
        create_sub_element(forma, "transferencia_iban", fp["transferencia_iban"])
        create_sub_element(forma, "transferencia_bic", fp["transferencia_bic"], default="CAGLESMMXXX")

    return prettify(root)

def _sanitize_token(tok: str) -> str:
    s = str(tok or "")
    for ch in ("\u200b","\u200f","\ufeff","\xa0"):
        s = s.replace(ch, " ")
    s = s.strip().strip('"').strip("'")
    s = re.sub(r"\s+", "", s)
    if re.fullmatch(r"\d+(\.0+)?", s):
        try:
            return str(int(float(s)))
        except Exception:
            return s
    if re.fullmatch(r"\d+([.,]\d+)?[eE][+\-]?\d+", s) or re.fullmatch(r"\d+[.,]\d+", s):
        try:
            val = float(s.replace(",", "."))
            return str(int(val)) if val.is_integer() else ("%.15f" % val).rstrip("0").rstrip(".")
        except Exception:
            return s
    return s

def send_proforma(xml_content, api_key, external_id, empresa, ejercicio, cliente_numero_documento,
                  api_email=None, api_url=None, use_offline_queue=False):
    """Envía la proforma a la API de Facturantia con manejo robusto de duplicados y errores de fecha."""
    import urllib.parse
    url = api_url or "https://www.facturantia.com/API/proformas_receptor.php"
    user_email = api_email or "facturacion@abalados.es"
    token_str = _sanitize_token(api_key)
    empresa_header = quitar_tildes_empresa(empresa)

    if not token_str:
        return {"external_id": external_id, "empresa": empresa, "status": "ERROR_SIN_API_KEY",
                "details": "Falta API Key", "pdf_url": None}

    headers = {
        "X-Usuario-Email": user_email,
        "X-Token": token_str,
        "X-Empresa-Nombre": empresa_header,
        "X-Empresa-Ejercicio": str(ejercicio),
        "X-Accion": "emitir",
        "Content-Type": "application/xml; charset=utf-8",
    }

    auid = "762687769"
    aen_encoded = urllib.parse.quote(empresa)
    aeid = str(external_id)
    acn = cliente_numero_documento
    predictive_pdf_url = f"https://www.facturantia.com/ver_afc_api.php?auid={auid}&aen={aen_encoded}&aej={ejercicio}&aeid={aeid}&acn={acn}"

    summary = {"external_id": external_id, "empresa": empresa, "status": "", "details": "", "pdf_url": None}

    try:
        resp = requests.post(url, data=xml_content, headers=headers, timeout=60)
        try:
            resp_text = resp.content.decode("utf-8")
        except UnicodeDecodeError:
            resp_text = resp.content.decode("iso-8859-1", errors="replace")

        try:
            stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            safe_id = re.sub(r'[^a-zA-Z0-9_-]', '_', str(external_id))
            os.makedirs(RESPONSE_DIR, exist_ok=True)
            raw_json_path = os.path.join(RESPONSE_DIR, f"api_{safe_id}_{stamp}.json")
            with open(raw_json_path, "w", encoding="utf-8") as f:
                f.write(resp_text)
            log(f"↳ API raw guardada en: {raw_json_path}")
        except Exception as io_err:
            log(f"⚠️ No se pudo guardar respuesta cruda: {io_err}")

        if resp.status_code != 200:
            summary["status"] = "API_ERROR"
            summary["details"] = f"HTTP {resp.status_code}: {resp_text}"
            return summary

        try:
            resp_json = resp.json()

            # ÉXITO
            if resp_json.get("estado_envio_facturantia") == "CORRECTO":
                summary["status"] = "ÉXITO"
                summary["details"] = resp_json
                procs = resp_json.get("proformas_procesadas", [])
                if procs and isinstance(procs, list) and procs[0].get("pdf"):
                    summary["pdf_url"] = procs[0].get("pdf")
                else:
                    summary["pdf_url"] = predictive_pdf_url
                
                # [NUEVO] Procesar enlaces QR tributarios si vienen en la respuesta
                # La API ahora devuelve un objeto JSON que relaciona external_id con enlace_qr
                enlaces_qr = resp_json.get("enlaces_qr", {})
                if isinstance(enlaces_qr, dict) and enlaces_qr:
                    # Guardar los enlaces QR en el summary para uso posterior
                    summary["enlaces_qr"] = enlaces_qr
                    # Si hay un enlace QR para este external_id específico, guardarlo también
                    if external_id in enlaces_qr:
                        summary["enlace_qr"] = enlaces_qr[external_id]
                        log(f"📱 Enlace QR tributario recibido para {external_id}")
                
                log(f"✅ Envío correcto para {external_id} ({empresa})")
                return summary

            # ATENCIÓN / DUPLICADOS / FECHA POSTERIOR
            if "mensaje_atencion" in resp_json:
                att = str(resp_json.get("mensaje_atencion", "") or "")
                att_clean = re.sub('<[^>]*>', '', att).lower()
                if ("fecha de emisión" in att_clean) and ("posterior" in att_clean):
                    summary["status"] = "ERROR_FECHA"
                    summary["details"] = "Error: Ya existen facturas con fecha posterior para esta serie."
                    summary["pdf_url"] = None
                    log(f"⛔ {external_id}: Error de Fecha de Emisión posterior detectado.")
                else:
                    summary["status"] = "DUPLICADO"
                    summary["details"] = att or "La factura ya existe (duplicada)."
                    summary["pdf_url"] = predictive_pdf_url
                    log(f"🔁 {external_id}: Factura duplicada detectada.")
                return summary

            # DUPLICADOS REALES
            if ("mensaje_error" in resp_json and
                ("duplicate entry" in str(resp_json.get("mensaje_error","")).lower() or
                 "ya existe" in str(resp_json.get("mensaje_error","")).lower())):
                summary["status"] = "DUPLICADO"
                summary["details"] = resp_json.get("mensaje_error", "La factura ya existe (duplicada).")
                summary["pdf_url"] = predictive_pdf_url
                log(f"🔁 {external_id}: Duplicado real detectado.")
                return summary

            # OTROS ERRORES
            error_msg = resp_json.get("mensaje_error", "Error API desconocido")
            error_msg_cleaned = re.sub('<[^>]*>', '', str(error_msg))
            if ("existen" in error_msg_cleaned.lower() and
                "facturas emitidas" in error_msg_cleaned.lower() and
                "fecha de emisión" in error_msg_cleaned.lower() and
                "posterior" in error_msg_cleaned.lower()):
                summary["status"] = "ERROR_FECHA"
                summary["details"] = "Error: Ya existen facturas con fecha posterior para esta serie."
                summary["pdf_url"] = None
                log(f"⛔ {external_id}: Error de Fecha posterior detectado.")
            else:
                summary["status"] = "ERROR_API_LÓGICO"
                summary["details"] = error_msg
                log(f"⚠️ {external_id}: Error API lógico -> {error_msg}")
            return summary

        except json.JSONDecodeError:
            if "CORRECTO" in resp_text.upper():
                summary["status"] = "ÉXITO"
                summary["details"] = resp_text
                summary["pdf_url"] = predictive_pdf_url
                return summary
            if "fecha de emisión" in resp_text.lower() and "posterior" in resp_text.lower():
                summary["status"] = "ERROR_FECHA"
                summary["details"] = "Error: Ya existen facturas con fecha posterior."
                summary["pdf_url"] = None
                return summary
            if "ya existe" in resp_text.lower():
                summary["status"] = "DUPLICADO"
                summary["details"] = "La factura ya existe (duplicada)."
                summary["pdf_url"] = predictive_pdf_url
                return summary

            summary["status"] = "API_ERROR_NO_JSON"
            summary["details"] = resp_text
            return summary

    except requests.exceptions.Timeout:
        summary["status"] = "ERROR_TIMEOUT"
        summary["details"] = "Timeout API (>60s)"
        log(f"⚠️ Timeout API (>60s) para {external_id}.")
    except requests.exceptions.RequestException as e:
        summary["status"] = "ERROR_CONEXIÓN"
        summary["details"] = str(e)
        log(f"❌ Error de conexión para {external_id}: {e}")
        
        # [NUEVO] Si hay error de conexión y está habilitado, añadir a cola offline
        if use_offline_queue:
            try:
                import offline_queue
                queue_id = offline_queue.add_to_queue(
                    xml_content, external_id, empresa, ejercicio, cliente_numero_documento, api_key
                )
                log(f"📦 Factura {external_id} añadida a cola offline (ID: {queue_id})")
                summary["status"] = "EN_COLA_OFFLINE"
                summary["details"] = f"Error de conexión. Añadida a cola offline (ID: {queue_id})"
            except Exception as queue_err:
                log(f"⚠️ Error añadiendo a cola offline: {queue_err}")

    return summary

def read_excel_any(path, **kwargs):
    lower = path.lower()
    if lower.endswith((".xlsx", ".xlsm")):
        return pd.read_excel(path, engine="openpyxl", **kwargs)
    return pd.read_excel(path, **kwargs)

def main(df_factura_historico=None, df_conceptos_historico=None):
    excel_path = os.environ.get("EXCEL_PATH", "Resumen FRAs 2025 aBalados Services_macro.xlsm")
    if not os.path.exists(excel_path):
        log(f"🔥 ERROR: No se encuentra el Excel: {excel_path}")
        return
    try:
        base_dir = os.path.dirname(excel_path)

        (
            df_factura, df_conceptos, df_forma_pago, df_conceptos_texto,
            df_factura_historico_leido, df_conceptos_historico_leido
        ) = macro_adapter.adapt_from_macro(excel_path)

        # Priorizar los dataframes pasados como argumentos (desde el worker)
        df_factura_historico = df_factura_historico if df_factura_historico is not None else df_factura_historico_leido
        df_conceptos_historico = df_conceptos_historico if df_conceptos_historico is not None else df_conceptos_historico_leido
        # --- [FIN MODIFICADO] ---

        # Normalizar SOLO la clave 'empresa_emisora' en todos los DFs para que casen los filtros
        df_factura['empresa_emisora'] = df_factura['empresa_emisora'].apply(quitar_tildes_empresa)
        df_conceptos['empresa_emisora'] = df_conceptos['empresa_emisora'].apply(quitar_tildes_empresa)
        df_forma_pago['empresa_emisora'] = df_forma_pago['empresa_emisora'].apply(quitar_tildes_empresa)
        if df_conceptos_texto is not None and not df_conceptos_texto.empty and 'empresa_emisora' in df_conceptos_texto.columns:
            df_conceptos_texto['empresa_emisora'] = df_conceptos_texto['empresa_emisora'].apply(quitar_tildes_empresa)
        df_factura['empresa_emisora'] = df_factura['empresa_emisora'].apply(quitar_tildes_empresa)
        for df in (df_factura, df_conceptos, df_forma_pago):
            df.columns = df.columns.str.strip()
        if df_conceptos_texto is not None and not df_conceptos_texto.empty:
            df_conceptos_texto.columns = df_conceptos_texto.columns.str.strip()
        df_factura["api_key"] = df_factura["api_key"].fillna("")

        summary_data = []
        for _, frow in df_factura.iterrows():
            num, empresa = frow["NumFactura"], frow["empresa_emisora"]
            ejercicio, cliente_doc, api_key = frow["ejercicio"], frow["cliente_numero_documento"], frow["api_key"]
            if not api_key:
                # --- CALCULAR IMPORTE PARA ERROR_SIN_API_KEY ---
                try:
                    df_tmp = df_conceptos[(df_conceptos["NumFactura"] == num) & (df_conceptos["empresa_emisora"] == empresa)].copy()
                    df_tmp["__base"] = df_tmp["unidades"].fillna(0) * df_tmp["base_unidad"].fillna(0)
                    df_tmp["__iva"] = df_tmp["__base"] * (df_tmp["porcentaje"].fillna(0) / 100.0)
                    def _ret_row(r):
                        try:
                            p = float(r.get("porcentaje_retenido", 0.0) or 0.0)
                            if p <= 0:
                                return 0.0
                            return float(r["__base"]) * (p/100.0)
                        except Exception:
                            return 0.0
                    df_tmp["__ret"] = df_tmp.apply(_ret_row, axis=1)
                    total_lineas = float((df_tmp["__base"] + df_tmp["__iva"] - df_tmp["__ret"]).sum())
                    suplidos = float(frow.get("total_suplidos", 0.0) or 0.0)
                    importe_total = round(total_lineas + suplidos, 2)
                except Exception:
                    importe_total = 0.0
                # --- FIN cálculo importe ---
                
                summary_data.append({"id": num, "empresa": empresa, "status": "ERROR_SIN_API_KEY", "details": "Falta API Key", "pdf_url": None, "cliente": frow.get("cliente_nombre",""), "importe": importe_total})
                continue

            df_f_single = pd.DataFrame([frow])
            df_c_single = df_conceptos[(df_conceptos["NumFactura"] == num) & (df_conceptos["empresa_emisora"] == empresa)]
            df_fp_single = df_forma_pago[(df_forma_pago["NumFactura"] == num) & (df_forma_pago["empresa_emisora"] == empresa)]
            df_txt_single = pd.DataFrame(columns=["NumFactura","empresa_emisora","descripcion","posicion"])
            if df_conceptos_texto is not None and not df_conceptos_texto.empty:
                df_txt_single = df_conceptos_texto[(df_conceptos_texto["NumFactura"] == num) & (df_conceptos_texto["empresa_emisora"] == empresa)].copy()

            # --- Calcular importe total para guardar en summary.json ---
            try:
                df_tmp = df_c_single.copy()
                df_tmp["__base"] = df_tmp["unidades"].fillna(0) * df_tmp["base_unidad"].fillna(0)
                df_tmp["__iva"] = df_tmp["__base"] * (df_tmp["porcentaje"].fillna(0) / 100.0)
                def _ret_row(r):
                    try:
                        p = float(r.get("porcentaje_retenido", 0.0) or 0.0)
                        if p <= 0:
                            return 0.0
                        return float(r["__base"]) * (p/100.0)
                    except Exception:
                        return 0.0
                df_tmp["__ret"] = df_tmp.apply(_ret_row, axis=1)
                total_lineas = float((df_tmp["__base"] + df_tmp["__iva"] - df_tmp["__ret"]).sum())
                suplidos = float(frow.get("total_suplidos", 0.0) or 0.0)
                importe_total = round(total_lineas + suplidos, 2)
            except Exception:
                importe_total = 0.0
            # --- FIN cálculo importe ---


            try:
                # --- [MODIFICADO] Combinar datos actuales e históricos para la búsqueda ---
                df_factura_busqueda = pd.concat([df_factura, df_factura_historico], ignore_index=True) if df_factura_historico is not None else df_factura
                df_conceptos_busqueda = pd.concat([df_conceptos, df_conceptos_historico], ignore_index=True) if df_conceptos_historico is not None else df_conceptos

                xml_bytes = create_xml_from_data(
                    df_f_single, df_c_single, df_fp_single, df_txt_single,
                    df_factura_all=df_factura_busqueda, df_conceptos_all=df_conceptos_busqueda
                )
                # --- [FIN MODIFICADO] ---

                # Guardar XML
                try:
                    empresa_safe = str(empresa).replace(" ", "_").replace(".", "")
                    # IDs sin ".0"
                    num_str = str(num)
                    try:
                        fnum = float(num_str)
                        if fnum.is_integer():
                            num_str = str(int(fnum))
                    except Exception:
                        pass
                    num_safe = num_str.replace("/", "_")
                    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    xml_filename_logs = os.path.join(LOG_DIR, f"{empresa_safe}_proforma_{num_safe}_{stamp}.xml")
                    xml_filename_resp = os.path.join(RESPONSE_DIR, f"xml_{num_safe}_{stamp}.xml")
                    with open(xml_filename_logs, "wb") as f:
                        f.write(xml_bytes)
                    with open(xml_filename_resp, "wb") as f:
                        f.write(xml_bytes)
                    log(f"💾 XML guardado en: {xml_filename_logs}")
                    log(f"💾 XML copiado en: {xml_filename_resp}")
                except Exception as io_err:
                    log(f"⚠️ No se pudo guardar el XML: {io_err}")

                # Validación XSD previa al envío
                try:
                    validate_xml_against_xsd(xml_bytes)
                except Exception as xsd_err:
                    summary_data.append({
                        "id": num,
                        "empresa": empresa,
                        "status": "ERROR_VALIDACION_XSD",
                        "details": str(xsd_err),
                        "pdf_url": None,
                        "cliente": frow.get("cliente_nombre",""),
                        "importe": importe_total
                    })
                    continue

                # [NUEVO] Procesamiento paralelo: preparar datos para envío
                invoice_data = {
                    "xml_bytes": xml_bytes,
                    "api_key": api_key,
                    "num": num,
                    "empresa": empresa,
                    "ejercicio": ejercicio,
                    "cliente_doc": cliente_doc,
                    "importe_total": importe_total,
                    "cliente": frow.get("cliente_nombre", ""),
                    "frow": frow
                }
                # [NUEVO] Verificar si hay conexión antes de enviar
                import requests
                has_connection = False
                try:
                    requests.get("https://www.facturantia.com", timeout=5)
                    has_connection = True
                except:
                    has_connection = False
                
                # Enviar con cola offline si no hay conexión
                use_offline = os.environ.get("USE_OFFLINE_QUEUE", "0") == "1"
                result = send_proforma(xml_bytes, api_key, num, empresa, ejercicio, cliente_doc, 
                                     use_offline_queue=(use_offline and not has_connection))
                details_json = result.get("details", {})
                procs = []
                if isinstance(details_json, dict):
                    procs = details_json.get("proformas_procesadas", [])
                if procs:
                    for pr in procs:
                        pdf_url = pr.get("pdf", result.get("pdf_url"))
                        # [NUEVO] Obtener enlace QR si está disponible
                        enlace_qr = None
                        if result.get("enlaces_qr") and pr.get("external_id"):
                            enlace_qr = result.get("enlaces_qr", {}).get(pr.get("external_id"))
                        elif result.get("enlace_qr"):
                            enlace_qr = result.get("enlace_qr")
                        
                        summary_data.append({
                            "id": pr.get("external_id"),
                            "empresa": empresa,
                            "status": pr.get("status","ERROR").upper(),
                            "details": pr.get("message",""),
                            "pdf_url": pdf_url,
                            "enlace_qr": enlace_qr,  # [NUEVO] Enlace QR tributario
                            "cliente": frow.get("cliente_nombre",""), 
                            "importe": importe_total  # <-- INCLUIR IMPORTE EN TODOS LOS CASOS
                        })
                else:
                    # [NUEVO] Obtener enlace QR si está disponible
                    enlace_qr = None
                    if result.get("enlaces_qr") and num:
                        enlace_qr = result.get("enlaces_qr", {}).get(num)
                    elif result.get("enlace_qr"):
                        enlace_qr = result.get("enlace_qr")
                    
                    summary_data.append({
                        "id": num,
                        "empresa": empresa,
                        "status": result.get("status","ERROR").upper(),
                        "details": result.get("details",""),
                        "pdf_url": result.get("pdf_url"),
                        "enlace_qr": enlace_qr,  # [NUEVO] Enlace QR tributario
                        "cliente": frow.get("cliente_nombre",""),
                        "importe": importe_total  # <-- INCLUIR IMPORTE EN TODOS LOS CASOS
                    })
            except Exception as e:
                summary_data.append({
                    "id": num,
                    "empresa": empresa,
                    "status": "ERROR_GENERACION_XML",
                    "details": str(e),
                    "pdf_url": None,
                    "cliente": frow.get("cliente_nombre",""),
                    "importe": importe_total  # <-- INCLUIR IMPORTE EN TODOS LOS CASOS
                })

    except Exception as e:
        log(f"❌ Error general: {e}")
        return

    os.makedirs(RESPONSE_DIR, exist_ok=True)
    summary_filename = os.path.join(RESPONSE_DIR, "summary.json")
    with open(summary_filename, "w", encoding="utf-8") as f:
        json.dump(summary_data, f, indent=4, ensure_ascii=False)
    log(f"📝 Resumen guardado en: {summary_filename}")

    # --- Post-proceso en Macro: marcar o borrar filas ---
    try:
        post_action = os.environ.get("POST_MACRO_ACTION", "MARK").upper()
        mark_rows_in_macro(excel_path, summary_data, estado_col=COL_ESTADO, keep_vba=True)
        if post_action == "DELETE_OK":
            delete_ok_rows_in_macro(excel_path, summary_data, keep_vba=True)
        log(f"🧾 Post-proceso Excel Macro completado ({post_action}).")
    except Exception as e:
        log(f"⚠️ No se pudo actualizar Macro: {e}")

if __name__ == "__main__":
    main()


def _remove_diacritics(s: str) -> str:
    nfkd = unicodedata.normalize("NFKD", str(s))
    return "".join(ch for ch in nfkd if not unicodedata.combining(ch))


def _strip_invisibles(s: str) -> str:
    if not isinstance(s, str):
        s = str(s or "")
    invisibles = ["\uFEFF","\u200B","\u200C","\u200D","\u2060","\u00AD"]
    for ch in invisibles:
        s = s.replace(ch, "")
    return "".join(ch for ch in s if unicodedata.category(ch) not in ("Cf","Cc"))


def _normalize_company_name(raw: str) -> str:
    s = unicodedata.normalize("NFKC", str(raw or ""))
    s = _strip_invisibles(s)
    s = s.replace("\u00A0", " ")
    s = re.sub(r"\s+", " ", s).strip()
    s = re.sub(r"\s+,", ",", s)
    s = re.sub(r",\s+", ", ", s)
    # Always strip diacritics unless KEEP_ACCENTS=1
    if os.getenv("KEEP_ACCENTS", "0") != "1":
        s = _remove_diacritics(s)
    return s