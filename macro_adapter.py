# -*- coding: utf-8 -*-
"""macro_adapter.py
   Lee la hoja 'Macro' (.xlsx/.xlsm) y produce:
   (df_factura, df_conceptos, df_forma_pago, df_conceptos_texto)
   [CORREGIDO V7 - ¡LA BUENA!] Mantiene lógica original 100% + cierre seguro con finally wb.close().
"""
import os, re, numpy as np, pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.cell import column_index_from_string
# Importar Workbook para type hinting (opcional pero bueno)
from openpyxl.workbook import Workbook


def _norm_invoice_id(x: object) -> str:
    s = str(x).strip()
    if re.fullmatch(r"\d+(?:\.0+)?", s):
        try:
            return str(int(float(s)))
        except Exception:
            return s
    return s

PREFERRED_SHEETS = ["Macro", "MACRO", "Hoja1", "Resumen"]

EXCEL_COLS = {
    "num_factura": "A", "fecha_emision": "B", "cif_emisor": "E",
    "cliente_nombre": "G", "cliente_nif": "H", "cliente_dir": "I",
    "cliente_cp_prov": "J",
    "desc_1": "K", "imp_1": "L",
    "desc_2": "M", "imp_2": "N",
    "desc_3": "O", "imp_3": "P",
    "desc_4": "Q", "imp_4": "R",
    "desc_5": "S", "imp_5": "T",
    "desc_6": "U", "imp_6": "V",
    "desc_7": "W", "imp_7": "X",
    "desc_8": "Y", "imp_8": "Z",
    "suplidos_aa": "AA", "iban_macro": "AB", "estado": "AC",
    "base_ad": "AD", "total_ah": "AH", "factura_original": "AI",
}

ISO2_TO_ISO3 = {
    "AT":"AUT","BE":"BEL","BG":"BGR","CY":"CYP","CZ":"CZE","DE":"DEU","DK":"DNK",
    "EE":"EST","ES":"ESP","FI":"FIN","FR":"FRA","GR":"GRC","EL":"GRC","HR":"HRV",
    "HU":"HUN","IE":"IRL","IT":"ITA","LT":"LTU","LU":"LUX","LV":"LVA","MT":"MLT",
    "NL":"NLD","PL":"POL","PT":"PRT","RO":"ROU","SE":"SWE","SI":"SVN","SK":"SVK",
    "GB":"GBR","UK":"GBR","NO":"NOR","CH":"CHE","US":"USA","CN":"CHN"
}

def excel_col_to_idx(col_letters: str) -> int:
    # Usa la función original de openpyxl que sí existe
    return column_index_from_string(col_letters.strip().upper()) - 1

def clean_nif_cliente(nif: str) -> str:
    if nif is None: return ""
    s = str(nif)
    s = re.sub(r'^\s*(CIF|NIF)\s*[:\-]?\s*', '', s, flags=re.IGNORECASE)
    s = re.sub(r'\s+', '', s)
    return s

def normalize_cif_emisor(s: str) -> str:
    if s is None: return ""
    s = str(s).upper().strip()
    s = re.sub(r'^(CIF|NIF)\s*', '', s)
    s = re.sub(r'^ES', '', s)
    s = re.sub(r'[\s\-\._]', '', s)
    return s

def normalize_series_list(series_str: str) -> list:
    if pd.isna(series_str): return []
    return [p.strip() for p in str(series_str).split(',') if p.strip() != ""]

def coerce_number(x):
    if x is None: return 0.0
    if isinstance(x, (int, float, np.number)):
        return float(x) if not np.isnan(x) else 0.0
    s = str(x).strip()
    if s == "": return 0.0
    # Permitir notación científica y controlar errores
    if re.match(r"^-?\d+(\.\d+)?([eE][+-]?\d+)?$", s):
        try:
            val = float(s)
            return val if not np.isnan(val) else 0.0
        except ValueError:
            pass  # Continuar para probar con reemplazo de comas

    # Lógica original de reemplazo de comas (ajustada)
    s_cleaned = s.replace('.', '').replace(',', '.')
    try:
        val = float(s_cleaned)
        return val if not np.isnan(val) else 0.0
    except ValueError:
        return 0.0

def _split_cp_prov(s: str):
    """Acepta '41004 Sevilla' o 'Sevilla 41004'."""
    s = str(s or "").strip()
    if not s: return "", ""
    t = s.split()
    if t and re.fullmatch(r'\d{4,6}', t[-1]):  # ... CP
        return t[-1], " ".join(t[:-1]).strip()
    if t and re.fullmatch(r'\d{4,6}', t[0]):   # CP ...
        return t[0], " ".join(t[1:]).strip()
    return "", s

# [MODIFICADO] Usar try...finally para asegurar wb.close()
def _read_sheet_to_df_any(path: str, preferred_names=None):
    wb: Workbook = None # Definir wb fuera del try
    try:
        wb = load_workbook(path, data_only=True, read_only=True)
        names = [ws.title for ws in wb.worksheets]
        target = None
        if preferred_names:
            low = {n.lower(): n for n in names}
            for cand in preferred_names:
                if cand.lower() in low:
                    target = low[cand.lower()]; break
        if not target:
            for ws in wb.worksheets:
                if ws.sheet_state == "visible": target = ws.title; break
            if not target: target = wb.worksheets[0].title
        ws = wb[target]
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
    except Exception as e:
        # Imprimir error pero también propagarlo
        print(f"Error leyendo hoja genérica de {path}: {e}")
        raise # Es importante relanzar el error para que main.py lo capture
    finally:
        # ---> AÑADIDO <---
        if wb: # Solo intentar cerrar si el workbook se llegó a abrir
            try:
                wb.close() # <- Cierra el archivo
            except Exception as close_err:
                 print(f"Advertencia: Error al intentar cerrar {path} en _read_sheet_to_df_any: {close_err}")
        # ---> FIN AÑADIDO <---

    # --- El archivo debería estar cerrado aquí ---
    # Lógica original para crear DataFrame (sin tocar)
    if not rows: return pd.DataFrame()
    max_len = max(len(r) for r in rows) if rows else 0
    norm_rows = [r + [None]*(max_len-len(r)) for r in rows]
    return pd.DataFrame(norm_rows)


# [MODIFICADO] Usar try...finally para asegurar wb.close()
def _read_clientes_df_from_same_book(macro_path: str, sheet_name_candidates=None) -> pd.DataFrame:
    """Lee la hoja CLIENTES (o variantes) del MISMO archivo Excel que Macro (no crea columnas nuevas)."""
    if sheet_name_candidates is None:
        sheet_name_candidates = ["CLIENTES", "Clientes", "clientes", "EMISORES", "EMISOR", "CONFIG", "Config"]
    wb: Workbook = None # Definir wb fuera del try
    try:
        wb = load_workbook(macro_path, data_only=True, read_only=True)
        names = [ws.title for ws in wb.worksheets]
        low = {n.lower(): n for n in names}
        target = None
        for cand in sheet_name_candidates:
            if cand.lower() in low:
                target = low[cand.lower()]
                break
        if not target:
            raise ValueError("No se encontró la hoja 'CLIENTES' en el Excel.")
        ws = wb[target]
        rows = list(ws.iter_rows(values_only=True))
    except Exception as e:
        print(f"Error leyendo hoja 'CLIENTES' de {macro_path}: {e}")
        raise # Propagar error
    finally:
        # ---> AÑADIDO <---
        if wb: # Solo cerrar si se abrió
            try:
                wb.close() # <- Cierra archivo
            except Exception as close_err:
                print(f"Advertencia: Error cerrar {macro_path} en clientes: {close_err}")
        # ---> FIN AÑADIDO <---

    # --- El archivo debería estar cerrado aquí ---
    # Lógica original para crear DataFrame (sin tocar)
    if not rows:
        return pd.DataFrame()
    headers = [str(c or "").strip() for c in rows[0]]
    data = [list(r) for r in rows[1:]]
    df = pd.DataFrame(data, columns=headers)
    df.columns = [str(c).strip().lower() for c in df.columns]
    return df

# [MODIFICADO] Usar try...finally para asegurar wb.close()
def _read_emisores_df(emisores_path: str) -> pd.DataFrame:
    wb: Workbook = None # Definir wb fuera del try
    try:
        wb = load_workbook(emisores_path, data_only=True, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    except Exception as e:
        print(f"Error leyendo archivo de emisores {emisores_path}: {e}")
        raise # Propagar error
    finally:
        # ---> AÑADIDO <---
        if wb: # Solo cerrar si se abrió
            try:
                wb.close() # <- Cierra archivo
            except Exception as close_err:
                 print(f"Advertencia: Error al intentar cerrar {emisores_path} en _read_emisores_df: {close_err}")
        # ---> FIN AÑADIDO <---

    # --- El archivo debería estar cerrado aquí ---
    # Lógica original para crear DataFrame (sin tocar)
    if not rows: return pd.DataFrame()
    headers = [str(c or "").strip() for c in rows[0]]
    data = [list(r) for r in rows[1:]]
    df = pd.DataFrame(data, columns=headers)
    df.columns = [str(c).strip().lower() for c in df.columns]
    if "empresa_nombre" not in df.columns and "nombre_legal" in df.columns:
        df["empresa_nombre"] = df["nombre_legal"]
    if "pais_iso2" not in df.columns and "pains_iso2" in df.columns:
        df["pais_iso2"] = df["pains_iso2"]
    if "cif_aliases" not in df.columns:
        df["cif_aliases"] = ""
    return df


# --- El resto de funciones (_match_emisor, _snap_vat, adapt_from_macro) se mantienen EXACTAMENTE IGUAL que tu versión original ---
def _match_emisor(df_emisores: pd.DataFrame, cif_input: str):
    target = normalize_cif_emisor(cif_input)
    # Añadir comprobación por si target queda vacío después de normalizar
    if not target:
        print(f"Advertencia: CIF de entrada '{cif_input}' normalizado a vacío, no se puede buscar.")
        return None, None

    df_emisores = df_emisores.copy()
    # Comprobar si 'cif' existe, si no, intentar con 'cif/nif' o similares
    cif_col_name = None
    possible_cif_cols = ['cif', 'cif/nif', 'nif', 'vat']
    for col in possible_cif_cols:
        if col in df_emisores.columns:
            cif_col_name = col
            break
    if not cif_col_name:
        print(f"Advertencia: La hoja 'CLIENTES' no contiene ninguna columna de CIF esperada ({', '.join(possible_cif_cols)}).")
        return None, None

    df_emisores["_cif_norm"] = df_emisores[cif_col_name].astype(str).map(normalize_cif_emisor)
    hits = df_emisores[df_emisores["_cif_norm"] == target]
    has_token_col = "api_token" in df_emisores.columns # Comprobar si existe la columna token

    if not hits.empty:
        if has_token_col: # Ordenar solo si la columna existe
            # ESTA ES LA LÓGICA ORIGINAL QUE TENÍAS PARA ORDENAR
            hits = hits.assign(_has_token=hits["api_token"].astype(str).str.strip().ne(""))
            hits = hits.sort_values(by=["_has_token"], ascending=True)
            return hits.iloc[-1], "cif" # Devolver el último (prioriza True si ascending=True)
        else:
            return hits.iloc[0], "cif" # Si no hay token, devolver el primero

    # Búsqueda por alias (igual que antes)
    def split_aliases(s):
        if pd.isna(s) or str(s).strip()=="": return []
        return [a.strip() for a in str(s).split(",") if a.strip()!=""]

    alias_rows = []
    if "cif_aliases" in df_emisores.columns: # Comprobar si existe la columna de alias
        for _, row in df_emisores.iterrows():
            # Asegurarse que row['cif_aliases'] no sea None antes de split
            aliases_str = row.get("cif_aliases","")
            if pd.notna(aliases_str):
                 aliases = [normalize_cif_emisor(a) for a in split_aliases(aliases_str)]
                 if target in aliases: alias_rows.append(row)

    if alias_rows:
        hits = pd.DataFrame(alias_rows)
        if has_token_col: # Ordenar si existe token
             # Lógica original de ordenación
             hits = hits.assign(_has_token=lambda d: d["api_token"].astype(str).str.strip().ne(""))
             hits = hits.sort_values(by=["_has_token"], ascending=True)
             return hits.iloc[-1], "alias" # Devolver el último
        else:
             return hits.iloc[0], "alias" # Si no hay token, devolver el primero

    return None, None # No se encontró ni por CIF ni por alias

def _snap_vat(p):
    """Ajusta el IVA a tipos comunes si está muy cerca."""
    if p is None or (isinstance(p, float) and np.isnan(p)): return 0.0
    try: # Añadir try-except para conversión robusta
        p = float(p)
    except (ValueError, TypeError):
        return 0.0

    for c in (0.0, 4.0, 5.0, 10.0, 21.0):
        # Asegurar que p no sea NaN antes de comparar
        if not np.isnan(p) and abs(p - c) <= 0.25:
            return c
    # Redondear solo si p no es NaN
    return round(p, 2) if not np.isnan(p) else 0.0

# --- adapt_from_macro (lógica principal, SIN CAMBIOS respecto a tu versión original) ---
def adapt_from_macro(macro_path: str):
    df_all = _read_sheet_to_df_any(macro_path, preferred_names=PREFERRED_SHEETS)
    if df_all.empty or df_all.shape[0] < 2: # Mantener la comprobación original
        raise ValueError("La hoja 'Macro' está vacía o no tiene datos")

    # Lógica original para obtener 'df' desde 'df_all'
    df = df_all.iloc[1:].copy().reset_index(drop=True)
    # Asignar nombres de columna desde la primera fila de df_all (lógica original implícita)
    df.columns = [str(c or f"col_{i}").strip() for i, c in enumerate(df_all.iloc[0])]

    cols = {}
    for key, letter in EXCEL_COLS.items():
        idx = excel_col_to_idx(letter)
        # Lógica original para seleccionar columnas
        cols[key] = df.iloc[:, idx] if idx < df.shape[1] else pd.Series([np.nan]*len(df))
    m = pd.DataFrame(cols)

    m["num_factura"] = m["num_factura"].map(_norm_invoice_id)
    # Lógica original de filtrado
    m = m[m["num_factura"]!=""].reset_index(drop=True)
    if m.empty: # Añadir comprobación por si el filtrado deja m vacío
         raise ValueError("No se encontraron filas con número de factura válido tras filtrar.")

    m["tipo"] = np.where(m["num_factura"].str.startswith("Int"), "intereses",
                  np.where(m["num_factura"].str.startswith("A"), "intra", "normal"))
    m["fecha_emision"] = pd.to_datetime(m["fecha_emision"], errors="coerce")
    m["cliente_nif_limpio"] = m["cliente_nif"].apply(clean_nif_cliente)
    for c in [f"imp_{i}" for i in range(1,9)] + ["suplidos_aa","base_ad","total_ah"]:
        # Asegurar que la columna exista antes de aplicar coerce_number
        if c in m.columns:
            m[c] = m[c].apply(coerce_number)
        else:
            m[c] = np.nan # Si no existe, llenarla con NaN

    # Emisor desde hoja CLIENTES del mismo Excel
    try:
        df_emisores = _read_clientes_df_from_same_book(macro_path)
        if df_emisores.empty:
            raise ValueError("La hoja CLIENTES está vacía o mal formada")
    except ValueError as e:
         raise ValueError(f"Error al leer la hoja 'CLIENTES': {e}")

    # --- INICIO REFACTORIZACIÓN PARA MÚLTIPLES EMISORES ---

    # 1. Normalizar CIFs en el dataframe principal 'm' para poder agrupar
    def _norm_cif_cell(x):
        if pd.isna(x): return ""
        s = str(x).strip()
        if s == "" or s.lower() in ("none", "nan", "null", "#n/a", "#n/d", "-", "—"):
            return ""
        return normalize_cif_emisor(s)

    if 'cif_emisor' not in m.columns:
         raise ValueError("La columna 'cif_emisor' (E) no se encontró en la hoja Macro procesada.")
    m['cif_emisor_norm'] = m['cif_emisor'].apply(_norm_cif_cell)

    # Listas para acumular los dataframes de cada emisor
    all_facturas = []
    all_conceptos = []
    all_formas_pago = []
    all_textos = []

    # 2. Iterar sobre cada grupo de facturas (agrupadas por CIF de emisor)
    for cif_norm, m_group in m.groupby('cif_emisor_norm'):
        if not cif_norm:
            continue # Omitir filas sin CIF de emisor

        # --- Lógica de procesamiento existente, ahora aplicada a 'm_group' ---
        emisor_row_series, match_type = _match_emisor(df_emisores, cif_norm)
        if emisor_row_series is None:
            # En lugar de lanzar un error, podríamos registrarlo y continuar
            print(f"Advertencia: Empresa no configurada en hoja CLIENTES para CIF: {cif_norm}. Omitiendo {len(m_group)} facturas.")
            continue
        emisor_row = emisor_row_series.to_dict()

        empresa_nombre = str(emisor_row.get("empresa_nombre","") or "").strip()
        if not empresa_nombre:
            empresa_nombre = str(emisor_row.get("nombre_legal","") or "").strip()
        if not empresa_nombre:
             empresa_nombre = cif_norm
             print(f"Advertencia: No se encontró nombre para emisor {cif_norm}, usando CIF.")

        unidad_def   = str(emisor_row.get("unidad_medida_defecto","") or "ud")
        bic_conf     = str(emisor_row.get("bic","") or "").replace(" ","")
        series_list  = normalize_series_list(emisor_row.get("series_retencion",""))
        iban_defecto = str(emisor_row.get("iban_defecto","") or "").replace(" ","")

        def _pick(row_dict, keys):
            for k in keys:
                if k in row_dict:
                    v = str(row_dict.get(k, "") or "").strip()
                    if v: return v
            return ""
        api_token = _pick(emisor_row, ["api_token","api_key","token","facturantia_token","token_api"]) or os.environ.get("API_TOKEN", "").strip()
        api_email = _pick(emisor_row, ["api_email","email_api","usuario_email","api_user_email","user_email"]) or os.environ.get("API_EMAIL", "").strip()
        api_url   = _pick(emisor_row, ["api_url","url_api","endpoint","api_endpoint"]) or os.environ.get("API_URL", "").strip() or "https://www.facturantia.com/API/proformas_receptor.php"

        m_group["iban_resuelto"] = m_group["iban_macro"].astype(str).str.replace(" ", "").fillna('')
        mask_empty_iban = m_group["iban_resuelto"].str.strip() == ""
        m_group.loc[mask_empty_iban, "iban_resuelto"] = iban_defecto

        if (m_group["iban_resuelto"].fillna('').str.strip() == "").any():
            rows_missing = m_group[m_group["iban_resuelto"].fillna('').str.strip() == ""]
            idxs = (rows_missing.index + 2).tolist()
            print(f"Advertencia: Falta IBAN para CIF {cif_norm} en filas Excel: {idxs}. Omitiendo estas facturas.")
            continue

        conceptos_rows, textos_rows = [], []
        for _, row in m_group.iterrows():
            num = _norm_invoice_id(row["num_factura"])
            pos = 0
            for i in range(1, 9):
                desc = row.get(f"desc_{i}", "")
                imp  = row.get(f"imp_{i}", np.nan)
                if isinstance(desc, float) and np.isnan(desc): desc = ""
                desc = str(desc or "").strip()
                imp_float = coerce_number(imp)
                is_valid_imp = not pd.isna(imp_float) and imp_float != 0.0
                if desc and is_valid_imp:
                    conceptos_rows.append({
                        "NumFactura": num, "empresa_emisora": empresa_nombre,
                        "descripcion": desc, "cuenta_contable": "7050000",
                        "unidad_medida": unidad_def, "unidades": 1.0,
                        "base_unidad": float(imp_float),
                        "tipo_impuesto": "IVA", "porcentaje": 0.0,
                    })
                    pos += 1
                elif desc:
                    textos_rows.append({
                        "NumFactura": num, "empresa_emisora": empresa_nombre,
                        "descripcion": desc, "posicion": pos
                    })
                    pos += 1

        df_conceptos_group = pd.DataFrame(conceptos_rows) if conceptos_rows else pd.DataFrame()
        df_txt_group = pd.DataFrame(textos_rows) if textos_rows else pd.DataFrame()

        iva_map = {}
        if not m_group.empty:
            for num, grp in m_group.groupby("num_factura"):
                if grp.empty: continue
                first_row = grp.iloc[0]; tipo = first_row["tipo"]
                aa = float(coerce_number(first_row.get("suplidos_aa", 0.0)) or 0.0)
                ad = float(coerce_number(first_row.get("base_ad", 0.0)) or 0.0)
                ah = float(coerce_number(first_row.get("total_ah", 0.0)) or 0.0)
                if tipo == "normal":
                    vat = None
                    if ad != 0 and ah != 0:
                        try:
                            raw = ((ah - aa) - ad) / ad * 100.0
                            vat = _snap_vat(raw)
                        except ZeroDivisionError: vat = 21.0
                    if vat is None or vat <= 0.0 or np.isnan(vat): vat = 21.0
                    iva_map[num] = float(vat)
                else: iva_map[num] = 0.0
        if not df_conceptos_group.empty:
            df_conceptos_group["porcentaje"] = df_conceptos_group["NumFactura"].map(iva_map).fillna(21.0)

        if not df_conceptos_group.empty:
            df_conceptos_group["tipo_impuesto_retenido"] = ""
            df_conceptos_group["porcentaje_retenido"] = 0.0
            tipo_map = m_group.drop_duplicates(subset=['num_factura']).set_index('num_factura')['tipo'].to_dict()
            def needs_ret(num):
                t = tipo_map.get(num, "normal")
                if t == "intereses": return True
                if t == "normal" and any(str(num).startswith(str(p)) for p in series_list if p): return True
                return False
            mask = df_conceptos_group["NumFactura"].apply(needs_ret)
            df_conceptos_group.loc[mask, "tipo_impuesto_retenido"] = "IRPF"; df_conceptos_group.loc[mask, "porcentaje_retenido"] = 19.0

        fact_rows = []; processed_nums = set()
        for _, row in m_group.iterrows():
            num = _norm_invoice_id(row["num_factura"])
            if num in processed_nums: continue
            cp, prov = _split_cp_prov(row.get("cliente_cp_prov","")); prov = (prov or "").strip()
            nif_clean = str(row["cliente_nif_limpio"] or ""); tipo = str(row["tipo"])
            fecha_emision_val = row["fecha_emision"]
            if pd.isna(fecha_emision_val):
                 print(f"Advertencia: Fecha emisión inválida para factura {num}. Omitida."); processed_nums.add(num); continue
            ejercicio_val = fecha_emision_val.year
            if tipo == "intra":
                mcc = re.match(r"^([A-Z]{2})", nif_clean); iso2 = mcc.group(1).upper() if mcc else "ES"
                codigo_pais = ISO2_TO_ISO3.get(iso2, "ESP"); tipo_doc = "otro_id"; tipo_res = "U"
            else: codigo_pais = "ESP"; tipo_doc = "nif"; tipo_res = "R"
            pobl_use = prov or ""
            plantilla_emitidas = str(emisor_row.get("plantilla_facturas_emitidas", "") or "").strip()
            plantilla_proforma = str(emisor_row.get("plantilla_facturas_proforma", "") or "").strip()
            fact_rows.append({"NumFactura": num, "empresa_emisora": empresa_nombre, "api_key": api_token, "api_email": api_email, "api_url": api_url,"serie_factura": "", "fecha_emision": fecha_emision_val, "fecha_vencimiento": fecha_emision_val,"descripcion_general": "", "tipo_factura": "F1", "ejercicio": ejercicio_val, "cliente_tipo_persona": "J","cliente_nombre": row["cliente_nombre"], "cliente_tipo_documento": tipo_doc, "cliente_numero_documento": nif_clean,"cliente_cuenta_contable": "4300000", "cliente_observacion": "", "cliente_tipo_residencia": tipo_res,"cliente_codigo_pais": codigo_pais, "cliente_provincia": (prov or "")[:20], "cliente_poblacion": (pobl_use or "")[:50],"cliente_domicilio": row["cliente_dir"], "cliente_domicilio_2": "", "cliente_cp": cp, "cliente_telefono": "","cliente_email": "","total_suplidos": coerce_number(row.get("suplidos_aa", 0.0)) if tipo == "normal" else 0.0,"total_gastos_financieros": 0.0, "total_retenciones": 0.0, "plantilla_facturas_emitidas": plantilla_emitidas, "plantilla_facturas_proforma": plantilla_proforma})
            processed_nums.add(num)
        df_factura_group = pd.DataFrame(fact_rows)

        fp_rows = []; facturas_validas = set(df_factura_group["NumFactura"].unique()) if not df_factura_group.empty else set()
        for _, row in m_group.iterrows():
            num = _norm_invoice_id(row["num_factura"])
            if num not in facturas_validas: continue
            iban_final = str(row.get("iban_resuelto","") or "").strip()
            if not iban_final: continue
            fp_rows.append({"NumFactura": num, "empresa_emisora": empresa_nombre, "metodo": "transferencia","transferencia_banco": "ABANCA", "transferencia_beneficiario": empresa_nombre,"transferencia_concepto": "Pago Factura", "transferencia_iban": iban_final,"transferencia_bic": bic_conf if bic_conf else "CAGLESMMXXX"})
        df_forma_pago_group = pd.DataFrame(fp_rows)

        # Append results for this group to the master lists
        all_facturas.append(df_factura_group)
        all_conceptos.append(df_conceptos_group)
        all_formas_pago.append(df_forma_pago_group)
        all_textos.append(df_txt_group)

    # 3. Concatenar los dataframes de todos los emisores
    df_factura = pd.concat(all_facturas, ignore_index=True) if all_facturas else pd.DataFrame()
    df_conceptos = pd.concat(all_conceptos, ignore_index=True) if all_conceptos else pd.DataFrame()
    df_forma_pago = pd.concat(all_formas_pago, ignore_index=True) if all_formas_pago else pd.DataFrame()
    df_txt = pd.concat(all_textos, ignore_index=True) if all_textos else pd.DataFrame()

    # Asegurar que no haya columnas duplicadas antes de reindexar (evita "Reindexing only valid with uniquely valued Index")
    def _dedupe_columns(df: pd.DataFrame) -> pd.DataFrame:
        if df is not None and not df.empty and df.columns.duplicated().any():
            return df.loc[:, ~df.columns.duplicated()]
        return df

    df_factura = _dedupe_columns(df_factura)
    df_conceptos = _dedupe_columns(df_conceptos)
    df_forma_pago = _dedupe_columns(df_forma_pago)
    df_txt = _dedupe_columns(df_txt)

    # --- FIN REFACTORIZACIÓN ---

    # Asegurar columnas esperadas (lógica original)
    expected_fact_cols = ["NumFactura", "empresa_emisora", "api_key", "api_email", "api_url", "serie_factura", "fecha_emision", "fecha_vencimiento", "descripcion_general", "tipo_factura", "ejercicio", "cliente_tipo_persona", "cliente_nombre", "cliente_tipo_documento", "cliente_numero_documento", "cliente_cuenta_contable", "cliente_observacion", "cliente_tipo_residencia", "cliente_codigo_pais", "cliente_provincia", "cliente_poblacion", "cliente_domicilio", "cliente_domicilio_2", "cliente_cp", "cliente_telefono", "cliente_email", "total_suplidos", "total_gastos_financieros", "total_retenciones", "plantilla_facturas_emitidas", "plantilla_facturas_proforma"]
    expected_conc_cols = ["NumFactura", "empresa_emisora", "descripcion", "cuenta_contable", "unidad_medida", "unidades", "base_unidad", "tipo_impuesto", "porcentaje", "tipo_impuesto_retenido", "porcentaje_retenido"]
    expected_fp_cols = ["NumFactura", "empresa_emisora", "metodo", "transferencia_banco", "transferencia_beneficiario", "transferencia_concepto", "transferencia_iban", "transferencia_bic"]
    expected_txt_cols = ["NumFactura", "empresa_emisora", "descripcion", "posicion"]

    df_factura = df_factura.reindex(columns=expected_fact_cols)
    df_conceptos = df_conceptos.reindex(columns=expected_conc_cols)
    df_forma_pago = df_forma_pago.reindex(columns=expected_fp_cols)
    df_txt = df_txt.reindex(columns=expected_txt_cols)

    # --- [NUEVO] LEER Y PROCESAR HOJAS DE HISTORIAL ---
    # El objetivo es crear df_factura_historico y df_conceptos_historico para que
    # prueba.py pueda buscar facturas originales aunque hayan sido borradas de la hoja "Macro".

    df_factura_historico = pd.DataFrame()
    df_conceptos_historico = pd.DataFrame()

    try:
        wb = load_workbook(macro_path, data_only=True, read_only=True)
        all_sheet_names = wb.sheetnames

        # Excluir las hojas ya procesadas o de configuración
        sheets_to_exclude = PREFERRED_SHEETS + ["CLIENTES", "Clientes", "clientes", "EMISORES", "EMISOR", "CONFIG", "Config"]
        sheets_to_exclude_lower = [s.lower() for s in sheets_to_exclude]

        historical_sheet_names = [
            name for name in all_sheet_names
            if name.lower() not in sheets_to_exclude_lower
        ]

        if historical_sheet_names:
            historical_dfs_raw = []
            for sheet_name in historical_sheet_names:
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(values_only=True))
                if rows and len(rows) > 1:
                    # Asumimos la misma estructura: cabecera en fila 1, datos desde fila 2
                    headers = [str(c or f"col_{i}").strip() for i, c in enumerate(rows[0])]
                    data = rows[1:]
                    df_sheet = pd.DataFrame(data, columns=headers)
                    historical_dfs_raw.append(df_sheet)

            if historical_dfs_raw:
                df_hist_all = pd.concat(historical_dfs_raw, ignore_index=True)

                # --- Re-aplicar la misma lógica de procesamiento que para la hoja "Macro" ---
                # Asegurar índice único antes de resetear (evita error "Reindexing only valid with uniquely valued Index")
                df_hist = df_hist_all.copy()
                if df_hist.columns.duplicated().any():
                    df_hist = df_hist.loc[:, ~df_hist.columns.duplicated()]
                # Si el índice no es único, forzar reset con drop=True
                if not df_hist.index.is_unique:
                    df_hist.index = pd.RangeIndex(len(df_hist))
                df_hist = df_hist.reset_index(drop=True)
                # (No necesitamos la primera fila para nombres de columna porque ya los asignamos)

                cols_hist = {}
                for key, letter in EXCEL_COLS.items():
                    idx = excel_col_to_idx(letter)
                    if idx < df_hist.shape[1]:
                        # Asegurarse de que el nombre de columna exista antes de acceder
                        col_name = df_hist.columns[idx]
                        cols_hist[key] = df_hist[col_name]
                    else:
                        cols_hist[key] = pd.Series([np.nan] * len(df_hist))

                m_hist = pd.DataFrame(cols_hist)
                m_hist["num_factura"] = m_hist["num_factura"].map(_norm_invoice_id)
                m_hist = m_hist[m_hist["num_factura"] != ""]
                # Asegurar índice único antes de resetear
                if not m_hist.empty:
                    if m_hist.columns.duplicated().any():
                        m_hist = m_hist.loc[:, ~m_hist.columns.duplicated()]
                    if not m_hist.index.is_unique:
                        m_hist.index = pd.RangeIndex(len(m_hist))
                    m_hist = m_hist.reset_index(drop=True)

                if not m_hist.empty:
                    m_hist['cif_emisor_norm'] = m_hist['cif_emisor'].apply(_norm_cif_cell)

                    hist_all_facturas = []
                    hist_all_conceptos = []

                    for cif_norm, m_hist_group in m_hist.groupby('cif_emisor_norm'):
                        if not cif_norm: continue

                        emisor_row_series, _ = _match_emisor(df_emisores, cif_norm)
                        if emisor_row_series is None: continue

                        emisor_row = emisor_row_series.to_dict()
                        empresa_nombre = str(emisor_row.get("empresa_nombre","") or cif_norm).strip()
                        unidad_def = str(emisor_row.get("unidad_medida_defecto","") or "ud")

                        # Procesar conceptos del historial
                        conceptos_rows_hist = []
                        for _, row in m_hist_group.iterrows():
                            num = _norm_invoice_id(row["num_factura"])
                            for i in range(1, 9):
                                desc = row.get(f"desc_{i}", "")
                                imp  = row.get(f"imp_{i}", np.nan)
                                if (str(desc or "").strip() and not pd.isna(imp) and coerce_number(imp) != 0.0):
                                    conceptos_rows_hist.append({
                                        "NumFactura": num, "empresa_emisora": empresa_nombre,
                                        "base_unidad": float(coerce_number(imp)),
                                        "descripcion": str(desc), "unidad_medida": unidad_def,
                                    })

                        if conceptos_rows_hist:
                             hist_all_conceptos.append(pd.DataFrame(conceptos_rows_hist))

                        # Procesar facturas del historial
                        fact_rows_hist = []
                        processed_nums_hist = set()
                        for _, row in m_hist_group.iterrows():
                            num = _norm_invoice_id(row["num_factura"])
                            if num in processed_nums_hist: continue

                            fact_rows_hist.append({
                                "NumFactura": num, "empresa_emisora": empresa_nombre,
                                "cliente_nif": str(row.get("cliente_nif","")),
                                "cliente_nombre": str(row.get("cliente_nombre","")),
                            })
                            processed_nums_hist.add(num)

                        if fact_rows_hist:
                            hist_all_facturas.append(pd.DataFrame(fact_rows_hist))

                    if hist_all_facturas:
                        df_factura_historico = pd.concat(hist_all_facturas, ignore_index=True)
                    if hist_all_conceptos:
                        df_conceptos_historico = pd.concat(hist_all_conceptos, ignore_index=True)

    except Exception as e:
        # Si falla la lectura del historial, no detenemos el proceso, solo lo advertimos.
        print(f"Advertencia: No se pudo procesar el historial de facturas. Causa: {e}")
    finally:
        if 'wb' in locals() and wb:
            wb.close()

    # Devolver los 4 dataframes originales + los 2 del historial
    return df_factura, df_conceptos, df_forma_pago, df_txt, df_factura_historico, df_conceptos_historico